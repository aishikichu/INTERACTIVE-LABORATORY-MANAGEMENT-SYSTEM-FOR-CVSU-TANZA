import os
import socket
import threading
import time
from datetime import datetime
import json
import serial
from PIL import Image, ImageTk
import ttkbootstrap as ttk
import tkinter as tk
from tkinter import messagebox, filedialog
import pandas as pd
from tkinter import messagebox, simpledialog
import openpyxl
import sqlite3
import sqlite3
import threading
def send_json_response(sock, payload: dict):
    sock.sendall(json.dumps(payload).encode("utf-8"))
import os

client_sockets = []
clients_by_id = {}
client_id_by_sock = {}
clients_lock = threading.Lock()
pc_timers = {}
_next_client_id = 1

conn = sqlite3.connect(os.path.join(os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "..")), 'teacher_management.db'))
c = conn.cursor()
c.execute("SELECT * FROM student_assignments")
rows = c.fetchall()
print("🧾 All assignments:")
for r in rows:
    print(r)
conn.close()

conn = sqlite3.connect(os.path.join(os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "..")), 'teacher_management.db'))
c = conn.cursor()
c.execute("PRAGMA table_info(students)")
for col in c.fetchall():
    print(col)
conn.close()

conn = sqlite3.connect(os.path.join(os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "..")), 'teacher_management.db'))
c = conn.cursor()
c.execute("PRAGMA table_info(students)")
print("📋 students table schema:")
for col in c.fetchall():
    print(col)
conn.close()
import sqlite3

conn = sqlite3.connect(os.path.join(os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "..")), 'teacher_management.db'))
c = conn.cursor()

c.execute("SELECT student_number, name, rfid_code FROM students")
rows = c.fetchall()

print("📋 All students in DB:")
for row in rows:
    print(row)

conn.close()

import sqlite3

conn = sqlite3.connect(os.path.join(os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "..")), 'teacher_management.db'))
c = conn.cursor()

rfid = "111000"
instructor = "Angel Duales"
subject = "DCIT 013"

# Check if the student exists
c.execute("SELECT student_number, name FROM students WHERE rfid_code = ?", (rfid,))

student = c.fetchone()
print("🧠 Student from RFID:", student)

if student:
    student_number = student[0]

    # Check if this student is assigned to the correct instructor and subject
    c.execute("""
              SELECT *
              FROM student_assignments
              WHERE student_number = ?
                AND instructor = ?
                AND subject = ?
              """, (student_number, instructor, subject))

    assignment = c.fetchall()
    print("📚 Matching assignment(s):", assignment)
else:
    print("❌ No student found with that RFID.")

conn.close()

conn = sqlite3.connect(os.path.join(os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "..")), 'teacher_management.db'))
c = conn.cursor()

# Add 'status' column if it doesn't exist
try:
    c.execute("ALTER TABLE instructors ADD COLUMN status TEXT DEFAULT 'Active'")
except sqlite3.OperationalError:
    print("✅ 'status' column already exists.")

# Add 'section' column if it doesn't exist
try:
    c.execute("ALTER TABLE instructors ADD COLUMN section TEXT DEFAULT 'Unassigned'")
except sqlite3.OperationalError:
    print("✅ 'section' column already exists.")

conn.commit()
conn.close()

print("✅ Instructors table updated with 'status' and 'section' columns.")

conn = sqlite3.connect(os.path.join(os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "..")), 'teacher_management.db'))
c = conn.cursor()

c.execute("SELECT * FROM quiz_scores")
quiz_scores = c.fetchall()

print("✅ Quiz Scores Table Data:")
for row in quiz_scores:
    print(row)

conn.close()

conn = sqlite3.connect(os.path.join(os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "..")), 'teacher_management.db'))
c = conn.cursor()

# Get existing columns in `quiz_scores`
c.execute("PRAGMA table_info(quiz_scores);")
existing_columns = [col[1] for col in c.fetchall()]  # Extract column names

# Add `quiz_number` if missing
if "quiz_number" not in existing_columns:
    c.execute("ALTER TABLE quiz_scores ADD COLUMN quiz_number INTEGER;")
    print("✅ quiz_number column added successfully!")
else:
    print("⚠️ quiz_number column already exists.")

# Add `date_taken` if missing
if "date_taken" not in existing_columns:
    c.execute("ALTER TABLE quiz_scores ADD COLUMN date_taken TEXT;")
    print("✅ date_taken column added successfully!")
else:
    print("⚠️ date_taken column already exists.")

conn.commit()
conn.close()

conn = sqlite3.connect(os.path.join(os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "..")), 'teacher_management.db'))
c = conn.cursor()

# Check students table structure
c.execute("PRAGMA table_info(students);")
print("Students Table:", c.fetchall())

# Check quizzes table structure
c.execute("PRAGMA table_info(quizzes);")
print("Quizzes Table:", c.fetchall())

conn.close()

conn = sqlite3.connect(os.path.join(os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "..")), 'teacher_management.db'))
c = conn.cursor()

# Check students
c.execute("SELECT * FROM students")
print("\n✅ Students Table:")
for row in c.fetchall():
    print(row)

# Check quizzes
c.execute("SELECT * FROM quizzes")
print("\n✅ Quizzes Table:")
for row in c.fetchall():
    print(row)

conn.close()

conn = sqlite3.connect(os.path.join(os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "..")), 'teacher_management.db'))
c = conn.cursor()

c.execute("SELECT * FROM quiz_scores")
quiz_scores = c.fetchall()

print("✅ Quiz Scores Table Data:")
for row in quiz_scores:
    print(row)

conn.close()


def fetch_quiz_scores(section, subject):
    try:
        conn = sqlite3.connect(os.path.join(os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "..")), 'school.db'))  # ✅ Must match server DB
        c = conn.cursor()
        c.execute("""
                  SELECT student_name, quiz_number, score, date_taken
                  FROM quiz_scores
                  WHERE section = ? AND subject = ?
                  """, (section, subject))
        rows = c.fetchall()
        conn.close()
        return rows
    except Exception as e:
        print(f"⚠️ Failed to fetch quiz scores: {e}")
        return []


def safe_filename(text):
    text = str(text).strip()
    return re.sub(r"[\\/:*?\"<>|]", "_", text)


def get_quiz_base_dir(teacher, section, subject):
    # quizzes/<teacher>/<section>/<subject>/
    return os.path.join(
        "../../quizzes",
        safe_filename(teacher),
        safe_filename(section),
        safe_filename(subject)
    )


def create_quiz_scores_table():
    conn = sqlite3.connect(os.path.join(os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "..")), 'school.db'))  # Make sure this matches your actual database file
    c = conn.cursor()

    # ✅ Create the `quiz_scores` table if it does not exist
    c.execute("""
              CREATE TABLE IF NOT EXISTS quiz_scores
              (
                  id
                  INTEGER
                  PRIMARY
                  KEY
                  AUTOINCREMENT,
                  student_name
                  TEXT
                  NOT
                  NULL,
                  quiz_number
                  INTEGER
                  NOT
                  NULL,
                  score
                  REAL
                  NOT
                  NULL,
                  section
                  TEXT
                  NOT
                  NULL,
                  subject
                  TEXT
                  NOT
                  NULL,
                  date_taken
                  TEXT
                  NOT
                  NULL
              )
              """)

    conn.commit()
    conn.close()
    print("✅ `quiz_scores` table created successfully!")


# Run this function once to create the table
create_quiz_scores_table()


def create_subjects_table():
    """Ensure the subjects table exists in the database."""
    conn = sqlite3.connect(os.path.join(os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "..")), 'teacher_management.db'))
    c = conn.cursor()
    c.execute('''CREATE TABLE IF NOT EXISTS subjects
                 (
                     id
                     INTEGER
                     PRIMARY
                     KEY
                     AUTOINCREMENT,
                     subject_name
                     TEXT
                     UNIQUE
                     NOT
                     NULL
                 )''')
    conn.commit()
    conn.close()


# Call this function when the program starts
create_subjects_table()


def debug_check_new_instructors():
    """Print all new instructors stored in the database to check usernames & hashed passwords."""
    conn = sqlite3.connect(os.path.join(os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "..")), 'teacher_management.db'))
    c = conn.cursor()
    c.execute(
        "SELECT id, name, username, password FROM instructors ORDER BY id DESC LIMIT 5")  # Show last 5 instructors
    instructors = c.fetchall()
    conn.close()

    print("✅ Checking Recently Added Instructors:")
    for inst in instructors:
        print(f"ID: {inst[0]}, Name: {inst[1]}, Username: {inst[2]}, Hashed Password: {inst[3]}")


# Run this once
debug_check_new_instructors()

import hashlib


def hash_password(password):
    """Hash the password using SHA-256."""
    return hashlib.sha256(password.encode()).hexdigest()


import sqlite3

conn = sqlite3.connect(os.path.join(os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "..")), 'teacher_management.db'))
c = conn.cursor()

# ✅ List all tables in the database
c.execute("SELECT name FROM sqlite_master WHERE type='table';")
tables = c.fetchall()
print("Tables in database:", tables)

conn.close()
import sqlite3

conn = sqlite3.connect(os.path.join(os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "..")), 'teacher_management.db'))
c = conn.cursor()

# ✅ Ensure the `instructor_sections` table exists
c.execute("""
          CREATE TABLE IF NOT EXISTS instructor_sections
          (
              id
              INTEGER
              PRIMARY
              KEY
              AUTOINCREMENT,
              instructor
              TEXT
              NOT
              NULL,
              section
              TEXT
              NOT
              NULL,
              subject
              TEXT
              NOT
              NULL
          )
          """)
conn.commit()
conn.close()

import base64


def encode_image_to_base64(path):
    try:
        with open(path, "rb") as image_file:
            return base64.b64encode(image_file.read()).decode('utf-8')
    except Exception:
        return None


def create_instructor_table():
    conn = sqlite3.connect(os.path.join(os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "..")), 'teacher_management.db'))
    c = conn.cursor()
    c.execute('''CREATE TABLE IF NOT EXISTS instructors
                 (
                     id
                     INTEGER
                     PRIMARY
                     KEY
                     AUTOINCREMENT,
                     name
                     TEXT,
                     instructor_id
                     TEXT
                     UNIQUE,
                     username
                     TEXT
                     UNIQUE,
                     password
                     TEXT,
                     section
                     TEXT
                 )''')
    conn.commit()
    conn.close()


def add_instructor(name, instructor_id, username, password, section):
    conn = sqlite3.connect(os.path.join(os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "..")), 'teacher_management.db'))
    c = conn.cursor()
    try:
        hashed_password = hash_password(password)  # ✅ Hash password before storing
        c.execute("INSERT INTO instructors (name, instructor_id, username, password, section) VALUES (?, ?, ?, ?, ?)",
                  (name, instructor_id, username, hashed_password, section))
        conn.commit()
        messagebox.showinfo("Success", "Instructor added successfully!")
    except sqlite3.IntegrityError:
        messagebox.showerror("Error", "Instructor ID or Username already exists!")
    conn.close()


def remove_instructor(instructor_id):
    conn = sqlite3.connect(os.path.join(os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "..")), 'teacher_management.db'))
    c = conn.cursor()
    c.execute("DELETE FROM instructors WHERE instructor_id = ?", (instructor_id,))
    conn.commit()
    conn.close()
    messagebox.showinfo("Success", "Instructor removed successfully!")


def get_instructors():
    conn = sqlite3.connect(os.path.join(os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "..")), 'teacher_management.db'))
    c = conn.cursor()
    c.execute("SELECT id, name, username, instructor_id, section FROM instructors")
    instructors = c.fetchall()
    conn.close()
    return instructors


def show_add_remove_instructor():
    action = simpledialog.askstring("Instructor Management", "Type 'add' to add or 'remove' to remove an instructor:")

    if action == 'add':
        name = simpledialog.askstring("Add Instructor", "Enter Instructor Name:")
        instructor_id = simpledialog.askstring("Add Instructor", "Enter Instructor ID:")
        section = simpledialog.askstring("Add Instructor", "Enter Assigned Section:")
        if name and instructor_id and section:
            add_instructor(name, instructor_id, section)

    elif action == 'remove':
        instructor_id = simpledialog.askstring("Remove Instructor", "Enter Instructor ID to Remove:")
        if instructor_id:
            remove_instructor(instructor_id)

    else:
        messagebox.showerror("Error", "Invalid action! Please type 'add' or 'remove'.")

create_instructor_table()


def create_tables():
    conn = sqlite3.connect(os.path.join(os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "..")), 'teacher_management.db'))
    c = conn.cursor()

    # Ensure teacher_subjects table exists
    c.execute('''CREATE TABLE IF NOT EXISTS teacher_subjects
                 (
                     id
                     INTEGER
                     PRIMARY
                     KEY
                     AUTOINCREMENT,
                     instructor
                     TEXT,
                     section
                     TEXT,
                     subject
                     TEXT
                 )''')

    # Check if default subjects exist
    c.execute("SELECT COUNT(*) FROM teacher_subjects WHERE subject IN ('DCIT 65', 'ITEC 111')")
    if c.fetchone()[0] == 0:
        # Insert preset subjects for testing
        c.execute("INSERT INTO teacher_subjects (instructor, section, subject) VALUES (?, ?, ?)",
                  ("Ms. Angel", "BSIT 4-1", "DCIT 65"))
        c.execute("INSERT INTO teacher_subjects (instructor, section, subject) VALUES (?, ?, ?)",
                  ("Ms. Angel", "BSIT 4-1", "ITEC 111"))
        conn.commit()

    conn.close()


def update_students_table():
    """Ensure students table includes an instructor column."""
    conn = sqlite3.connect(os.path.join(os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "..")), 'teacher_management.db'))
    c = conn.cursor()

    # Add instructor column if it doesn't exist
    c.execute("PRAGMA table_info(students)")
    columns = [row[1] for row in c.fetchall()]
    if "instructor" not in columns:
        c.execute("ALTER TABLE students ADD COLUMN instructor TEXT")
        conn.commit()

    conn.close()
update_students_table()

def add_student(name, student_number, rfid_code, section, subject, instructor):
    conn = sqlite3.connect(os.path.join(os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "..")), 'teacher_management.db'))
    c = conn.cursor()
    try:
        # Insert into students table (basic info only)
        c.execute("""
                  INSERT
                  OR IGNORE INTO students (name, student_number, rfid_code)
            VALUES (?, ?, ?)
                  """, (name, student_number, rfid_code))

        # Insert into assignments (if not already assigned)
        c.execute("""
                  SELECT 1
                  FROM student_assignments
                  WHERE student_number = ? AND section = ? AND subject = ? AND instructor = ?
                  """, (student_number, section, subject, instructor))
        if not c.fetchone():
            c.execute("""
                      INSERT INTO student_assignments (student_number, section, subject, instructor)
                      VALUES (?, ?, ?, ?)
                      """, (student_number, section, subject, instructor))
        conn.commit()
    except Exception as e:
        print("❌ Error in add_student:", e)
    finally:
        conn.close()
def remove_student_by_rfid(rfid_code):
    conn = sqlite3.connect(os.path.join(os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "..")), 'teacher_management.db'))
    c = conn.cursor()
    c.execute("DELETE FROM students WHERE rfid_code=?", (rfid_code,))
    conn.commit()
    conn.close()
def get_students(section, subject, teacher):
    conn = sqlite3.connect(os.path.join(os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "..")), 'teacher_management.db'))
    c = conn.cursor()

    c.execute("""
              SELECT s.student_id, s.name, s.student_number, s.rfid_code
              FROM students s
                       JOIN student_assignments sa ON s.student_number = sa.student_number
              WHERE sa.section = ?
                AND sa.subject = ?
                AND sa.instructor = ?
              """, (section, subject, teacher))

    students = c.fetchall()
    conn.close()
    return students

def log_rfid_scan(student_id):
    conn = sqlite3.connect(os.path.join(os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "..")), 'teacher_management.db'))
    c = conn.cursor()
    c.execute("INSERT INTO rfid_logs (student_id) VALUES (?)", (student_id,))
    conn.commit()
    conn.close()

class ToolTip:
    def __init__(self, widget, text):
        self.widget = widget
        self.text = text
        self.tip_window = None
        widget.bind("<Enter>", self.show_tip)
        widget.bind("<Leave>", self.hide_tip)

    def show_tip(self, event):
        x, y, _cx, _cy = self.widget.bbox("insert")
        x = x + self.widget.winfo_rootx() + 25
        y = y + self.widget.winfo_rooty() + 25
        self.tip_window = tw = tk.Toplevel(self.widget)
        tw.wm_overrideredirect(True)
        tw.wm_geometry(f"+{x}+{y}")
        label = tk.Label(tw, text=self.text, justify=tk.LEFT,
                         background="#ffffe0", relief=tk.SOLID, borderwidth=1,
                         font=("tahoma", "8", "normal"))
        label.pack(ipadx=1)

    def hide_tip(self, event):
        tw = self.tip_window
        self.tip_window = None
        if tw:
            tw.destroy()

def create_rfid_logs_table():
    conn = sqlite3.connect(os.path.join(os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "..")), 'rfid_logs.db'))
    c = conn.cursor()
    c.execute('''CREATE TABLE IF NOT EXISTS rfid_logs
                 (
                     id
                     INTEGER
                     PRIMARY
                     KEY
                     AUTOINCREMENT,
                     student_id
                     TEXT,
                     timestamp
                     DATETIME
                     DEFAULT
                     CURRENT_TIMESTAMP
                 )''')
    conn.commit()
    conn.close()

def log_rfid_scan(student_id):
    conn = sqlite3.connect(os.path.join(os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "..")), 'rfid_logs.db'))
    c = conn.cursor()
    c.execute("INSERT INTO rfid_logs (student_id) VALUES (?)", (student_id,))
    conn.commit()
    conn.close()

def read_rfid():
    try:
        ser = serial.Serial('COM3', 9600, timeout=1)  # Update 'COM3' to the correct port
        time.sleep(2)  # Wait for the serial connection to initialize
        while True:
            if ser.in_waiting > 0:
                rfid_data = ser.readline().decode('utf-8').strip()
                if rfid_data:
                    print(f"RFID data: {rfid_data}")
                    log_rfid_scan(rfid_data)
                    messagebox.showinfo("RFID Scan", f"RFID {rfid_data} logged successfully")
    except serial.SerialException as e:
        messagebox.showerror("Error", f"Serial exception: {e}")
    except Exception as e:
        messagebox.showerror("Error", f"Exception: {e}")


def start_rfid_scanner():
    rfid_thread = threading.Thread(target=read_rfid)
    rfid_thread.daemon = True
    rfid_thread.start()


window = ttk.Window(themename="flatly")  # Light theme for a cleaner look
window.title('Teacher Management System')
window.state('zoomed')
window.resizable(True, True)
window.minsize(900, 700)  # Prevent shrinking below 900x700

def toggle_fullscreen(event=None):
    is_fullscreen = window.attributes('-fullscreen')
    window.attributes('-fullscreen', not is_fullscreen)  # Toggle fullscreen
window.bind('<F11>', toggle_fullscreen)
window.grid_columnconfigure(0, weight=1)
window.grid_rowconfigure(0, weight=1)

def resize_image(event):
    new_width = max(event.width, 900)
    new_height = max(event.height, 700)
    image = original_bg_image.resize((new_width, new_height), Image.Resampling.LANCZOS)
    bg_image = ImageTk.PhotoImage(image)
    canvas.create_image(0, 0, anchor=tk.NW, image=bg_image)
    canvas.image = bg_image  # Keep a reference to prevent garbage collection

script_dir = os.path.dirname(os.path.abspath(__file__))
bg_path = os.path.join(script_dir, "bg.jpg")
if not os.path.exists(bg_path):
    print(f"⚠️ Warning: bg.jpg not found at {bg_path}")
    # Fallback or create a placeholder if needed, but for now just try open
original_bg_image = Image.open(bg_path)

# ✅ Create Canvas and Set Background Image
canvas = tk.Canvas(window, width=900, height=700)  # Start with min size
canvas.grid(row=0, column=0, rowspan=4, columnspan=3, sticky="nsew")
canvas.bind("<Configure>", resize_image)  # Bind resize event to function

# Add a lighter-colored frame in the middle (no alpha transparency)
square_frame = tk.Frame(window, width=700, height=500, bg='#F0F0F0', bd=3, relief='solid', highlightthickness=2)
square_frame.place(relx=0.5, rely=0.45, anchor='center')  # Centered frame

# Pink glow for the frame
square_frame.configure(highlightbackground='#006400', highlightcolor='#FF8C00')

# Create a custom style for the button with pink color
style = ttk.Style()

# Ensure it applies to all buttons
style.configure("TButton",
                background='#FF8C00',  # Orange background
                foreground='white',  # White text
                borderwidth=2,
                relief='flat',
                font=("Arial", 14, 'bold'))

# Add hover effects
style.map("TButton",
          background=[("active", "#3BB143")],  # Darker orange when hovered
          foreground=[("active", "white")])  # Keep white text


# Function to clear the grid and remove all widgets from the frame
def clear_grid():
    # Clear widgets inside the square_frame
    for widget in square_frame.winfo_children():
        widget.destroy()

    # Clear all widgets from the main window (including any frames or labels)
    for widget in window.winfo_children():
        if isinstance(widget, ttk.Frame) or isinstance(widget, ttk.Label):
            widget.place_forget()  # Use place_forget to hide any placed widgets
            widget.destroy()  # Ensure they are fully removed


def show_teacher_selection():
    clear_grid()

    title_label = ttk.Label(window, text="Welcome to Instructors Client of CvSU Tanza Laboratory",
                            font=("Helvetica", 24, "bold"), bootstyle="light")
    title_label.place(relx=0.5, rely=0.07, anchor='center')
    title_label.configure(background='white', foreground='#8A2BE2')

    subtitle_label = ttk.Label(window, text="Empowering Future Educators",
                               font=("Helvetica", 18), bootstyle="light")
    subtitle_label.place(relx=0.5, rely=0.12, anchor='center')
    subtitle_label.configure(background='#D1BAFF', foreground='white')

    global rfid_entry

    # ✅ RFID Entry Field
    rfid_label = ttk.Label(square_frame, text="Scan Your RFID:", font=("Helvetica", 14, "bold"))
    rfid_label.grid(row=0, column=0, padx=10, pady=10, sticky="w")

    rfid_entry = ttk.Entry(square_frame, font=("Helvetica", 14))

    rfid_entry.grid(row=0, column=1, padx=10, pady=10, sticky="w")
    rfid_entry.focus_set()
    rfid_entry.bind("<Return>", lambda event: validate_rfid_login(rfid_entry.get().strip()))

    # ✅ Custom Bootstrap-like Button Styles
    style = ttk.Style()

    style.configure("Primary.TButton",
                    background='#007BFF',  # Bootstrap Primary Blue
                    foreground='white',
                    font=("Arial", 14, "bold"),
                    borderwidth=3,
                    relief="raised",
                    padding=(10, 5))

    style.map("Primary.TButton",
              background=[("active", "#0056b3")],  # Darker blue on hover
              relief=[("pressed", "sunken")])  # Pressed effect

    style.configure("Secondary.TButton",
                    background='#6C757D',  # Bootstrap Secondary Gray
                    foreground='white',
                    font=("Arial", 14, "bold"),
                    borderwidth=3,
                    relief="raised",
                    padding=(10, 5))

    style.map("Secondary.TButton",
              background=[("active", "#5a6268")],  # Darker gray on hover
              relief=[("pressed", "sunken")])  # Pressed effect

    # ✅ RFID Login Button (Styled)
    login_button = ttk.Button(square_frame, text="Login", style="Primary.TButton",
                              width=20, command=lambda: validate_rfid_login(rfid_entry.get().strip()))
    login_button.grid(row=1, column=0, columnspan=2, pady=10)


# Global Variables to Track the Currently Logged-in Instructor
current_instructor = None
current_section = None
current_subject = None


def admin_login():
    """Admin login window with password or RFID authentication."""
    login_window = tk.Toplevel()
    login_window.title("Admin Login")
    login_window.geometry("400x400")

    ttk.Label(login_window, text="Admin Login", font=("Helvetica", 18, "bold")).pack(pady=15)

    # Username Field
    ttk.Label(login_window, text="Username:", font=("Helvetica", 12)).pack(pady=5)
    username_entry = ttk.Entry(login_window, font=("Helvetica", 12))
    username_entry.pack(pady=5)

    # Password Field
    ttk.Label(login_window, text="Password:", font=("Helvetica", 12)).pack(pady=5)
    password_entry = ttk.Entry(login_window, font=("Helvetica", 12), show="•")
    password_entry.pack(pady=5)

    # RFID Entry Field
    ttk.Label(login_window, text="Or Tap RFID:", font=("Helvetica", 12)).pack(pady=5)
    rfid_entry = ttk.Entry(login_window, font=("Helvetica", 12))
    rfid_entry.pack(pady=5)

    def validate_login():
        """Validates login using username & password."""
        username = username_entry.get().strip()
        password = password_entry.get().strip()

        if not (username and password):
            messagebox.showerror("Error", "Username and password are required!")
            return

        conn = sqlite3.connect(os.path.join(os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "..")), 'teacher_management.db'))
        c = conn.cursor()
        c.execute("SELECT * FROM admins WHERE username=? AND password=?", (username, password))
        admin = c.fetchone()
        conn.close()

        if admin:
            messagebox.showinfo("Success", "Login successful!")
            login_window.destroy()
            admin_management()  # ✅ Open admin panel after successxz`
        else:
            messagebox.showerror("Error", "Invalid credentials!")

    def validate_rfid():
        """Validates login using RFID."""
        rfid_code = rfid_entry.get().strip()

        if not rfid_code:
            messagebox.showerror("Error", "Please enter an RFID code!")
            return

        conn = sqlite3.connect(os.path.join(os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "..")), 'teacher_management.db'))
        c = conn.cursor()
        c.execute("SELECT * FROM admins WHERE rfid_code=?", (rfid_code,))
        admin = c.fetchone()
        conn.close()

        if admin:
            messagebox.showinfo("Success", "RFID Login successful!")
            login_window.destroy()
            admin_management()  # ✅ Open admin panel after success
        else:
            messagebox.showerror("Error", "RFID not recognized!")

    # Login Buttons
    ttk.Button(login_window, text="Login", bootstyle="success", command=validate_login).pack(pady=10)
    ttk.Button(login_window, text="Login with RFID", bootstyle="primary", command=validate_rfid).pack(pady=5)

    # Signup Option
    ttk.Button(login_window, text="Sign Up", bootstyle="info", command=admin_signup).pack(pady=5)

    # Cancel Button
    ttk.Button(login_window, text="Cancel", bootstyle="danger", command=login_window.destroy).pack(pady=10)


def admin_signup():
    """Admin Sign-Up Window."""
    signup_window = tk.Toplevel()
    signup_window.title("Admin Sign-Up")
    signup_window.geometry("400x400")

    ttk.Label(signup_window, text="Admin Sign-Up", font=("Helvetica", 18, "bold")).pack(pady=15)

    ttk.Label(signup_window, text="Username:", font=("Helvetica", 12)).pack(pady=5)
    username_entry = ttk.Entry(signup_window, font=("Helvetica", 12))
    username_entry.pack(pady=5)

    ttk.Label(signup_window, text="Password:", font=("Helvetica", 12)).pack(pady=5)
    password_entry = ttk.Entry(signup_window, font=("Helvetica", 12), show="•")
    password_entry.pack(pady=5)

    ttk.Label(signup_window, text="Confirm Password:", font=("Helvetica", 12)).pack(pady=5)
    confirm_password_entry = ttk.Entry(signup_window, font=("Helvetica", 12), show="•")
    confirm_password_entry.pack(pady=5)

    ttk.Label(signup_window, text="RFID Code:", font=("Helvetica", 12)).pack(pady=5)
    rfid_entry = ttk.Entry(signup_window, font=("Helvetica", 12))
    rfid_entry.pack(pady=5)

    def register_admin():
        """Registers a new admin in the database."""
        username = username_entry.get().strip()
        password = password_entry.get().strip()
        confirm_password = confirm_password_entry.get().strip()
        rfid_code = rfid_entry.get().strip()

        if not (username and password and confirm_password and rfid_code):
            messagebox.showerror("Error", "All fields are required!")
            return

        if password != confirm_password:
            messagebox.showerror("Error", "Passwords do not match!")
            return

        conn = sqlite3.connect(os.path.join(os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "..")), 'teacher_management.db'))
        c = conn.cursor()

        try:
            c.execute("""
                      CREATE TABLE IF NOT EXISTS admins
                      (
                          id
                          INTEGER
                          PRIMARY
                          KEY
                          AUTOINCREMENT,
                          username
                          TEXT
                          UNIQUE
                          NOT
                          NULL,
                          password
                          TEXT
                          NOT
                          NULL,
                          rfid_code
                          TEXT
                          UNIQUE
                          NOT
                          NULL
                      )
                      """)
            c.execute("INSERT INTO admins (username, password, rfid_code) VALUES (?, ?, ?)",
                      (username, password, rfid_code))
            conn.commit()
            messagebox.showinfo("Success", "Admin registered successfully!")
            signup_window.destroy()
        except sqlite3.IntegrityError:
            messagebox.showerror("Error", "Username or RFID already exists!")
        finally:
            conn.close()

    ttk.Button(signup_window, text="Register", bootstyle="success", command=register_admin).pack(pady=10)
    ttk.Button(signup_window, text="Cancel", bootstyle="danger", command=signup_window.destroy).pack(pady=10)


def validate_rfid_login(rfid_code):
    """Validate login using RFID for instructor OR admin."""
    global current_instructor, current_section, current_subject

    if not rfid_code:
        messagebox.showerror("Login Error", "Please scan your RFID!")
        return

    conn = sqlite3.connect(os.path.join(os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "..")), 'teacher_management.db'))
    c = conn.cursor()

    # 🏫 Check if RFID belongs to an instructor
    c.execute("SELECT name, status FROM instructors WHERE rfid_code = ?", (rfid_code,))
    instructor = c.fetchone()

    if instructor:
        name, status = instructor
        if status != "Active":
            conn.close()
            messagebox.showerror("Login Error", "Your account is currently INACTIVE. Please contact Admin.")
            return

        conn.close()
        current_instructor = name
        print(f"✅ Logged in as Instructor: {current_instructor}")
        show_sections(current_instructor)
        return

    # 🛡 Check if RFID belongs to an admin
    c.execute("SELECT * FROM admins WHERE rfid_code = ?", (rfid_code,))
    admin = c.fetchone()
    conn.close()

    if admin:
        print(f"✅ Logged in as Admin")
        admin_management()
        return

    messagebox.showerror("Login Error", "RFID not recognized!")


def update_instructor_buttons():
    """ Refreshes instructor buttons dynamically while keeping the Add/Remove buttons visible """
    # Remove only instructor buttons, keeping the Add/Remove buttons
    for widget in square_frame.winfo_children():
        if isinstance(widget, ttk.Button) and widget.cget("text") not in ["Add Instructor", "Remove Instructor"]:
            widget.destroy()

    instructors = get_instructors()  # Fetch from the database

    for i, instructor in enumerate(instructors):
        btn = ttk.Button(square_frame, text=instructor[1], style="Transparent.TButton",
                         width=20, padding=(10, 5), command=lambda name=instructor[1]: show_sections(name))
        btn.grid(row=i // 3, column=i % 3, padx=20, pady=10)


def add_instructor_prompt():
    """ Pop-up to add an instructor without assigning a section """
    name = simpledialog.askstring("Add Instructor", "Enter Instructor Name:")
    instructor_id = simpledialog.askstring("Add Instructor", "Enter Instructor ID:")

    if name and instructor_id:
        add_instructor(name, instructor_id, section="")  # Empty section since instructors can access all
        update_instructor_buttons()  # Refresh UI
    else:
        messagebox.showerror("Error", "Please fill in all fields!")


def remove_instructor_prompt():
    """ Pop-up to remove an instructor """
    instructors = get_instructors()
    instructor_names = [inst[1] for inst in instructors]

    selected_name = simpledialog.askstring("Remove Instructor", "Enter Instructor Name to Remove:")

    if selected_name:
        for inst in instructors:
            if inst[1] == selected_name:
                remove_instructor(inst[2])  # Uses instructor_id for deletion
                update_instructor_buttons()  # Refresh UI without clearing Add/Remove buttons
                return

        messagebox.showerror("Error", "Instructor not found!")


import hashlib


def instructor_signup():
    """Instructor registration popup (now using RFID instead of username/password)."""
    signup_window = tk.Toplevel()
    signup_window.title("Instructor RFID Registration")
    signup_window.geometry("400x300")

    ttk.Label(signup_window, text="Name:", font=("Helvetica", 12)).pack(pady=5)
    name_entry = ttk.Entry(signup_window, font=("Helvetica", 12))
    name_entry.pack(pady=5)

    ttk.Label(signup_window, text="Instructor ID:", font=("Helvetica", 12)).pack(pady=5)
    instructor_id_entry = ttk.Entry(signup_window, font=("Helvetica", 12))
    instructor_id_entry.pack(pady=5)

    ttk.Label(signup_window, text="Scan RFID:", font=("Helvetica", 12)).pack(pady=5)
    rfid_entry = ttk.Entry(signup_window, font=("Helvetica", 12))
    rfid_entry.pack(pady=5)
    rfid_entry.focus_set()  # Autofocus on RFID field

    def register_instructor():
        """Registers an instructor with RFID."""
        name = name_entry.get().strip()
        instructor_id = instructor_id_entry.get().strip()
        rfid_code = rfid_entry.get().strip()

        if not (name and instructor_id and rfid_code):
            messagebox.showerror("Error", "All fields are required!")
            return

        conn = sqlite3.connect(os.path.join(os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "..")), 'teacher_management.db'))
        c = conn.cursor()

        try:
            # ✅ Ensure RFID column exists
            c.execute("""
                      CREATE TABLE IF NOT EXISTS instructors
                      (
                          id
                          INTEGER
                          PRIMARY
                          KEY
                          AUTOINCREMENT,
                          name
                          TEXT
                          NOT
                          NULL,
                          instructor_id
                          TEXT
                          UNIQUE
                          NOT
                          NULL,
                          rfid_code
                          TEXT
                          UNIQUE
                          NOT
                          NULL
                      )
                      """)

            # ✅ Insert Instructor with RFID instead of Username/Password
            c.execute("INSERT INTO instructors (name, instructor_id, rfid_code) VALUES (?, ?, ?)",
                      (name, instructor_id, rfid_code))
            conn.commit()
            messagebox.showinfo("Success", f"Instructor {name} registered successfully!")

        except sqlite3.IntegrityError:
            messagebox.showerror("Error", "Instructor ID or RFID already exists!")
        finally:
            conn.close()

    ttk.Button(signup_window, text="Register", command=register_instructor).pack(pady=10)


def update_instructor_dropdown():
    """Refresh the instructor dropdown list after adding a new instructor."""
    global instructor_var, instructor_dropdown

    conn = sqlite3.connect(os.path.join(os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "..")), 'teacher_management.db'))
    c = conn.cursor()
    c.execute("SELECT name FROM instructors")  # ✅ Fetch the latest instructor names
    instructors = c.fetchall()
    conn.close()

    instructor_names = [inst[0] for inst in instructors]  # Extract names

    instructor_dropdown["values"] = instructor_names  # ✅ Update dropdown options
    if instructor_names:
        instructor_var.set(instructor_names[-1])  # ✅ Auto-select the newest instructor

def view_students():
    student_window = tk.Toplevel(window)
    student_window.title("Student Viewer")

    width = 1370
    height = 720

    screen_width = student_window.winfo_screenwidth()
    screen_height = student_window.winfo_screenheight()

    x = int((screen_width / 2) - (width / 2))
    y = int((screen_height / 2) - (height / 2))

    student_window.geometry(f"{width}x{height}+{x}+{y}")

    student_window.resizable(False, False)

    student_window.transient(window)
    student_window.focus_force()

    student_window.configure(bg="#F5F7FA")

    style = ttk.Style()

    style.configure(
        "Modern.TButton",
        font=("Helvetica", 11, "bold"),
        padding=10
    )

    title_label = tk.Label(
        student_window,
        text="📚 Student Management Viewer",
        font=("Helvetica", 24, "bold"),
        bg="#F5F7FA",
        fg="#4B0082"
    )

    title_label.pack(pady=20)

    # =========================
    # FILTERS FRAME
    # =========================
    filters_frame = ttk.Frame(student_window)
    filters_frame.configure(padding=15)
    filters_frame.pack(pady=5)

    name_filter = tk.StringVar()
    section_filter = tk.StringVar()
    subject_filter = tk.StringVar()
    instructor_filter = tk.StringVar()

    ttk.Label(filters_frame, text="Name/ID:").grid(row=0, column=0, padx=5)

    ttk.Entry(
        filters_frame,
        textvariable=name_filter
    ).grid(row=0, column=1, padx=5)

    ttk.Label(filters_frame, text="Section:").grid(row=0, column=2, padx=5)

    section_combo = ttk.Combobox(
        filters_frame,
        textvariable=section_filter,
        state="readonly"
    )

    section_combo.grid(row=0, column=3, padx=5)

    ttk.Label(filters_frame, text="Subject:").grid(row=0, column=4, padx=5)

    subject_combo = ttk.Combobox(
        filters_frame,
        textvariable=subject_filter,
        state="readonly"
    )

    subject_combo.grid(row=0, column=5, padx=5)

    ttk.Label(filters_frame, text="Instructor:").grid(row=0, column=6, padx=5)

    instructor_combo = ttk.Combobox(
        filters_frame,
        textvariable=instructor_filter,
        state="readonly"
    )

    instructor_combo.grid(row=0, column=7, padx=5)

    # =========================
    # TABLE FRAME
    # =========================
    table_frame = ttk.Frame(student_window)
    table_frame.pack(fill="both", expand=True, padx=10, pady=10)

    columns = (
        "Name",
        "Student Number",
        "RFID",
        "Section",
        "Subject",
        "Instructor",
        "Actions"
    )

    tree = ttk.Treeview(
        table_frame,
        columns=columns,
        show="headings"
    )

    style.configure(
        "Treeview",
        rowheight=35,
        font=("Arial", 11)
    )

    style.configure(
        "Treeview.Heading",
        font=("Helvetica", 12, "bold")
    )

    for col in columns[:-1]:
        tree.heading(col, text=col)
        tree.column(col, anchor="center", width=140)

    tree.column("Actions", width=150, anchor="center")

    # Scrollbar
    scrollbar = ttk.Scrollbar(
        table_frame,
        orient="vertical",
        command=tree.yview
    )

    tree.configure(yscrollcommand=scrollbar.set)

    tree.pack(side="left", fill="both", expand=True)
    scrollbar.pack(side="right", fill="y")

    # =========================
    # POPULATE DROPDOWNS
    # =========================
    def populate_dropdowns():

        conn = sqlite3.connect(
            os.path.join(
                os.path.abspath(
                    os.path.join(os.path.dirname(__file__), "..", "..")
                ),
                'teacher_management.db'
            )
        )

        c = conn.cursor()

        c.execute("SELECT DISTINCT section FROM student_assignments")

        section_combo["values"] = [
            ""
        ] + [row[0] for row in c.fetchall()]

        c.execute("SELECT DISTINCT subject FROM student_assignments")

        subject_combo["values"] = [
            ""
        ] + [row[0] for row in c.fetchall()]

        c.execute("SELECT DISTINCT instructor FROM student_assignments")

        instructor_combo["values"] = [
            ""
        ] + [
            row[0].strip()
            for row in c.fetchall()
            if row[0]
        ]

        conn.close()

    # =========================
    # REFRESH TABLE
    # =========================
    def refresh_table():

        for row in tree.get_children():
            tree.delete(row)

        query = """
            SELECT s.name,
                   s.student_number,
                   s.rfid_code,
                   sa.section,
                   sa.subject,
                   sa.instructor
            FROM students s
            JOIN student_assignments sa
                ON s.student_number = sa.student_number
            WHERE 1 = 1
        """

        filters = []
        params = []

        if name_filter.get().strip():
            filters.append(
                "(s.name LIKE ? OR s.student_number LIKE ?)"
            )

            params.extend([
                f"%{name_filter.get().strip()}%"
            ] * 2)

        if section_filter.get().strip():
            filters.append("sa.section = ?")
            params.append(section_filter.get())

        if subject_filter.get().strip():
            filters.append("sa.subject = ?")
            params.append(subject_filter.get())

        if instructor_filter.get().strip():
            filters.append("LOWER(sa.instructor) = ?")

            params.append(
                instructor_filter.get().strip().lower()
            )

        if filters:
            query += " AND " + " AND ".join(filters)

        conn = sqlite3.connect(
            os.path.join(
                os.path.abspath(
                    os.path.join(os.path.dirname(__file__), "..", "..")
                ),
                'teacher_management.db'
            )
        )

        c = conn.cursor()

        c.execute(query, params)

        students = c.fetchall()

        for index, student in enumerate(students):

            tag = "evenrow" if index % 2 == 0 else "oddrow"

            tree.insert(
                "",
                "end",
                values=student + ("Double Click to Edit",),
                tags=(tag,)
            )

        tree.tag_configure(
            "evenrow",
            background="#F8F8F8"
        )

        tree.tag_configure(
            "oddrow",
            background="#EDEDED"
        )

        conn.close()

    # =========================
    # EDIT WINDOW
    # =========================
    def show_edit_window(data):

        name, number, rfid, section, subject, instructor = data[:6]

        edit_win = tk.Toplevel(student_window)
        edit_win.title("Edit Student")

        width = 620
        height = 520

        screen_width = edit_win.winfo_screenwidth()
        screen_height = edit_win.winfo_screenheight()

        x = int((screen_width / 2) - (width / 2))
        y = int((screen_height / 2) - (height / 2))

        edit_win.geometry(f"{width}x{height}+{x}+{y}")

        edit_win.configure(bg="#F5F7FA")
        edit_win.resizable(False, False)

        # IMPORTANT FIX
        edit_win.transient(student_window)
        edit_win.grab_set()

        main_frame = ttk.Frame(edit_win, padding=20)
        main_frame.pack(fill="both", expand=True)

        title_label = tk.Label(
            main_frame,
            text="✏ Edit Student",
            font=("Helvetica", 18, "bold"),
            bg="#F5F7FA",
            fg="#4B0082"
        )

        title_label.pack(pady=(0, 20))

        entries = {}

        fields = [
            ("Name", name),
            ("Student Number", number),
            ("RFID", rfid),
            ("Section", section),
            ("Subject", subject),
            ("Instructor", instructor)
        ]

        for label_text, value in fields:

            field_frame = ttk.Frame(main_frame)
            field_frame.pack(fill="x", pady=8)

            ttk.Label(
                field_frame,
                text=label_text + ":",
                width=15
            ).pack(side="left", padx=5)

            entry = ttk.Entry(field_frame, width=35)

            entry.insert(0, value)

            entry.pack(side="left", padx=5)

            entries[label_text] = entry

        # =========================
        # SAVE CHANGES
        # =========================
        def save_changes():

            new_values = {
                k: v.get().strip()
                for k, v in entries.items()
            }

            if not all(new_values.values()):

                messagebox.showerror(
                    "Error",
                    "All fields are required.",
                    parent=edit_win
                )

                return

            conn = sqlite3.connect(
                os.path.join(
                    os.path.abspath(
                        os.path.join(
                            os.path.dirname(__file__),
                            "..",
                            ".."
                        )
                    ),
                    'teacher_management.db'
                )
            )

            c = conn.cursor()

            try:

                c.execute("""
                    UPDATE students
                    SET name=?,
                        rfid_code=?
                    WHERE student_number=?
                """, (
                    new_values["Name"],
                    new_values["RFID"],
                    number
                ))

                c.execute("""
                    UPDATE student_assignments
                    SET section=?,
                        subject=?,
                        instructor=?
                    WHERE student_number=?
                """, (
                    new_values["Section"],
                    new_values["Subject"],
                    new_values["Instructor"],
                    number
                ))

                conn.commit()

                edit_win.destroy()

                refresh_table()

                messagebox.showinfo(
                    "Success",
                    "Student updated successfully!",
                    parent=student_window
                )

            finally:
                conn.close()

        # =========================
        # DELETE STUDENT
        # =========================
        def delete_student():

            confirm = messagebox.askyesno(
                "Confirm Delete",
                f"Delete student:\n\n{name}?",
                parent=edit_win
            )

            if not confirm:
                return

            conn = sqlite3.connect(
                os.path.join(
                    os.path.abspath(
                        os.path.join(
                            os.path.dirname(__file__),
                            "..",
                            ".."
                        )
                    ),
                    'teacher_management.db'
                )
            )

            c = conn.cursor()

            try:

                c.execute(
                    "DELETE FROM students WHERE student_number=?",
                    (number,)
                )

                c.execute(
                    "DELETE FROM student_assignments WHERE student_number=?",
                    (number,)
                )

                conn.commit()

                edit_win.destroy()

                refresh_table()

                messagebox.showinfo(
                    "Deleted",
                    "Student removed successfully!",
                    parent=student_window
                )

            finally:
                conn.close()

        # =========================
        # BUTTONS
        # =========================
        button_frame = ttk.Frame(main_frame)
        button_frame.pack(pady=20)

        ttk.Button(
            button_frame,
            text="💾 Save Changes",
            command=save_changes
        ).grid(row=0, column=0, padx=10)

        ttk.Button(
            button_frame,
            text="🗑 Remove",
            command=delete_student
        ).grid(row=0, column=1, padx=10)

        ttk.Button(
            button_frame,
            text="❌ Cancel",
            command=edit_win.destroy
        ).grid(row=0, column=2, padx=10)

    # =========================
    # DOUBLE CLICK EVENT
    # =========================
    def on_tree_select(event):

        selected = tree.focus()

        if not selected:
            return

        values = tree.item(selected, "values")

        show_edit_window(values)

    # Bind double click
    tree.bind("<Double-1>", on_tree_select)

    # =========================
    # FILTER EVENTS
    # =========================
    for var in (
        name_filter,
        section_filter,
        subject_filter,
        instructor_filter
    ):
        var.trace(
            "w",
            lambda *args: refresh_table()
        )

    # =========================
    # BUTTON FRAME
    # =========================
    button_frame = ttk.Frame(student_window)
    button_frame.pack(pady=10)

    ttk.Button(
        button_frame,
        text="🔄 Refresh Table",
        style="Modern.TButton",
        command=refresh_table
    ).grid(row=0, column=0, padx=10)

    ttk.Button(
        button_frame,
        text="❌ Close",
        style="Modern.TButton",
        command=student_window.destroy
    ).grid(row=0, column=1, padx=10)

    # =========================
    # INITIAL LOAD
    # =========================
    populate_dropdowns()
    refresh_table()

def assign_selected_instructors():
    if not selected_instructors:
        messagebox.showerror("Error", "Please select an instructor!")
        return
    selected_names = [selected_instructors[0]]  # Single selection only


import re  # Add this at the top if it's not already there

import re


def normalize_section(s):
    """Standardize spacing to make parsing reliable."""
    s = s.strip().upper()
    s = re.sub(r"\s+", " ", s)  # collapse multiple spaces
    s = re.sub(r"BSIT\s*(\d)\s*-\s*(\d)", r"BSIT \1-\2", s)  # fix inconsistent dashes/spaces
    return s


def parse_section(s):
    # This matches "BSIT 1-2", "BSIT1-2", etc., ignoring spaces
    match = re.search(r"BSIT\s*(\d)[^\d]?(\d)", s)
    if match:
        return int(match.group(1)), int(match.group(2))
    return (999, 999)


# Connect to DB
conn = sqlite3.connect(os.path.join(os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "..")), 'teacher_management.db'))
c = conn.cursor()
c.execute("SELECT DISTINCT section FROM teacher_subjects")
raw_sections = [normalize_section(row[0]) for row in c.fetchall() if row[0]]
conn.close()

# Remove duplicates after normalization
raw_sections = list(set(raw_sections))

# Sort properly
sections = sorted(raw_sections, key=parse_section)

print("Raw sections:")
for s in raw_sections:
    print(f"  '{s}' parsed as {parse_section(s)}")

print("\nSorted sections:")
for s in sections:
    print(f"  {s}")

print("Raw sections from DB:", raw_sections)
import sqlite3

conn = sqlite3.connect(os.path.join(os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "..")), 'teacher_management.db'))
c = conn.cursor()

import pandas as pd


def open_section_management_popup():
    popup = tk.Toplevel(window)
    popup.title("Manage or Create Sections")

    width = 1100
    height = 970

    # ✅ Get screen size
    screen_width = popup.winfo_screenwidth()
    screen_height = popup.winfo_screenheight()

    # ✅ Compute center position
    x = int((screen_width / 2) - (width / 2))
    y = int((screen_height / 2) - (height / 2))

    # ✅ Set final geometry WITH position
    popup.geometry(f"{width}x{height}+{x}+{y}")

    # ✅ Prevent resizing
    popup.resizable(False, False)

    # ✅ Keep popup on top
    popup.transient(window)

    # ✅ Disable main window interaction
    popup.grab_set()

    # ✅ Focus popup
    popup.focus_force()

    ttk.Label(popup, text="Existing Section-Subject Assignments", font=("Helvetica", 14, "bold")).pack(pady=10)

    frame = ttk.Frame(popup)
    frame.pack(fill="both", expand=True, padx=10, pady=5)

    tree = ttk.Treeview(
        frame,
        columns=("section", "subject"),
        show="headings",
        height=18,
        selectmode="extended"
    )
    tree.heading("section", text="Section")
    tree.heading("subject", text="Subject")
    tree.pack(side="left", fill="both", expand=True)

    scrollbar = ttk.Scrollbar(frame, orient="vertical", command=tree.yview)
    scrollbar.pack(side="right", fill="y")

    tree.configure(yscrollcommand=scrollbar.set)

    def refresh_tree():
        ttk.Label(
            popup,
            text="💡 Tip: Hold CTRL to select multiple items individually.\n"
                 "Hold SHIFT to select a range of rows.\n"
                 "You can also drag your mouse to highlight multiple assignments.",
            foreground="gray",
            font=("Arial", 10, "italic"),
            justify="center"
        ).pack(pady=(0, 10))
        for i in tree.get_children():
            tree.delete(i)
        conn = sqlite3.connect(os.path.join(os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "..")), 'teacher_management.db'))
        c = conn.cursor()
        c.execute("SELECT DISTINCT section, subject FROM teacher_subjects")
        rows = c.fetchall()
        conn.close()
        for section, subject in rows:
            tree.insert("", "end", values=(section, subject))

    refresh_tree()

    ttk.Separator(popup, orient="horizontal").pack(fill="x", pady=10)

    # Entry form frame
    entry_frame = ttk.Frame(popup)
    entry_frame.pack(pady=10)

    ttk.Label(entry_frame, text="New Section:").grid(row=0, column=0, padx=5, pady=5)
    all_sections = [f"BSIT {year}-{sec}" for year in range(1, 5) for sec in range(1, 5)]
    new_section_var = tk.StringVar()
    section_dropdown = ttk.Combobox(entry_frame, textvariable=new_section_var, values=all_sections, state="readonly",
                                    width=18)
    section_dropdown.grid(row=0, column=1, padx=5)

    ttk.Label(entry_frame, text="New Subject:").grid(row=0, column=2, padx=5, pady=5)
    new_subject_var = tk.StringVar()
    subject_entry = ttk.Entry(entry_frame, textvariable=new_subject_var, width=20)
    subject_entry.grid(row=0, column=3, padx=5)

    def add_section_subject():
        section = new_section_var.get().strip()
        subject = new_subject_var.get().strip()
        if not section or not subject:
            messagebox.showerror("Error", "Both section and subject are required.")
            return
        conn = sqlite3.connect(os.path.join(os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "..")), 'teacher_management.db'))
        c = conn.cursor()
        c.execute("INSERT INTO teacher_subjects (instructor, section, subject) VALUES (?, ?, ?)",
                  ("Admin", section, subject))
        conn.commit()
        conn.close()
        refresh_tree()
        new_section_var.set("")
        new_subject_var.set("")
        messagebox.showinfo("Success", "New section and subject added.")

    # 📥 Excel upload
    def upload_from_excel():
        file_path = filedialog.askopenfilename(
            filetypes=[("Excel Files", "*.xlsx *.xls")]
        )

        if not file_path:
            return

        try:
            df = pd.read_excel(file_path)

            # ✅ Validate columns
            if "Section" not in df.columns or "Subject" not in df.columns:
                messagebox.showerror(
                    "Error",
                    "Excel must contain 'Section' and 'Subject' columns."
                )
                return

            # ✅ Collect preview rows
            preview_rows = []

            for _, row in df.iterrows():
                section = str(row["Section"]).strip()
                subject = str(row["Subject"]).strip()

                if section and subject:
                    preview_rows.append(f"• {section} → {subject}")

            if not preview_rows:
                messagebox.showerror(
                    "Error",
                    "No valid Section/Subject rows found."
                )
                return

            # ✅ Preview message
            preview_text = "\n".join(preview_rows[:15])

            if len(preview_rows) > 15:
                preview_text += f"\n\n...and {len(preview_rows) - 15} more."

            confirm = messagebox.askyesno(
                "Confirm Upload",
                f"The following assignments will be added:\n\n"
                f"{preview_text}\n\n"
                f"Do you want to continue?"
            )

            if not confirm:
                return
            conn = sqlite3.connect(
                os.path.join(
                    os.path.abspath(
                        os.path.join(os.path.dirname(__file__), "..", "..")
                    ),
                    'teacher_management.db'
                )
            )

            c = conn.cursor()
            count = 0
            for _, row in df.iterrows():
                section = str(row["Section"]).strip()
                subject = str(row["Subject"]).strip()

                if section and subject:
                    c.execute("""
                        INSERT INTO teacher_subjects
                        (instructor, section, subject)
                        VALUES (?, ?, ?)
                    """, ("Admin", section, subject))

                    count += 1
            conn.commit()
            conn.close()
            refresh_tree()
            messagebox.showinfo(
                "Success",
                f"{count} assignments added successfully!"
            )
        except Exception as e:
            messagebox.showerror(
                "Error",
                f"Failed to read Excel file:\n{e}"
            )
    def download_excel_template():
        try:
            desktop_path = os.path.join(os.path.expanduser("~"), "Desktop")

            # Create folder
            folder_path = os.path.join(
                desktop_path,
                "Section Subject Templates"
            )

            os.makedirs(folder_path, exist_ok=True)

            # Template file path
            file_path = os.path.join(
                folder_path,
                "Section_Subject_Template.xlsx"
            )

            # Sample template
            template_df = pd.DataFrame({
                "Section": [
                    "BSIT 1-1",
                    "BSIT 1-2",
                    "BSIT 2-1"
                ],
                "Subject": [
                    "DCIT 013",
                    "COSC 77",
                    "ITEC 114"
                ]
            })

            template_df.to_excel(file_path, index=False)

            # Open folder automatically
            os.startfile(folder_path)

            messagebox.showinfo(
                "Template Downloaded",
                f"Excel template saved successfully!\n\nLocation:\n{file_path}"
            )

        except Exception as e:
            messagebox.showerror("Error", str(e))

    # 🔥 Remove Selected
    def remove_selected_assignments():
        selected_items = tree.selection()
        if not selected_items:
            messagebox.showwarning("Warning", "Please select at least one assignment to remove.")
            return
        selected_data = [tree.item(item, "values") for item in selected_items]
        confirm_msg = "Are you sure you want to delete the following assignments?\n\n"
        confirm_msg += "\n".join([f"• {sec} - {subj}" for sec, subj in selected_data])
        confirm = messagebox.askyesno("Confirm Deletion", confirm_msg)
        if confirm:
            conn = sqlite3.connect(os.path.join(os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "..")), 'teacher_management.db'))
            c = conn.cursor()
            for section, subject in selected_data:
                c.execute("DELETE FROM teacher_subjects WHERE section = ? AND subject = ?", (section, subject))
            conn.commit()
            conn.close()
            refresh_tree()
            messagebox.showinfo("Deleted", "Selected assignments have been removed.")

    # Buttons frame (single placement)
    button_frame = ttk.Frame(popup)
    button_frame.pack(pady=10)

    ttk.Button(button_frame, text="➕ Add Section + Subject", command=add_section_subject).grid(row=0, column=0, padx=5)
    ttk.Button(button_frame, text="🗑 Remove Selected", command=remove_selected_assignments).grid(row=0, column=1,
                                                                                                 padx=5)
    ttk.Button(
        button_frame,
        text="📥 Upload From Excel",
        command=upload_from_excel
    ).grid(row=0, column=2, padx=5)

    ttk.Button(
        button_frame,
        text="📤 Download Excel Template",
        command=download_excel_template
    ).grid(row=0, column=3, padx=5)


def manage_pc_clients():
    import tkinter as tk
    from tkinter import ttk, simpledialog, messagebox
    import sqlite3
    import os
    
    pc_window = tk.Toplevel()
    pc_window.title("Manage PC Client Numbers")
    pc_window.geometry("700x400")
    pc_window.transient()
    
    tree = ttk.Treeview(pc_window, columns=("MAC Address", "PC Number", "Hostname"), show='headings')
    tree.heading("MAC Address", text="MAC Address")
    tree.heading("PC Number", text="PC Number")
    tree.heading("Hostname", text="Hostname")
    tree.column("MAC Address", width=250)
    tree.column("PC Number", width=150)
    tree.column("Hostname", width=250)
    tree.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
    
    def refresh_pcs():
        for i in tree.get_children():
            tree.delete(i)
        conn = sqlite3.connect(os.path.join(os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "..")), 'teacher_management.db'))
        c = conn.cursor()
        c.execute("SELECT mac_address, pc_number, hostname FROM pc_assignments")
        for row in c.fetchall():
            tree.insert('', tk.END, values=row)
        conn.close()
        
    refresh_pcs()
    
    def edit_pc():
        sel = tree.selection()
        if not sel:
            messagebox.showwarning("Warning", "Select a PC to edit.")
            return
        item = tree.item(sel[0])
        mac, old_num, host = item['values']
        new_num = simpledialog.askstring("Edit PC Number", f"Enter new PC Number for {host} ({mac}):", initialvalue=old_num)
        if new_num:
            conn = sqlite3.connect(os.path.join(os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "..")), 'teacher_management.db'))
            c = conn.cursor()
            c.execute("UPDATE pc_assignments SET pc_number=? WHERE mac_address=?", (new_num, mac))
            conn.commit()
            conn.close()
            refresh_pcs()
            
    def delete_pc():
        sel = tree.selection()
        if not sel:
            return
        item = tree.item(sel[0])
        mac = item['values'][0]
        if messagebox.askyesno("Delete", f"Delete assignment for MAC {mac}?"):
            conn = sqlite3.connect(os.path.join(os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "..")), 'teacher_management.db'))
            c = conn.cursor()
            c.execute("DELETE FROM pc_assignments WHERE mac_address=?", (mac,))
            conn.commit()
            conn.close()
            refresh_pcs()

    btn_frame = ttk.Frame(pc_window)
    btn_frame.pack(pady=10)
    ttk.Button(btn_frame, text="Edit Selected PC Number", command=edit_pc).pack(side=tk.LEFT, padx=10)
    ttk.Button(btn_frame, text="Remove PC", command=delete_pc).pack(side=tk.LEFT, padx=10)


def admin_management():
    import tkinter as tk
    from tkinter import ttk, messagebox
    import sqlite3

    admin_window = tk.Toplevel(window)
    admin_window.title("Admin Management")

    width = 1200
    height = 800

    screen_width = admin_window.winfo_screenwidth()
    screen_height = admin_window.winfo_screenheight()

    x = int((screen_width / 2) - (width / 2))
    y = int((screen_height / 2) - (height / 2))

    admin_window.geometry(f"{width}x{height}+{x}+{y}")

    admin_window.resizable(False, False)

    # ✅ Keep linked to main app
    admin_window.transient(window)

    # ❌ REMOVE grab_set()

    # ✅ Focus window
    admin_window.focus_force()
    conn = sqlite3.connect(os.path.join(os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "..")), 'teacher_management.db'))
    c = conn.cursor()
    c.execute("""
              CREATE TABLE IF NOT EXISTS instructors
              (
                  id
                  INTEGER
                  PRIMARY
                  KEY
                  AUTOINCREMENT,
                  name
                  TEXT
                  NOT
                  NULL,
                  rfid_code
                  TEXT
                  UNIQUE
                  NOT
                  NULL,
                  instructor_id
                  TEXT
                  UNIQUE
                  NOT
                  NULL,
                  status
                  TEXT
                  DEFAULT
                  'Active',
                  section
                  TEXT
                  DEFAULT
                  'N/A'
              )
              """)
    conn.commit()
    conn.close()

    main_frame = ttk.Frame(admin_window, padding=10)
    main_frame.pack(fill="both", expand=True)

    ttk.Label(main_frame, text="Instructor Management", font=("Helvetica", 18, "bold")).pack(pady=10)

    header_frame = ttk.Frame(main_frame)
    header_frame.pack(fill="x", padx=20)

    headers = ["Name", "RFID Code", "Instructor ID", "Status", "Registered Section", ""]  # last column for Edit
    for i, title in enumerate(headers):
        header_frame.grid_columnconfigure(i, weight=1, uniform="column")
        ttk.Label(header_frame, text=title, font=("Helvetica", 12, "bold")).grid(
            row=0, column=i, padx=5, pady=5, sticky="nsew"
        )

    # =========================
    # SCROLLABLE CONTENT AREA
    # =========================

    table_container = ttk.Frame(main_frame)
    table_container.pack(fill="both", expand=True, padx=20)

    canvas = tk.Canvas(
        table_container,
        highlightthickness=0
    )

    scrollbar = ttk.Scrollbar(
        table_container,
        orient="vertical",
        command=canvas.yview
    )

    scrollable_frame = ttk.Frame(canvas)

    # Create window inside canvas
    canvas_window = canvas.create_window(
        (0, 0),
        window=scrollable_frame,
        anchor="nw"
    )

    # Configure scrolling
    canvas.configure(yscrollcommand=scrollbar.set)

    canvas.pack(side="left", fill="both", expand=True)
    scrollbar.pack(side="right", fill="y")

    # Update scrollregion automatically
    def configure_scrollregion(event):
        canvas.configure(
            scrollregion=canvas.bbox("all")
        )

    scrollable_frame.bind("<Configure>", configure_scrollregion)

    # Make frame width follow canvas width
    def resize_frame(event):
        canvas.itemconfig(
            canvas_window,
            width=event.width
        )

    canvas.bind("<Configure>", resize_frame)

    # Mousewheel scrolling
    def _on_mousewheel(event):

        # ONLY SCROLL if needed
        bbox = canvas.bbox("all")

        if bbox and bbox[3] > canvas.winfo_height():
            canvas.yview_scroll(
                int(-1 * (event.delta / 120)),
                "units"
            )

    canvas.bind_all("<MouseWheel>", _on_mousewheel)

    # THIS becomes your content frame now
    content_frame = scrollable_frame
    assign_selected_btn = ttk.Button(main_frame, text="🧩 Manage or Create Sections", width=30,
                                     command=open_section_management_popup)

    assign_selected_btn.pack(pady=10)

    for i in range(6):
        content_frame.grid_columnconfigure(i, weight=1, uniform="column")

    def load_instructor_rows():
        global selected_instructors
        selected_instructors = []

        for widget in content_frame.winfo_children():
            widget.destroy()

        conn = sqlite3.connect(os.path.join(os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "..")), 'teacher_management.db'))
        c = conn.cursor()
        c.execute("SELECT name, rfid_code, instructor_id, status FROM instructors")
        instructors = c.fetchall()
        c.execute("SELECT instructor, section, subject FROM teacher_subjects")
        assignments = c.fetchall()
        conn.close()

        assignment_map = {}
        for instructor, section, subject in assignments:
            assignment_map.setdefault(instructor, []).append(f"{section} {subject}")

        for idx, (name, rfid, ins_id, status) in enumerate(instructors, start=1):
            # Name
            ttk.Label(content_frame, text=name, anchor="center").grid(row=idx, column=0, padx=5, pady=2, sticky="ew")

            # RFID
            ttk.Label(content_frame, text=rfid, anchor="center").grid(row=idx, column=1, padx=5, pady=2, sticky="ew")

            # Instructor ID
            ttk.Label(content_frame, text=ins_id, anchor="center").grid(row=idx, column=2, padx=5, pady=2, sticky="ew")

            # Status Button
            ttk.Button(content_frame, text=status,
                       command=lambda iid=ins_id, current=status: toggle_status(iid, current)).grid(
                row=idx, column=3, padx=5, pady=2, sticky="ew")

            # Section Dropdown
            section_list = assignment_map.get(name, ["N/A"])
            section_var = tk.StringVar(value=section_list[0])
            section_dropdown = ttk.Combobox(content_frame, textvariable=section_var, values=section_list,
                                            state="readonly")
            section_dropdown.grid(row=idx, column=4, padx=5, pady=2, sticky="ew")

            # Action Buttons Frame
            action_frame = ttk.Frame(content_frame)
            action_frame.grid(row=idx, column=5, padx=5, pady=2, sticky="ew")

            # Edit Button
            ttk.Button(action_frame, text="✏ Edit",
                       command=lambda iid=ins_id: confirm_edit(iid)).pack(side="left", padx=2)

            # Remove Button
            ttk.Button(action_frame, text="🗑",
                       command=lambda iid=ins_id: confirm_remove(iid)).pack(side="left", padx=2)

    def confirm_edit(instructor_id):
        if messagebox.askyesno("Confirm Edit", "Edit this instructor's information?"):
            edit_instructor(instructor_id)

    def confirm_remove(instructor_id):
        if messagebox.askyesno("Confirm Removal", "Are you sure you want to remove this instructor?"):
            remove_instructor(instructor_id)
            load_instructor_rows()  # Refresh UI after deletion

    def toggle_status(instructor_id, current_status):
        new_status = "Inactive" if current_status == "Active" else "Active"
        conn = sqlite3.connect(os.path.join(os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "..")), 'teacher_management.db'))
        c = conn.cursor()
        c.execute("UPDATE instructors SET status=? WHERE instructor_id=?", (new_status, instructor_id))
        conn.commit()
        conn.close()
        load_instructor_rows()

    def edit_instructor(instructor_id):
        edit_window = tk.Toplevel()
        edit_window.title("Edit Instructor")

        conn = sqlite3.connect(os.path.join(os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "..")), 'teacher_management.db'))
        c = conn.cursor()
        c.execute("SELECT name, rfid_code, section FROM instructors WHERE instructor_id = ?", (instructor_id,))
        result = c.fetchone()
        conn.close()

        if result is None:
            messagebox.showerror("Error", "Instructor not found.")
            return

        name, rfid_code, current_section = result

        # Main Info
        ttk.Label(edit_window, text="Name:").grid(row=0, column=0, padx=10, pady=5, sticky="e")
        name_entry = ttk.Entry(edit_window)
        name_entry.insert(0, name)
        name_entry.grid(row=0, column=1, padx=10, pady=5)

        ttk.Label(edit_window, text="RFID Code:").grid(row=1, column=0, padx=10, pady=5, sticky="e")
        rfid_entry = ttk.Entry(edit_window)
        rfid_entry.insert(0, rfid_code)
        rfid_entry.grid(row=1, column=1, padx=10, pady=5)

        # Fetch all sections and subjects
        conn = sqlite3.connect(os.path.join(os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "..")), 'teacher_management.db'))
        c = conn.cursor()
        c.execute("SELECT DISTINCT section FROM teacher_subjects")
        raw_sections = [row[0] for row in c.fetchall() if row[0]]
        sections = sorted(raw_sections, key=parse_section)
        c.execute("SELECT DISTINCT subject FROM teacher_subjects")
        subjects = [row[0] for row in c.fetchall() if row[0]]
        conn.close()

        # Combined Section-Subject Listbox
        ttk.Label(edit_window, text="Assign Section-Subject:").grid(row=2, column=0, padx=10, pady=5, sticky="ne")
        combo_listbox = tk.Listbox(edit_window, selectmode="multiple", height=10, exportselection=False, width=35)

        conn = sqlite3.connect(os.path.join(os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "..")), 'teacher_management.db'))
        c = conn.cursor()
        c.execute(
            "SELECT DISTINCT section, subject FROM teacher_subjects WHERE section IS NOT NULL AND subject IS NOT NULL")
        section_subject_pairs = sorted(set((row[0], row[1]) for row in c.fetchall()),
                                       key=lambda x: (parse_section(x[0]), x[1]))
        conn.close()

        for sec, sub in section_subject_pairs:
            combo_listbox.insert(tk.END, f"{sec} - {sub}")

        combo_listbox.grid(row=2, column=1, padx=10, pady=5)

        def download_template():
            try:
                import os
                import subprocess
                from pathlib import Path

                instructor_name = name_entry.get().strip()

                if not instructor_name:
                    messagebox.showerror("Error", "Instructor name is empty.")
                    return

                # ✅ Desktop path
                desktop_path = Path.home() / "Desktop"

                # ✅ Main folder
                main_folder = desktop_path / "Instructor Files"

                # ✅ Subfolder
                sub_folder = main_folder / "Assign Section - Subject"

                # ✅ Create folders if not existing
                sub_folder.mkdir(parents=True, exist_ok=True)

                # ✅ File name
                file_name = f"{instructor_name}_Section_Subject_Template.xlsx"

                # ✅ Full file path
                file_path = sub_folder / file_name

                # ✅ Template content
                template_data = pd.DataFrame({
                    "Section": ["BSIT 1-1", "BSIT 1-2"],
                    "Subject": ["DCIT 013", "COSC 77"]
                })

                # ✅ Save excel
                template_data.to_excel(file_path, index=False)

                # ✅ Open folder automatically
                os.startfile(sub_folder)

                messagebox.showinfo(
                    "Success",
                    f"Template created successfully!\n\nSaved at:\n{file_path}"
                )

            except Exception as e:
                messagebox.showerror("Error", str(e))

        def upload_template():
            try:
                file_path = filedialog.askopenfilename(
                    filetypes=[("Excel Files", "*.xlsx")]
                )

                if not file_path:
                    return

                df = pd.read_excel(file_path)

                required_columns = ["Section", "Subject"]

                for col in required_columns:
                    if col not in df.columns:
                        messagebox.showerror(
                            "Error",
                            f"Missing column: {col}"
                        )
                        return

                preview_list = []

                for _, row in df.iterrows():
                    section = str(row["Section"]).strip()
                    subject = str(row["Subject"]).strip()

                    if not section or not subject:
                        continue

                    preview_list.append(f"{section} - {subject}")

                if not preview_list:
                    messagebox.showerror("Error", "No valid assignments found.")
                    return

                # ✅ Preview Window
                preview_window = tk.Toplevel()
                preview_window.title("Preview Uploaded Assignments")
                preview_window.geometry("500x500")
                preview_window.transient(edit_window)
                preview_window.grab_set()

                ttk.Label(
                    preview_window,
                    text="Assignments To Upload",
                    font=("Helvetica", 14, "bold")
                ).pack(pady=10)

                listbox = tk.Listbox(
                    preview_window,
                    font=("Arial", 11),
                    width=50,
                    height=18
                )
                listbox.pack(padx=10, pady=10, fill="both", expand=True)

                for item in preview_list:
                    listbox.insert(tk.END, item)

                ttk.Label(
                    preview_window,
                    text=f"Total Assignments: {len(preview_list)}",
                    foreground="green",
                    font=("Arial", 10, "bold")
                ).pack(pady=5)

                # ✅ Confirm Upload
                def confirm_upload():
                    conn = sqlite3.connect("../../teacher_management.db")
                    c = conn.cursor()

                    added_count = 0

                    for item in preview_list:
                        section, subject = item.split(" - ", 1)

                        c.execute("""
                            INSERT OR IGNORE INTO teacher_subjects
                            (instructor, section, subject)
                            VALUES (?, ?, ?)
                        """, (
                            name_entry.get().strip(),
                            section,
                            subject
                        ))

                        added_count += 1

                    conn.commit()
                    conn.close()

                    refresh_assignments(name_entry.get().strip())

                    preview_window.destroy()

                    messagebox.showinfo(
                        "Success",
                        f"{added_count} assignments uploaded successfully!"
                    )

                # ✅ Buttons
                btn_frame = ttk.Frame(preview_window)
                btn_frame.pack(pady=10)

                ttk.Button(
                    btn_frame,
                    text="✅ Confirm Upload",
                    command=confirm_upload
                ).grid(row=0, column=0, padx=10)

                ttk.Button(
                    btn_frame,
                    text="❌ Cancel",
                    command=preview_window.destroy
                ).grid(row=0, column=1, padx=10)

            except Exception as e:
                messagebox.showerror("Error", str(e))

        def save_changes():
            new_name = name_entry.get().strip()
            new_rfid = rfid_entry.get().strip()

            selected_combos = [combo_listbox.get(i) for i in combo_listbox.curselection()]

            if not new_name or not new_rfid:
                messagebox.showerror("Error", "Name and RFID Code cannot be empty.")
                return

            conn = sqlite3.connect(os.path.join(os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "..")), 'teacher_management.db'))
            c = conn.cursor()

            # Update instructor base info
            c.execute("""
                      UPDATE instructors
                      SET name      = ?,
                          rfid_code = ?
                      WHERE instructor_id = ?
                      """, (new_name, new_rfid, instructor_id))

            # Only update assignments if any are selected
            if selected_combos:
                # Assign selected combinations (FIXED — removed duplicate VALUES)
                for combo in selected_combos:
                    section, subject = combo.split(" - ", 1)
                    c.execute("""
                              INSERT
                              OR IGNORE INTO teacher_subjects (instructor, section, subject)
                              VALUES (?, ?, ?)
                              """, (new_name, section, subject))

                # If only one section (from selected combos), set it as primary
                unique_sections = set(combo.split(" - ", 1)[0] for combo in selected_combos)
                if len(unique_sections) == 1:
                    c.execute("""
                              UPDATE instructors
                              SET section = ?
                              WHERE instructor_id = ?
                              """, (list(unique_sections)[0], instructor_id))

            conn.commit()
            conn.close()

            messagebox.showinfo("Success", "Instructor updated and assignments saved.")
            edit_window.destroy()
            load_instructor_rows()

        template_frame = ttk.Frame(edit_window)
        template_frame.grid(row=3, column=0, columnspan=2, pady=10)

        ttk.Button(
            template_frame,
            text="📥 Download Template",
            command=download_template
        ).pack(side="left", padx=5)

        ttk.Button(
            template_frame,
            text="📤 Upload Template",
            command=upload_template
        ).pack(side="left", padx=5)
        ttk.Button(edit_window, text="Save", command=save_changes).grid(row=4, column=0, columnspan=2, pady=10)

        # Assignments display
        ttk.Label(edit_window, text="Assigned Sections & Subjects:").grid(row=5, column=0, columnspan=2, pady=(10, 0))
        assignment_frame = ttk.Frame(edit_window)
        assignment_frame.grid(row=6, column=0, columnspan=2, padx=10, pady=5)

        def refresh_assignments(current_instructor_name):
            for widget in assignment_frame.winfo_children():
                widget.destroy()

            conn = sqlite3.connect(os.path.join(os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "..")), 'teacher_management.db'))
            c = conn.cursor()
            c.execute("SELECT section, subject FROM teacher_subjects WHERE instructor = ?", (current_instructor_name,))
            assignments = c.fetchall()
            conn.close()

            if not assignments:
                ttk.Label(assignment_frame, text="No assignments yet.").grid(row=0, column=0, sticky="w")
                return

            # 💡 TIP (row 0)
            ttk.Label(
                assignment_frame,
                text="💡 Tip: Click multiple Section–Subject pairs to remove them simultaneously.",
                foreground="gray", font=("Arial", 10, "italic")
            ).grid(row=0, column=0, columnspan=2, sticky="w", padx=5, pady=(0, 5))

            # 📝 Selection label (row 1)
            ttk.Label(assignment_frame, text="Select Assignments to Remove:").grid(
                row=1, column=0, sticky="w", padx=5, pady=(0, 5))

            # 📜 Listbox (row 2)
            listbox = tk.Listbox(assignment_frame, selectmode="multiple", height=8, width=45, exportselection=False)
            listbox.grid(row=2, column=0, padx=5, pady=5)

            assignment_options = [f"{section} - {subject}" for section, subject in assignments]
            for item in assignment_options:
                listbox.insert(tk.END, item)

            # 🗑 Remove button (row 2, next to listbox)
            def remove_selected_assignments():
                selected_indices = listbox.curselection()
                if not selected_indices:
                    messagebox.showwarning("No Selection", "Please select one or more assignments to remove.")
                    return

                selected_items = [listbox.get(i) for i in selected_indices]
                confirm = messagebox.askyesno("Confirm", f"Remove the following?\n\n" + "\n".join(selected_items))
                if not confirm:
                    return

                with sqlite3.connect(os.path.join(os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "..")), 'teacher_management.db')) as conn:
                    c = conn.cursor()
                    for item in selected_items:
                        section, subject = item.split(" - ", 1)
                        c.execute("DELETE FROM teacher_subjects WHERE instructor = ? AND section = ? AND subject = ?",
                                  (current_instructor_name, section, subject))
                    conn.commit()

                refresh_assignments(current_instructor_name)
                messagebox.showinfo("Removed", "✅ Successfully removed:\n\n" + "\n".join(selected_items))

            ttk.Button(assignment_frame, text="🗑 Remove Selected", command=remove_selected_assignments).grid(
                row=2, column=1, padx=10, pady=5, sticky="n")

        def remove_assignment(instructor, section, subject):
            conn = sqlite3.connect(os.path.join(os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "..")), 'teacher_management.db'))
            c = conn.cursor()
            c.execute("DELETE FROM teacher_subjects WHERE instructor = ? AND section = ? AND subject = ?",
                      (instructor, section, subject))
            conn.commit()
            conn.close()
            refresh_assignments(name)

        refresh_assignments(name)

    def confirm_remove(instructor_id):
        if messagebox.askyesno("Confirm", "Are you sure you want to remove this instructor?"):
            conn = sqlite3.connect(os.path.join(os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "..")), 'teacher_management.db'))
            c = conn.cursor()
            c.execute("DELETE FROM instructors WHERE instructor_id=?", (instructor_id,))
            conn.commit()
            conn.close()
            load_instructor_rows()

    def register_instructor():
        name = name_entry.get().strip()
        rfid = rfid_entry.get().strip()
        ins_id = instructor_id_entry.get().strip()
        if not (name and rfid and ins_id):
            messagebox.showerror("Error", "All fields are required.")
            return
        conn = sqlite3.connect(os.path.join(os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "..")), 'teacher_management.db'))
        c = conn.cursor()
        try:
            c.execute("INSERT INTO instructors (name, rfid_code, instructor_id) VALUES (?, ?, ?)",
                      (name, rfid, ins_id))
            conn.commit()
            load_instructor_rows()
            messagebox.showinfo("Success", "Instructor added successfully.")
        except sqlite3.IntegrityError:
            messagebox.showerror("Error", "RFID or Instructor ID already exists.")
        finally:
            conn.close()

    form_frame = ttk.LabelFrame(main_frame, text="Add New Instructor")
    form_frame.pack(pady=20, fill="x", padx=20)

    ttk.Label(form_frame, text="Name:").grid(row=0, column=0, padx=5, pady=5)
    name_entry = ttk.Entry(form_frame)
    name_entry.grid(row=0, column=1, padx=5, pady=5)

    ttk.Label(form_frame, text="RFID Code:").grid(row=0, column=2, padx=5, pady=5)
    rfid_entry = ttk.Entry(form_frame)
    rfid_entry.grid(row=0, column=3, padx=5, pady=5)

    ttk.Label(form_frame, text="Instructor ID:").grid(row=0, column=4, padx=5, pady=5)
    instructor_id_entry = ttk.Entry(form_frame)
    instructor_id_entry.grid(row=0, column=5, padx=5, pady=5)

    ttk.Button(form_frame, text="Add Instructor", command=register_instructor).grid(row=0, column=6, padx=10)

    action_frame = ttk.Frame(main_frame)
    action_frame.pack(pady=10)
    ttk.Button(action_frame, text="📚 Manage Students", command=manage_students, width=30).pack(pady=3)
    ttk.Button(action_frame, text="🔍 View Students", command=view_students, width=30).pack(pady=3)
    ttk.Button(action_frame, text="👑 Admin Settings", command=admin_settings, width=30).pack(pady=3)
    ttk.Button(action_frame, text="💻 Manage PC Clients", command=manage_pc_clients, width=30).pack(pady=3)
    ttk.Button(action_frame, text="📤 Export All Scores by Instructor", width=30,
               command=export_scores_by_instructor_ui).pack(pady=3)
    ttk.Button(
        action_frame,
        text="🧹 Clean Instructor Data for New Semester",
        width=30,
        command=open_clean_data_popup
    ).pack(pady=3)

    load_instructor_rows()


def open_clean_data_popup():
    popup = tk.Toplevel(window)
    popup.title("Clean Instructor Data")
    width = 420
    height = 420
    screen_width = popup.winfo_screenwidth()
    screen_height = popup.winfo_screenheight()

    x = int((screen_width / 2) - (width / 2))
    y = int((screen_height / 2) - (height / 2))

    popup.geometry(f"{width}x{height}+{x}+{y}")
    popup.configure(bg="white")
    popup.resizable(False, False)
    popup.transient(window)
    popup.grab_set()
    popup.focus_force()
    popup.lift()
    popup.attributes("-topmost", True)

    popup.after(
        200,
        lambda: popup.attributes("-topmost", False)
    )
    title_label = tk.Label(
        popup,
        text="🧹 Clean Instructor Data",
        font=("Helvetica", 18, "bold"),
        bg="white",
        fg="#4B0082"
    )

    title_label.pack(pady=(25, 15))

    tk.Label(popup, text="Select Instructor to Clean:", bg="white").pack(pady=(10, 0))
    instructor_var = tk.StringVar()
    instructor_dropdown = ttk.Combobox(popup, textvariable=instructor_var, state="readonly", width=40)
    instructor_dropdown.pack(pady=5)

    # Load instructors
    try:
        conn = sqlite3.connect(os.path.join(os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "..")), 'teacher_management.db'))
        c = conn.cursor()
        c.execute("SELECT DISTINCT name FROM instructors WHERE name != 'Admin'")
        instructors = [row[0] for row in c.fetchall()]
        conn.close()
        instructor_dropdown["values"] = instructors
    except Exception as e:
        messagebox.showerror("Error", f"Failed to load instructors: {e}")
        return

    tk.Label(popup, text="Select Year to Clear:", bg="white").pack(pady=(10, 0))
    year_var = tk.StringVar()
    year_dropdown = ttk.Combobox(popup, textvariable=year_var, state="readonly", width=10)
    year_dropdown.pack(pady=5)

    def load_available_years(event):
        selected = instructor_var.get()
        if not selected:
            return
        try:
            conn1 = sqlite3.connect(os.path.join(os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "..")), 'school.db'))
            c1 = conn1.cursor()
            c1.execute("SELECT DISTINCT SUBSTR(date_taken, 1, 4) FROM activity_scores WHERE instructor = ?",
                       (selected,))
            years1 = {row[0] for row in c1.fetchall() if row[0]}
            c1.execute("SELECT DISTINCT SUBSTR(date, 1, 4) FROM attendance WHERE instructor = ?", (selected,))
            years2 = {row[0] for row in c1.fetchall() if row[0]}
            conn1.close()

            conn2 = sqlite3.connect(os.path.join(os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "..")), 'student_quiz_scores.db'))
            c2 = conn2.cursor()
            c2.execute("SELECT DISTINCT SUBSTR(date_taken, 1, 4) FROM quiz_scores")
            years3 = {row[0] for row in c2.fetchall() if row[0]}
            conn2.close()

            all_years = sorted(years1.union(years2).union(years3))
            year_dropdown["values"] = all_years
        except Exception as e:
            messagebox.showerror("Error", f"Failed to load years: {e}")

    instructor_dropdown.bind("<<ComboboxSelected>>", load_available_years)

    def confirm_clean():
        instructor = instructor_var.get()
        year = year_var.get()
        if not instructor or not year:
            messagebox.showwarning("Missing Info", "Please select both instructor and year.")
            return
        if messagebox.askyesno("Confirm", f"Clear all data for {instructor} in {year}?"):
            clear_instructor_student_data(instructor, year)
            popup.destroy()

    def confirm_remove_students_only():
        instructor = instructor_var.get()
        if not instructor:
            messagebox.showwarning("Missing Info", "Please select an instructor.")
            return

        try:
            conn = sqlite3.connect(os.path.join(os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "..")), 'teacher_management.db'))
            c = conn.cursor()
            c.execute("""
                      SELECT s.name, s.student_number, s.rfid_code
                      FROM students s
                               JOIN student_assignments a ON s.student_number = a.student_number
                      WHERE a.instructor = ?
                      """, (instructor,))
            students = c.fetchall()
            conn.close()

            if not students:
                messagebox.showinfo("No Students", f"No students assigned to {instructor}.")
                return

            # 🔹 Create confirmation preview window
            preview_win = tk.Toplevel(popup)
            preview_win.title("Confirm Student Removal")
            preview_win.geometry("500x400")
            preview_win.config(bg="white")

            tk.Label(preview_win, text=f"Students assigned to {instructor}:", font=("Helvetica", 12, "bold"),
                     bg="white").pack(pady=10)

            # 🔹 Scrollable preview list
            frame = tk.Frame(preview_win, bg="white")
            frame.pack(fill="both", expand=True, padx=10, pady=(0, 10))

            text = tk.Text(frame, wrap="none", font=("Consolas", 10), bg="#f9f9f9")
            text.pack(side="left", fill="both", expand=True)

            scrollbar = ttk.Scrollbar(frame, orient="vertical", command=text.yview)
            scrollbar.pack(side="right", fill="y")
            text.configure(yscrollcommand=scrollbar.set)

            text.insert(tk.END, f"{'Name':<25} {'Student #':<15} {'RFID'}\n")
            text.insert(tk.END, "-" * 60 + "\n")
            for name, number, rfid in students:
                text.insert(tk.END, f"{name:<25} {number:<15} {rfid}\n")

            text.config(state="disabled")

            # 🔹 Confirm + Cancel Buttons
            btn_frame = ttk.Frame(preview_win)
            btn_frame.pack(pady=10)

            ttk.Button(btn_frame, text="✅ Confirm Remove", style="success.TButton",
                       command=lambda: [remove_all_students_from_instructor(instructor), preview_win.destroy(),
                                        popup.destroy()]).pack(side="left", padx=10)
            ttk.Button(btn_frame, text="❌ Cancel", style="danger.TButton", command=preview_win.destroy).pack(
                side="left", padx=10)

        except Exception as e:
            messagebox.showerror("Error", f"Error loading assigned students: {e}")

    # Buttons
    ttk.Button(popup, text="🧹 Clean Scores & Attendance", command=confirm_clean).pack(pady=10)
    ttk.Button(popup, text="❌ Remove All Assigned Students", command=confirm_remove_students_only).pack(pady=5)


def clear_instructor_student_data(instructor_name, year_to_clear):
    try:
        instructor_name = instructor_name.strip()

        # 🔹 Connect to databases
        conn_school = sqlite3.connect(os.path.join(os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "..")), 'school.db'), timeout=5.0)
        print("🔍 school.db locked?", os.path.exists("school.db-journal"))
        conn_teacher = sqlite3.connect(os.path.join(os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "..")), 'teacher_management.db'), timeout=5.0)

        c_school = conn_school.cursor()
        c_teacher = conn_teacher.cursor()

        # 🔹 Delete activity scores
        c_school.execute("""
                         DELETE
                         FROM activity_scores
                         WHERE instructor = ?
                           AND date_taken LIKE ?
                         """, (instructor_name, f"{year_to_clear}-%"))

        # 🔹 Delete attendance
        c_school.execute("""
                         DELETE
                         FROM attendance
                         WHERE instructor = ? AND date LIKE ?
                         """, (instructor_name, f"{year_to_clear}-%"))

        # 🔹 Delete quiz scores from school.db
        c_teacher.execute("SELECT section, subject FROM teacher_subjects WHERE instructor = ?", (instructor_name,))
        assignments = c_teacher.fetchall()

        for section, subject in assignments:
            c_school.execute("""
                             DELETE
                             FROM quiz_scores
                             WHERE section = ? AND subject = ? AND date_taken LIKE ?
                             """, (section.strip(), subject.strip(), f"{year_to_clear}-%"))

        # 🔹 Delete student assignments
        c_teacher.execute("""
                          DELETE
                          FROM student_assignments
                          WHERE instructor = ?
                          """, (instructor_name,))

        # 🔹 Delete students with no assignments left
        c_teacher.execute("SELECT DISTINCT student_number FROM students")
        all_students = [row[0] for row in c_teacher.fetchall()]

        for student_number in all_students:
            c_teacher.execute("SELECT COUNT(*) FROM student_assignments WHERE student_number = ?", (student_number,))
            if c_teacher.fetchone()[0] == 0:
                c_teacher.execute("DELETE FROM students WHERE student_number = ?", (student_number,))

        # ✅ Commit and close
        conn_school.commit()
        conn_teacher.commit()
        conn_school.close()
        conn_teacher.close()

        messagebox.showinfo("Cleaned", f"✅ {instructor_name}'s data for {year_to_clear} has been cleared.")

    except Exception as e:
        messagebox.showerror("Error", f"Failed to clear data: {e}")


def grouped_rows_with_headers(df, group_keys, display_prefix):
    rows = []
    grouped = df.groupby(group_keys)
    for keys, group in grouped:
        header = f"{display_prefix} {keys[0]}: {keys[1]}" if isinstance(keys, tuple) else f"{display_prefix} {keys}"
        rows.append([header])  # Insert group header
        rows.append(group.columns.tolist())  # Insert column headers
        rows.extend(group.values.tolist())  # Insert group rows
        rows.append([])  # Blank line between groups
    return pd.DataFrame(rows)


def export_quiz_scores_by_instructor(instructor_name):
    try:
        # Step 1: Get instructor's section-subject assignments
        conn_assign = sqlite3.connect(os.path.join(os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "..")), 'teacher_management.db'))
        c_assign = conn_assign.cursor()
        c_assign.execute("""
                         SELECT section, subject
                         FROM teacher_subjects
                         WHERE instructor = ?
                         """, (instructor_name,))
        assignments = c_assign.fetchall()
        conn_assign.close()

        if not assignments:
            messagebox.showinfo("No Data", f"{instructor_name} has no assigned section/subject.")
            return

        import pandas as pd
        from datetime import datetime

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"{instructor_name}_Exported_Scores_{timestamp}.xlsx"
        writer = pd.ExcelWriter(filename, engine='openpyxl')

        conn_scores = sqlite3.connect(os.path.join(os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "..")), 'school.db'))

        c_scores = conn_scores.cursor()
        quiz_results = []

        for section, subject in assignments:
            c_scores.execute("""
                             SELECT student_name, quiz_number, score, date_taken
                             FROM quiz_scores
                             WHERE section = ? AND subject = ?
                             """, (section, subject))
            rows = c_scores.fetchall()
            for row in rows:
                quiz_results.append((section, subject, *row))

        conn_scores.close()

        if quiz_results:
            df_quiz = pd.DataFrame(quiz_results, columns=[
                "Section", "Subject", "Student Name", "Quiz Number", "Score", "Date Taken"])
            df_quiz.sort_values(by=["Subject", "Section", "Quiz Number", "Student Name"], inplace=True)
            df_quiz.to_excel(writer, sheet_name="Quiz Scores", index=False)

        # Step 3: Fetch activity scores from school.db
        conn_acts = sqlite3.connect(os.path.join(os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "..")), 'school.db'))
        c_acts = conn_acts.cursor()
        activity_results = []

        for section, subject in assignments:
            c_acts.execute("""
                           SELECT student_name, activity_number, activity_name, score, date_taken
                           FROM activity_scores
                           WHERE instructor = ? AND section = ? AND subject = ?
                           """, (instructor_name, section, subject))
            rows = c_acts.fetchall()
            for row in rows:
                activity_results.append((section, subject, *row))

        conn_acts.close()

        if activity_results:
            df_acts = pd.DataFrame(activity_results, columns=[
                "Section", "Subject", "Student Name", "Activity Number", "Activity Name", "Score", "Date Taken"])
            df_acts.sort_values(by=["Subject", "Section", "Activity Number", "Student Name"], inplace=True)
            df_acts.to_excel(writer, sheet_name="Activity Scores", index=False)

        # Save Excel file
        writer.close()
        messagebox.showinfo("Exported", f"Scores exported to {filename}")

    except Exception as e:
        messagebox.showerror("Export Failed", f"Error: {e}")


def export_scores_by_instructor_ui():

    popup = tk.Toplevel(window)

    popup.title("Export Scores by Instructor")

    # =========================
    # WINDOW SIZE
    # =========================
    width = 550
    height = 490

    # =========================
    # CENTER WINDOW
    # =========================
    screen_width = popup.winfo_screenwidth()
    screen_height = popup.winfo_screenheight()

    x = int((screen_width / 2) - (width / 2))
    y = int((screen_height / 2) - (height / 2))

    popup.geometry(f"{width}x{height}+{x}+{y}")
    popup.configure(bg="white")

    popup.resizable(False, False)
    popup.transient(window)
    popup.grab_set()
    popup.focus_force()
    popup.lift()

    popup.attributes("-topmost", True)

    popup.after(
        200,
        lambda: popup.attributes("-topmost", False)
    )

    title_label = tk.Label(
        popup,
        text="📊 Export Scores by Instructor",
        font=("Helvetica", 18, "bold"),
        bg="white",
        fg="#4B0082"
    )

    title_label.pack(pady=(25, 15))

    tk.Label(popup, text="Export All Scores for Instructor", font=("Helvetica", 18, "bold"),
             fg="#4B0082", bg="white").pack(pady=20)

    tk.Label(popup, text="Select Instructor:", font=("Helvetica", 12), bg="white").pack()

    # ✅ Define inside this function
    instructor_var = tk.StringVar()
    instructor_dropdown = ttk.Combobox(popup, textvariable=instructor_var, state="readonly", width=50)
    instructor_dropdown.pack(pady=10)

    # Fetch instructors (excluding Admin)
    try:
        conn = sqlite3.connect(os.path.join(os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "..")), 'teacher_management.db'))
        c = conn.cursor()
        c.execute("SELECT DISTINCT name FROM instructors WHERE name != 'Admin'")
        names = [row[0] for row in c.fetchall()]
        instructor_dropdown['values'] = names
        conn.close()
    except Exception as e:
        messagebox.showerror("Error", f"Failed to load instructors: {e}")
        return

    # ✅ This is now inside the same function so it can access `instructor_var`
    def export_for_selected():
        selected = instructor_var.get()
        if not selected:
            messagebox.showwarning("No Selection", "Please select an instructor.")
        else:
            export_all_scores_for_instructor(selected)

    ttk.Button(popup, text="📤 Export All Scores", command=export_for_selected).pack(pady=10, ipadx=10, ipady=6)
    ttk.Button(popup, text="❌ Close", command=popup.destroy).pack(pady=5)


def remove_all_students_from_instructor(instructor_name):
    try:
        conn = sqlite3.connect(os.path.join(os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "..")), 'teacher_management.db'))
        c = conn.cursor()
        c.execute("DELETE FROM student_assignments WHERE instructor = ?", (instructor_name,))
        conn.commit()
        conn.close()
        messagebox.showinfo("Success", f"All student assignments for instructor '{instructor_name}' have been removed.")
    except Exception as e:
        messagebox.showerror("Error", f"Failed to remove students from instructor:\n{e}")


def clean_instructor_data(instructor_name, year):
    try:
        # Clean activity_scores and attendance in school.db
        conn = sqlite3.connect(os.path.join(os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "..")), 'school.db'))
        c = conn.cursor()

        # Delete activity scores
        c.execute("""
                  DELETE
                  FROM activity_scores
                  WHERE instructor = ?
                    AND date_taken LIKE ?
                  """, (instructor_name, f"{year}-%"))

        # Delete attendance records
        c.execute("""
                  DELETE
                  FROM attendance
                  WHERE instructor = ? AND date LIKE ?
                  """, (instructor_name, f"{year}-%"))

        conn.commit()
        conn.close()

        # Clean quiz scores in separate DB
        conn_q = sqlite3.connect(os.path.join(os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "..")), 'student_quiz_scores.db'))
        c_q = conn_q.cursor()

        c_q.execute("""
                    DELETE
                    FROM quiz_scores
                    WHERE date_taken LIKE ?
                    """, (f"{year}-%",))

        conn_q.commit()
        conn_q.close()

        messagebox.showinfo("Cleanup Complete", f"All data for year {year} has been removed for {instructor_name}.")

    except Exception as e:
        messagebox.showerror("Error", f"Failed to clean data: {e}")


def export_all_scores_for_instructor(instructor_name):
    def grouped_rows_with_headers(df, group_keys, display_prefix):
        rows = []
        grouped = df.groupby(group_keys)
        for keys, group in grouped:
            header = f"{display_prefix} {keys[0]}: {keys[1]}" if isinstance(keys, tuple) else f"{display_prefix} {keys}"
            rows.append([header])
            rows.append(group.columns.tolist())
            rows.extend(group.values.tolist())
            rows.append([])
        return pd.DataFrame(rows)

    try:
        print(f"🟣 Starting export for: {instructor_name}")
        import os
        # 🔹 Fetch assignments
        conn_assign = sqlite3.connect(os.path.join(os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "..")), 'teacher_management.db'))
        c_assign = conn_assign.cursor()
        c_assign.execute("SELECT section, subject FROM teacher_subjects WHERE instructor = ?",
                         (instructor_name.strip(),))
        assignments = c_assign.fetchall()
        conn_assign.close()

        if not assignments:
            messagebox.showinfo("No Data", f"{instructor_name} has no assigned section/subject.")
            return
        import os
        # 🔸 Quiz Scores
        quiz_results = []
        conn_q = sqlite3.connect(os.path.join(os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "..")), 'school.db'))

        c_q = conn_q.cursor()
        for section, subject in assignments:
            c_q.execute("""
                        SELECT student_name, quiz_number, score, date_taken
                        FROM quiz_scores
                        WHERE section = ? AND subject = ?
                        """, (section.strip(), subject.strip()))
            rows = c_q.fetchall()
            for row in rows:
                quiz_results.append((section, subject, *row))
        conn_q.close()

        quiz_df = pd.DataFrame(quiz_results, columns=[
            "Section", "Subject", "Student Name", "Quiz Number", "Score", "Date Taken"
        ])
        import os
        # 🔸 Activity Scores
        # 🔸 Activity Scores (FIXED)
        activity_results = []
        conn_a = sqlite3.connect(os.path.join(os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "..")), 'school.db'))
        c_a = conn_a.cursor()

        for section, subject in assignments:
            c_a.execute("""
                        SELECT student_name, activity_number, activity_name, score, date_taken
                        FROM activity_scores
                        WHERE section = ? AND subject = ?
                        """, (section.strip(), subject.strip()))

            rows = c_a.fetchall()
            for row in rows:
                activity_results.append((section, subject, *row))

        conn_a.close()

        activity_df = pd.DataFrame(activity_results, columns=[
            "Section", "Subject", "Student Name",
            "Activity Number", "Activity Name", "Score", "Date Taken"
        ])
        import os
        # 🔸 Attendance Records
        attendance_results = []
        conn_att = sqlite3.connect(os.path.join(os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "..")), 'school.db'))
        c_att = conn_att.cursor()
        for section, subject in assignments:
            c_att.execute("""
                          SELECT date, student_name, section, time_in, subject
                          FROM attendance
                          WHERE instructor = ? AND section = ? AND subject = ?
                          """, (instructor_name.strip(), section.strip(), subject.strip()))
            rows = c_att.fetchall()
            for row in rows:
                attendance_results.append((section, subject, *row))
        conn_att.close()

        attendance_df = pd.DataFrame(attendance_results, columns=[
            "Section", "Subject", "Date", "Student Name", "Section (Dup)", "Time In", "Subject (Dup)"
        ])

        # 🔹 Export to Excel
        from datetime import datetime
        import os
        # 📁 Desktop export folder
        desktop = os.path.join(os.path.expanduser("~"), "Desktop")
        export_folder = os.path.join(desktop, "Exported All Scores")

        # Create folder if it doesn't exist
        os.makedirs(export_folder, exist_ok=True)

        # 📄 File naming: Instructor + Year + Date
        today = datetime.now()
        year = today.year
        date_str = today.strftime("%Y_%m_%d")

        safe_name = instructor_name.replace(" ", "_")
        filename = f"{safe_name}_All_Scores_{year}_{date_str}.xlsx"
        file_path = os.path.join(export_folder, filename)

        with pd.ExcelWriter(file_path, engine='openpyxl') as writer:

            # Quiz Scores
            # 🔹 Export Quiz Scores
            if not quiz_df.empty:
                quiz_df["Quiz Number"] = pd.to_numeric(quiz_df["Quiz Number"], errors='coerce')
                quiz_df.sort_values(by=["Section", "Subject", "Quiz Number", "Student Name"], inplace=True)

                quiz_grouped_rows = []

                outer_group = quiz_df.groupby(["Section", "Subject"])
                for (section, subject), group in outer_group:
                    quiz_grouped_rows.append([f"{subject} — {section}"])  # Section-Subject Header
                    inner_group = group.groupby("Quiz Number")
                    for quiz_num, quiz_group in inner_group:
                        quiz_grouped_rows.append([f"Quiz {int(quiz_num)}"])  # Quiz Header
                        quiz_grouped_rows.append(["Student Name", "Score", "Date Taken"])  # Column headers
                        for _, row in quiz_group.iterrows():
                            quiz_grouped_rows.append([row["Student Name"], row["Score"], row["Date Taken"]])
                        quiz_grouped_rows.append([])  # space after each quiz
                    quiz_grouped_rows.append([])  # space after each section-subject

                quiz_export_df = pd.DataFrame(quiz_grouped_rows)

                quiz_export_df.to_excel(writer, sheet_name="Quiz Scores", index=False, header=False)

            else:
                pd.DataFrame([["No quiz scores found"]], columns=["Note"]).to_excel(
                    writer, sheet_name="Quiz Scores", index=False
                )

            # Activity Scores
            # 🔹 Export Activity Scores
            if not activity_df.empty:
                activity_df["Activity Number"] = pd.to_numeric(activity_df["Activity Number"], errors='coerce')
                activity_df.sort_values(by=["Subject", "Section", "Activity Number", "Student Name"], inplace=True)

                # Group by Section and Subject first, then Activity Number + Name
                activity_grouped_rows = []

                outer_group = activity_df.groupby(["Section", "Subject"])
                for (section, subject), group in outer_group:
                    activity_grouped_rows.append([f"{subject} — {section}"])  # Section-Subject Header
                    inner_group = group.groupby(["Activity Number", "Activity Name"])
                    for (act_num, act_name), act_group in inner_group:
                        activity_grouped_rows.append([f"Activity {act_num}: {act_name}"])  # Activity Header
                        activity_grouped_rows.append(["Student Name", "Score", "Date Taken"])  # Column headers
                        for _, row in act_group.iterrows():
                            activity_grouped_rows.append([row["Student Name"], row["Score"], row["Date Taken"]])
                        activity_grouped_rows.append([])  # space after each activity
                    activity_grouped_rows.append([])  # space after each section-subject

                activity_export_df = pd.DataFrame(activity_grouped_rows)
                activity_export_df.to_excel(writer, sheet_name="Activity Scores", index=False, header=False)

            else:
                pd.DataFrame([["No activity scores found"]], columns=["Note"]).to_excel(
                    writer, sheet_name="Activity Scores", index=False
                )

            # Attendance
            if not attendance_df.empty:
                attendance_df.drop(columns=["Section (Dup)", "Subject (Dup)"], inplace=True)
                attendance_df.sort_values(by=["Date", "Subject", "Section", "Student Name"], inplace=True)

                grouped_attendance_rows = []
                grouped = attendance_df.groupby("Date")

                for date, group in grouped:
                    grouped_attendance_rows.append([f"{date}"])
                    grouped_attendance_rows.append(group.columns.tolist())
                    grouped_attendance_rows.extend(group.values.tolist())
                    grouped_attendance_rows.append([])

                grouped_attendance_df = pd.DataFrame(grouped_attendance_rows)
                grouped_attendance_df.to_excel(writer, sheet_name="Attendance Records", index=False, header=False)
            else:
                pd.DataFrame([["No attendance records found"]], columns=["Note"]).to_excel(writer,
                                                                                           sheet_name="Attendance Records",
                                                                                           index=False)

        print(f"✅ Exported to: {filename}")
        messagebox.showinfo(
            "Export Successful",
            f"All scores exported successfully!\n\n"
            f"📁 Folder: Exported All Scores\n"
            f"📄 File: {filename}"
        )
        # 📂 Open export folder automatically
        try:
            if os.name == "nt":  # Windows
                subprocess.Popen(f'explorer "{export_folder}"')
            elif os.name == "posix":  # macOS / Linux
                subprocess.Popen(["open", export_folder])
        except Exception as e:
            print(f"⚠ Could not open folder: {e}")


    except Exception as e:
        print(f"❌ Exception during export: {e}")
        messagebox.showerror("Export Failed", f"Error: {e}")


def admin_settings():
    """Admin Settings - View, Add, Edit, Remove Admins."""

    settings_window = tk.Toplevel(window)
    settings_window.title("Admin Settings")

    width = 900
    height = 650

    # =========================
    # CENTER WINDOW
    # =========================
    screen_width = settings_window.winfo_screenwidth()
    screen_height = settings_window.winfo_screenheight()

    x = int((screen_width / 2) - (width / 2))
    y = int((screen_height / 2) - (height / 2))

    settings_window.geometry(f"{width}x{height}+{x}+{y}")

    # =========================
    # WINDOW SETTINGS
    # =========================
    settings_window.resizable(False, False)

    # Keep above parent
    settings_window.transient(window)

    # Focus window
    settings_window.focus_force()

    ttk.Label(
        settings_window,
        text="Admin Management Settings",
        font=("Helvetica", 20, "bold")
    ).pack(pady=15)

    ttk.Label(
        settings_window,
        text="Manage admin accounts and RFID access.",
        font=("Helvetica", 10),
        foreground="gray"
    ).pack(pady=(0, 10))

    # =========================
    # TABLE FRAME
    # =========================
    table_frame = ttk.Frame(settings_window)
    table_frame.pack(fill="both", expand=True, padx=20, pady=10)

    columns = ("ID", "Username", "RFID Code")

    tree = ttk.Treeview(
        table_frame,
        columns=columns,
        show="headings",
        height=15
    )

    # Headings
    tree.heading("ID", text="Admin ID")
    tree.heading("Username", text="Username")
    tree.heading("RFID Code", text="RFID Code")

    # Column Sizes
    tree.column("ID", width=100, anchor="center")
    tree.column("Username", width=300, anchor="center")
    tree.column("RFID Code", width=300, anchor="center")

    # Scrollbar
    scrollbar = ttk.Scrollbar(
        table_frame,
        orient="vertical",
        command=tree.yview
    )

    tree.configure(yscrollcommand=scrollbar.set)

    tree.pack(side="left", fill="both", expand=True)
    scrollbar.pack(side="right", fill="y")

    # =========================
    # LOAD ADMINS
    # =========================
    def load_admins():
        tree.delete(*tree.get_children())

        conn = sqlite3.connect(
            os.path.join(
                os.path.abspath(
                    os.path.join(os.path.dirname(__file__), "..", "..")
                ),
                'teacher_management.db'
            )
        )

        c = conn.cursor()

        # PASSWORD REMOVED
        c.execute("SELECT id, username, rfid_code FROM admins")

        rows = c.fetchall()

        conn.close()

        for row in rows:
            tree.insert("", "end", values=row)

    # =========================
    # ADD ADMIN
    # =========================
    def add_admin():

        popup = tk.Toplevel(settings_window)
        popup.title("Add Admin")

        popup.geometry("400x300")
        popup.resizable(False, False)

        popup.transient(settings_window)
        popup.grab_set()
        popup.focus_force()

        ttk.Label(
            popup,
            text="Add New Admin",
            font=("Helvetica", 16, "bold")
        ).pack(pady=15)

        form_frame = ttk.Frame(popup)
        form_frame.pack(pady=10)

        ttk.Label(form_frame, text="Username:").grid(row=0, column=0, padx=10, pady=10)
        username_entry = ttk.Entry(form_frame, width=30)
        username_entry.grid(row=0, column=1)

        ttk.Label(form_frame, text="RFID Code:").grid(row=1, column=0, padx=10, pady=10)
        rfid_entry = ttk.Entry(form_frame, width=30)
        rfid_entry.grid(row=1, column=1)

        def save_admin():

            username = username_entry.get().strip()
            rfid_code = rfid_entry.get().strip()

            if not username or not rfid_code:
                messagebox.showerror("Error", "All fields are required.")
                return

            conn = sqlite3.connect(
                os.path.join(
                    os.path.abspath(
                        os.path.join(os.path.dirname(__file__), "..", "..")
                    ),
                    'teacher_management.db'
                )
            )

            c = conn.cursor()

            try:
                # PASSWORD REMOVED
                c.execute("""
                    INSERT INTO admins (username, rfid_code)
                    VALUES (?, ?)
                """, (username, rfid_code))

                conn.commit()

                messagebox.showinfo(
                    "Success",
                    "Admin added successfully!"
                )

                popup.destroy()
                load_admins()

            except sqlite3.IntegrityError:
                messagebox.showerror(
                    "Error",
                    "Username or RFID already exists!"
                )

            finally:
                conn.close()

        ttk.Button(
            popup,
            text="💾 Save Admin",
            command=save_admin
        ).pack(pady=20)

    # =========================
    # EDIT ADMIN
    # =========================
    def edit_selected():
        selected = tree.focus()

        if not selected:
            messagebox.showerror("Error", "Please select an admin to edit!")
            return

        values = tree.item(selected, "values")

        admin_id = values[0]
        current_username = values[1]
        current_rfid = values[2]

        # =========================
        # Edit Window
        # =========================
        edit_window = tk.Toplevel(settings_window)
        edit_window.title("Edit Admin")

        width = 450
        height = 300

        screen_width = edit_window.winfo_screenwidth()
        screen_height = edit_window.winfo_screenheight()

        x = int((screen_width / 2) - (width / 2))
        y = int((screen_height / 2) - (height / 2))

        edit_window.geometry(f"{width}x{height}+{x}+{y}")

        edit_window.resizable(False, False)

        # Keep above parent only
        edit_window.transient(settings_window)

        # Focus only THIS popup
        edit_window.grab_set()

        # =========================
        # UI
        # =========================
        main_frame = ttk.Frame(edit_window, padding=20)
        main_frame.pack(fill="both", expand=True)

        ttk.Label(
            main_frame,
            text="Edit Admin",
            font=("Helvetica", 18, "bold")
        ).pack(pady=(0, 20))

        # Username
        ttk.Label(main_frame, text="Username:").pack(anchor="w")

        username_var = tk.StringVar(value=current_username)

        username_entry = ttk.Entry(
            main_frame,
            textvariable=username_var,
            width=35
        )
        username_entry.pack(pady=5)

        # RFID
        ttk.Label(main_frame, text="RFID Code:").pack(anchor="w", pady=(10, 0))

        rfid_var = tk.StringVar(value=current_rfid)

        rfid_entry = ttk.Entry(
            main_frame,
            textvariable=rfid_var,
            width=35
        )
        rfid_entry.pack(pady=5)

        # =========================
        # Save Function
        # =========================
        def save_changes():
            new_username = username_var.get().strip()
            new_rfid = rfid_var.get().strip()

            if not new_username or not new_rfid:
                messagebox.showerror(
                    "Error",
                    "All fields are required."
                )
                return

            confirm = messagebox.askyesno(
                "Confirm Update",
                f"Save changes for:\n\n"
                f"Username: {new_username}\n"
                f"RFID: {new_rfid}"
            )

            if not confirm:
                return

            conn = sqlite3.connect(
                os.path.join(
                    os.path.abspath(
                        os.path.join(
                            os.path.dirname(__file__),
                            "..",
                            ".."
                        )
                    ),
                    'teacher_management.db'
                )
            )

            c = conn.cursor()

            try:
                c.execute("""
                    UPDATE admins
                    SET username = ?,
                        rfid_code = ?
                    WHERE id = ?
                """, (new_username, new_rfid, admin_id))

                conn.commit()

                messagebox.showinfo(
                    "Success",
                    "Admin updated successfully!"
                )

                edit_window.destroy()

                load_admins()

            except sqlite3.IntegrityError:
                messagebox.showerror(
                    "Error",
                    "Username or RFID already exists!"
                )

            finally:
                conn.close()

        # =========================
        # Buttons
        # =========================
        button_frame = ttk.Frame(main_frame)
        button_frame.pack(pady=20)

        ttk.Button(
            button_frame,
            text="💾 Save Changes",
            command=save_changes
        ).grid(row=0, column=0, padx=10)

        ttk.Button(
            button_frame,
            text="❌ Cancel",
            command=edit_window.destroy
        ).grid(row=0, column=1, padx=10)
    def remove_selected():

        selected = tree.focus()

        if not selected:
            messagebox.showwarning(
                "No Selection",
                "Please select an admin."
            )
            return

        values = tree.item(selected, "values")

        admin_id = values[0]
        username = values[1]

        confirm = messagebox.askyesno(
            "Confirm Removal",
            f"Remove admin:\n\n{username}?"
        )

        if not confirm:
            return

        conn = sqlite3.connect(
            os.path.join(
                os.path.abspath(
                    os.path.join(os.path.dirname(__file__), "..", "..")
                ),
                'teacher_management.db'
            )
        )

        c = conn.cursor()

        c.execute(
            "DELETE FROM admins WHERE id=?",
            (admin_id,)
        )

        conn.commit()
        conn.close()

        load_admins()

        messagebox.showinfo(
            "Removed",
            "Admin removed successfully!"
        )

    # =========================
    # BUTTONS
    # =========================
    button_frame = ttk.Frame(settings_window)
    button_frame.pack(pady=15)

    ttk.Button(
        button_frame,
        text="➕ Add Admin",
        command=add_admin
    ).grid(row=0, column=0, padx=10)

    ttk.Button(
        button_frame,
        text="✏️ Edit Selected",
        command=edit_selected
    ).grid(row=0, column=1, padx=10)

    ttk.Button(
        button_frame,
        text="🗑 Remove Selected",
        command=remove_selected
    ).grid(row=0, column=2, padx=10)

    # =========================
    # LOAD DATA
    # =========================
    load_admins()


def manage_teacher_assignments(selected_instructors=None):
    """Window for assigning sections and subjects to selected instructors."""
    assign_window = tk.Toplevel()
    assign_window.title("Assign Sections & Subjects")
    assign_window.geometry("600x500")

    ttk.Label(assign_window, text="Assign Sections & Subjects", font=("Helvetica", 14, "bold")).pack(pady=10)

    if not selected_instructors:
        messagebox.showerror("Error", "No instructors selected!")
        assign_window.destroy()
        return

    # Show selected instructors (just label them)
    selected_label = tk.Label(assign_window, text="Selected Instructors:", font=("Helvetica", 12, "bold"))
    selected_label.pack(pady=5)

    for instructor in selected_instructors:
        tk.Label(assign_window, text=instructor, font=("Arial", 10)).pack()

    # Select Year Level
    ttk.Label(assign_window, text="Year Level:", font=("Helvetica", 12)).pack(pady=2)
    year_var = tk.StringVar()
    year_dropdown = ttk.Combobox(assign_window, textvariable=year_var, values=["1", "2", "3", "4"], state="readonly")
    year_dropdown.pack(pady=5)

    # Select Section (updates depending on Year)
    ttk.Label(assign_window, text="Section:", font=("Helvetica", 12)).pack(pady=2)
    section_var = tk.StringVar()
    section_dropdown = ttk.Combobox(assign_window, textvariable=section_var, state="disabled")
    section_dropdown.pack(pady=5)

    def update_sections(event=None):
        year_selected = year_var.get()
        if year_selected:
            section_dropdown["state"] = "readonly"
            section_dropdown["values"] = [
                f"BSIT {year_selected}-1",
                f"BSIT {year_selected}-2",
                f"BSIT {year_selected}-3",
                f"BSIT {year_selected}-4"
            ]

    year_dropdown.bind("<<ComboboxSelected>>", update_sections)

    # Subject Entry
    ttk.Label(assign_window, text="Subject:", font=("Helvetica", 12)).pack(pady=2)
    subject_var = tk.StringVar()
    subject_entry = ttk.Entry(assign_window, textvariable=subject_var, font=("Helvetica", 12))
    subject_entry.pack(pady=5)

    def add_assignments():
        section = section_var.get()
        subject = subject_var.get().strip()

        if not (section and subject):
            messagebox.showerror("Error", "Please complete Section and Subject!")
            return

        conn = sqlite3.connect(os.path.join(os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "..")), 'teacher_management.db'))
        c = conn.cursor()

        for instructor in selected_instructors:
            # Check if already exists
            c.execute("SELECT * FROM teacher_subjects WHERE instructor=? AND section=? AND subject=?",
                      (instructor, section, subject))
            if not c.fetchone():
                c.execute("INSERT INTO teacher_subjects (instructor, section, subject) VALUES (?, ?, ?)",
                          (instructor, section, subject))

        conn.commit()
        conn.close()

        messagebox.showinfo("Success", f"Assigned {len(selected_instructors)} instructor(s) to {section} - {subject}!")
        assign_window.destroy()

    ttk.Button(assign_window, text="Assign", command=add_assignments).pack(pady=20)


def manage_students():
    """Student management window to add/remove students manually or via Excel."""
    student_window = tk.Toplevel(window)
    student_window.title("Student Management")

    width = 900
    height = 750

    screen_width = student_window.winfo_screenwidth()
    screen_height = student_window.winfo_screenheight()

    x = (screen_width // 2) - (width // 2)
    y = (screen_height // 2) - (height // 2)

    student_window.geometry(f"{width}x{height}+{x}+{y}")

    student_window.resizable(False, False)

    # Make popup behave properly
    student_window.transient(window)
    student_window.grab_set()
    student_window.focus_force()

    # Keep above main window
    student_window.lift()
    student_window.attributes("-topmost", True)

    # Remove permanent topmost after opening
    student_window.after(
        200,
        lambda: student_window.attributes("-topmost", False)
    )

    ttk.Label(student_window, text="Manage Students", font=("Helvetica", 14, "bold")).pack(pady=10)

    # **Instructor Selection**
    ttk.Label(student_window, text="Instructor:", font=("Helvetica", 12)).pack(pady=2)
    instructor_var = tk.StringVar()
    instructor_dropdown = ttk.Combobox(student_window, textvariable=instructor_var, state="readonly")
    # Fetch instructors from teacher_subjects
    conn = sqlite3.connect(os.path.join(os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "..")), 'teacher_management.db'))
    c = conn.cursor()
    c.execute("SELECT name FROM instructors")
    instructors = [row[0] for row in c.fetchall()]
    conn.close()

    if instructors:
        instructor_dropdown["values"] = instructors
    else:
        instructor_dropdown["values"] = ["No Instructors Found"]
        instructor_var.set("No Instructors Found")

    instructor_dropdown.pack(pady=5)

    # **Year Level Selection**
    ttk.Label(student_window, text="Year Level:", font=("Helvetica", 12)).pack(pady=2)
    year_var = tk.StringVar()
    year_dropdown = ttk.Combobox(student_window, textvariable=year_var, values=["1", "2", "3", "4"], state="readonly")
    year_dropdown.pack(pady=5)

    # **Section Selection (Now Two Columns: Year & Section)**
    ttk.Label(student_window, text="Section:", font=("Helvetica", 12)).pack(pady=2)
    section_var = tk.StringVar()
    section_dropdown = ttk.Combobox(student_window, textvariable=section_var, state="disabled")
    section_dropdown.pack(pady=5)

    # **Subject Selection (Filtered by Instructor & Section)**
    ttk.Label(student_window, text="Subject:", font=("Helvetica", 12)).pack(pady=2)
    subject_var = tk.StringVar()
    subject_dropdown = ttk.Combobox(student_window, textvariable=subject_var, state="disabled")
    subject_dropdown.pack(pady=5)

    # **Update Sections Based on Year Level**
    def update_sections(event=None):
        year_selected = year_var.get()
        if year_selected:
            section_dropdown["state"] = "readonly"
            section_dropdown["values"] = [
                f"BSIT {year_selected}-1",
                f"BSIT {year_selected}-2",
                f"BSIT {year_selected}-3",
                f"BSIT {year_selected}-4"
            ]

    # **Update Subjects Based on Instructor & Section**
    def update_subjects(event=None):
        instructor_selected = instructor_var.get()
        section_selected = section_var.get()

        if instructor_selected and section_selected:
            conn = sqlite3.connect(os.path.join(os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "..")), 'teacher_management.db'))
            c = conn.cursor()
            c.execute("""
                      SELECT DISTINCT subject
                      FROM teacher_subjects
                      WHERE instructor = ? AND section = ?
                      """, (instructor_selected, section_selected))
            subjects = [row[0] for row in c.fetchall()]
            conn.close()

            if subjects:
                subject_dropdown["state"] = "readonly"
                subject_dropdown["values"] = subjects
            else:
                subject_dropdown["state"] = "disabled"
                subject_dropdown.set("")
                messagebox.showwarning("No Subjects Found", "No subjects assigned to this instructor.")

    instructor_dropdown.bind("<<ComboboxSelected>>", update_subjects)
    year_dropdown.bind("<<ComboboxSelected>>", update_sections)
    section_dropdown.bind("<<ComboboxSelected>>", update_subjects)

    # **Student Fields**
    ttk.Label(student_window, text="Student Name:", font=("Helvetica", 12)).pack(pady=2)
    student_name_entry = ttk.Entry(student_window, font=("Helvetica", 12))
    student_name_entry.pack(pady=5)

    ttk.Label(student_window, text="Student Number:", font=("Helvetica", 12)).pack(pady=2)
    student_number_entry = ttk.Entry(student_window, font=("Helvetica", 12))
    student_number_entry.pack(pady=5)

    ttk.Label(student_window, text="RFID Code:", font=("Helvetica", 12)).pack(pady=2)
    rfid_entry = ttk.Entry(student_window, font=("Helvetica", 12))
    rfid_entry.pack(pady=5)

    def add_student():
        student_name = student_name_entry.get().strip()
        student_number = student_number_entry.get().strip()
        rfid_code = rfid_entry.get().strip()
        section = section_var.get()
        subject = subject_var.get()
        instructor = instructor_var.get()

        if not (student_name and student_number and rfid_code and section and subject and instructor):
            messagebox.showerror("Error", "All fields are required!")
            return

        try:
            with sqlite3.connect(os.path.join(os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "..")), 'teacher_management.db')) as conn:
                c = conn.cursor()

                # ✅ Add to students table (basic info)
                c.execute("""
                          INSERT
                          OR IGNORE INTO students (name, student_number, rfid_code)
                    VALUES (?, ?, ?)
                          """, (student_name, student_number, rfid_code))

                # ✅ Check if already assigned
                c.execute("""
                          SELECT 1
                          FROM student_assignments
                          WHERE student_number = ? AND section = ? AND subject = ? AND instructor = ?
                          """, (student_number, section, subject, instructor))

                if not c.fetchone():
                    # ✅ Assign to instructor's section/subject
                    c.execute("""
                              INSERT INTO student_assignments (student_number, section, subject, instructor)
                              VALUES (?, ?, ?, ?)
                              """, (student_number, section, subject, instructor))
                    conn.commit()
                    messagebox.showinfo("Success", f"Student {student_name} added successfully!")
                else:
                    messagebox.showwarning("Already Assigned",
                                           "This student is already assigned to that section and subject.")
        except sqlite3.OperationalError as e:
            messagebox.showerror("Database Locked", f"Database is currently locked:\n{e}")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to add student:\n{e}")

    add_student_button = ttk.Button(student_window, text="Add Student", command=add_student)
    add_student_button.pack(pady=10)
    add_student_button = ttk.Button(student_window, text="Upload Students from Excel", command=upload_students_excel)
    add_student_button.pack(pady=10)
    add_student_button = ttk.Button(student_window, text="Download Excel Template", command=download_excel_template)
    add_student_button.pack(pady=10)


import openpyxl
from tkinter import filedialog, messagebox


def download_excel_template():

    try:

        # =========================
        # IMPORTS (SAFE INSIDE FUNC)
        # =========================
        import subprocess
        from datetime import datetime

        # =========================
        # DESKTOP PATH
        # =========================
        desktop = os.path.join(
            os.path.expanduser("~"),
            "Desktop"
        )

        # =========================
        # TEMPLATE FOLDER
        # =========================
        template_folder = os.path.join(
            desktop,
            "Student Folder",
            "Excel Templates"
        )

        os.makedirs(
            template_folder,
            exist_ok=True
        )

        # =========================
        # AUTO FILE NAME
        # =========================
        current_date = datetime.now().strftime(
            "%Y-%m-%d_%H-%M-%S"
        )

        file_name = (
            f"Student_Upload_Template_"
            f"{current_date}.xlsx"
        )

        file_path = os.path.join(
            template_folder,
            file_name
        )

        # =========================
        # CREATE WORKBOOK
        # =========================
        wb = openpyxl.Workbook()

        ws = wb.active

        ws.title = "Student Upload"

        # =========================
        # HEADERS
        # =========================
        headers = [
            "Student Name",
            "Student Number",
            "RFID",
            "Section",
            "Subject",
            "Instructor (Optional)"
        ]

        ws.append(headers)

        # =========================
        # HEADER STYLE
        # =========================
        for col in range(1, len(headers) + 1):

            cell = ws.cell(
                row=1,
                column=col
            )

            cell.font = openpyxl.styles.Font(
                bold=True
            )

        # =========================
        # SAMPLE ROWS
        # =========================
        for _ in range(10):

            ws.append([
                "",
                "",
                "",
                "",
                "",
                ""
            ])

        # =========================
        # COLUMN WIDTHS
        # =========================
        column_widths = {
            "A": 30,
            "B": 25,
            "C": 20,
            "D": 20,
            "E": 25,
            "F": 30
        }

        for col, width in column_widths.items():

            ws.column_dimensions[col].width = width
        ws_list = wb.create_sheet(
            "Valid Lists"
        )

        ws_list.append([
            "Instructor",
            "Section",
            "Subject"
        ])

        # Bold valid-list headers
        for cell in ws_list[1]:

            cell.font = openpyxl.styles.Font(
                bold=True
            )
        try:

            conn = sqlite3.connect(
                os.path.join(
                    os.path.abspath(
                        os.path.join(
                            os.path.dirname(__file__),
                            "..",
                            ".."
                        )
                    ),
                    "teacher_management.db"
                )
            )

            c = conn.cursor()

            c.execute("""
                SELECT instructor, section, subject
                FROM teacher_subjects
                ORDER BY instructor, section, subject
            """)

            rows = c.fetchall()

            for inst, sec, sub in rows:

                ws_list.append([
                    inst,
                    sec,
                    sub
                ])

            conn.close()

        except Exception as db_error:

            print(
                "⚠ Failed loading valid lists:",
                db_error
            )

        wb.save(file_path)

        messagebox.showinfo(
            "Template Downloaded",
            f"Excel template saved successfully!\n\n{file_path}"
        )
        subprocess.Popen(
            f'explorer "{os.path.realpath(template_folder)}"'
        )

    except Exception as e:

        messagebox.showerror(
            "Download Failed",
            f"Failed to generate template:\n\n{e}"
        )
def upload_students_excel():
    file_path = filedialog.askopenfilename(
        title="Select Excel File",
        filetypes=[("Excel Files", "*.xlsx *.xls")]
    )

    if not file_path:
        return

    try:
        df = pd.read_excel(file_path)
        # ✅ normalize headers + remove fully-empty rows
        df.columns = [str(c).strip() for c in df.columns]
        df = df.dropna(how="all")

        # ✅ also drop rows where required fields are missing
        for col in ["Student Name", "Student Number", "RFID"]:
            if col in df.columns:
                df[col] = df[col].astype(str).str.strip()

        df = df[df["Student Name"].astype(str).str.strip().ne("")]
        df = df[df["Student Number"].astype(str).str.strip().ne("")]
        df = df[df["RFID"].astype(str).str.strip().ne("")]

    except Exception as e:
        messagebox.showerror("Error", f"Failed to read file:\n{e}")
        return

    required_columns = {"Student Name", "Student Number", "RFID"}
    if not required_columns.issubset(set(df.columns)):
        messagebox.showerror("Error", "Excel must contain: Student Name, Student Number, RFID.")
        return

    preview_window = tk.Toplevel()
    preview_window.title("📄 Students to be Added")
    preview_window.geometry("900x600")
    preview_window.configure(bg="#F39C12")  # ✅ orange background

    # ✅ Center white card
    card = tk.Frame(preview_window, bg="white", bd=2, relief="ridge")
    card.place(relx=0.5, rely=0.5, anchor="center", width=820, height=520)

    # Header
    tk.Label(card, text="Upload Students From Excel", font=("Arial", 16, "bold"),
             bg="white").pack(pady=(15, 5))

    tk.Label(card, text=f"📁 File: {file_path}", wraplength=780, justify="center",
             font=("Arial", 10), fg="blue", bg="white").pack(pady=(0, 5))

    tk.Label(card, text=f"👥 Students Found: {len(df)}", font=("Arial", 12, "bold"),
             bg="white").pack(pady=(0, 10))

    # ✅ Scroll preview (optional but nice)
    preview_box = tk.Frame(card, bg="white")
    preview_box.pack(fill="both", expand=False, padx=20, pady=(0, 10))

    canvas = tk.Canvas(preview_box, bg="#f7f7f7", height=180, highlightthickness=1, highlightbackground="#ddd")
    vsb = ttk.Scrollbar(preview_box, orient="vertical", command=canvas.yview)
    canvas.configure(yscrollcommand=vsb.set)

    vsb.pack(side="right", fill="y")
    canvas.pack(side="left", fill="both", expand=True)

    scroll_frame = tk.Frame(canvas, bg="#f7f7f7")
    canvas.create_window((0, 0), window=scroll_frame, anchor="nw")
    scroll_frame.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))

    for i, row in df.iterrows():
        name = str(row.get("Student Name", "")).strip()
        sn = str(row.get("Student Number", "")).strip()
        rfid = str(row.get("RFID", "")).strip()
        ttk.Label(scroll_frame, text=f"{i + 1}. {name} | {sn} | {rfid}",
                  font=("Arial", 11)).pack(anchor="w", padx=10, pady=2)

    # ✅ Controls area (centered)
    controls = tk.Frame(card, bg="white")
    controls.pack(pady=(10, 0))

    # ---- Instructor (Optional) ----
    tk.Label(controls, text="👩‍🏫 Instructor (Optional):", bg="white",
             font=("Arial", 11, "bold")).grid(row=0, column=0, padx=10, pady=5, sticky="e")

    instructor_var = tk.StringVar()
    instructor_dropdown = ttk.Combobox(controls, textvariable=instructor_var, width=35, state="readonly")
    instructor_dropdown.grid(row=0, column=1, padx=10, pady=5)

    # Load instructors
    conn = sqlite3.connect(os.path.join(os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "..")), 'teacher_management.db'))
    c = conn.cursor()
    c.execute("SELECT DISTINCT name FROM instructors WHERE name != 'Admin'")
    instructors = [r[0] for r in c.fetchall()]
    conn.close()

    # ✅ THIS is where you add it:
    instructor_dropdown["values"] = [""] + instructors  # blank = optional
    instructor_var.set("")  # start blank

    # ---- Section ----
    tk.Label(controls, text="🏫 Section:", bg="white",
             font=("Arial", 11, "bold")).grid(row=1, column=0, padx=10, pady=5, sticky="e")

    section_var = tk.StringVar()
    section_dropdown = ttk.Combobox(controls, textvariable=section_var, width=35, state="readonly")
    section_dropdown.grid(row=1, column=1, padx=10, pady=5)

    # ---- Subject ----
    tk.Label(controls, text="📘 Subject:", bg="white",
             font=("Arial", 11, "bold")).grid(row=2, column=0, padx=10, pady=5, sticky="e")

    subject_var = tk.StringVar()
    subject_dropdown = ttk.Combobox(controls, textvariable=subject_var, width=35, state="readonly")
    subject_dropdown.grid(row=2, column=1, padx=10, pady=5)

    # ✅ section -> subjects map (combo-safe)
    section_subject_map = {}

    def refresh_section_subject_lists():
        section_subject_map.clear()
        subject_dropdown["values"] = []
        subject_var.set("")

        inst = instructor_var.get().strip()

        conn = sqlite3.connect(os.path.join(os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "..")), 'teacher_management.db'))
        c = conn.cursor()

        if inst:
            c.execute("SELECT DISTINCT section, subject FROM teacher_subjects WHERE instructor=?", (inst,))
        else:
            c.execute("SELECT DISTINCT section, subject FROM teacher_subjects")

        rows = c.fetchall()
        conn.close()

        for sec, sub in rows:
            section_subject_map.setdefault(sec, []).append(sub)

        sections = sorted(section_subject_map.keys())
        section_dropdown["values"] = sections
        section_var.set(sections[0] if sections else "")

        # auto-load subject list for the first section
        update_subjects()

    def update_subjects(event=None):
        sec = section_var.get().strip()
        subs = sorted(section_subject_map.get(sec, []))
        subject_dropdown["values"] = subs
        subject_var.set(subs[0] if subs else "")

    instructor_dropdown.bind("<<ComboboxSelected>>", lambda e: refresh_section_subject_lists())
    section_dropdown.bind("<<ComboboxSelected>>", update_subjects)

    # initial load (no instructor filter)
    refresh_section_subject_lists()

    # ✅ Upload button row
    btns = tk.Frame(card, bg="white")
    btns.pack(pady=20)
    def upload_to_db():
        instructor = instructor_var.get().strip()  # may be ""
        section = section_var.get().strip()
        subject = subject_var.get().strip()

        if not section or not subject:
            messagebox.showwarning("Missing Info", "Please select Section and Subject.")
            return

        try:
            conn = sqlite3.connect(os.path.join(os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "..")), 'teacher_management.db'))
            c = conn.cursor()
            added = 0

            for _, row in df.iterrows():
                name = str(row.get("Student Name", "")).strip()
                student_number = str(row.get("Student Number", "")).strip()
                rfid = str(row.get("RFID", "")).strip()

                if not (name and student_number and rfid):
                    continue

                c.execute("""
                          INSERT
                          OR IGNORE INTO students (name, student_number, rfid_code)
                    VALUES (?, ?, ?)
                          """, (name, student_number, rfid))

                c.execute("""
                          INSERT
                          OR IGNORE INTO student_assignments (student_number, section, subject, instructor)
                    VALUES (?, ?, ?, ?)
                          """, (student_number, section, subject, instructor if instructor else "N/A"))

                added += 1

            conn.commit()
            conn.close()

            preview_window.destroy()
            messagebox.showinfo("✅ Upload Complete",
                                f"Uploaded: {added}\nAssigned: {section} - {subject}\nInstructor: {instructor if instructor else 'N/A'}")
        except Exception as e:
            messagebox.showerror("Upload Error", f"Failed to upload students:\n{e}")

    ttk.Button(btns, text="✅ Upload These Students", command=upload_to_db).grid(row=0, column=0, padx=20, ipadx=10,
                                                                                ipady=5)
    ttk.Button(btns, text="❌ Cancel", command=preview_window.destroy).grid(row=0, column=1, padx=20, ipadx=10, ipady=5)


def show_sections(teacher):
    """Show assigned sections and subjects for the logged-in instructor."""
    clear_grid()

    # 🟧 Orange Button Style Setup
    style = ttk.Style()
    style.configure("Orange.TButton", font=("Helvetica", 12, "bold"), background="#FF8C00", foreground="white")
    style.map("Orange.TButton",
              background=[("active", "#E07B00")],
              foreground=[("active", "white")])

    # 🔳 Enlarged Main Frame in the Center
    frame = ttk.Frame(window, padding=30, borderwidth=2, relief="ridge", style="TFrame")
    frame.place(relx=0.5, rely=0.5, anchor="center", width=600, height=450)

    # ✅ Title Label (WRAPS if name is long)
    title_label = ttk.Label(
        frame,
        text=f"Welcome Professor {teacher}",
        font=("Helvetica", 22, "bold"),
        anchor="center",
        justify="center",
        wraplength=540  # must be slightly smaller than frame width
    )
    title_label.pack(pady=(10, 20), fill="x")

    # ✅ Connect to DB & Get Assigned Sections/Subjects
    conn = sqlite3.connect(os.path.join(os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "..")), 'teacher_management.db'))
    c = conn.cursor()
    c.execute("SELECT section, subject FROM teacher_subjects WHERE instructor = ?", (teacher,))
    assigned_data = c.fetchall()
    conn.close()

    if not assigned_data:
        messagebox.showwarning("No Assignments", "No sections or subjects assigned to you. Contact admin.")
        return show_teacher_selection()

    # ✅ Section Dropdown
    ttk.Label(frame, text="Select Section:", font=("Helvetica", 16)).pack(pady=(0, 5))
    section_var = tk.StringVar()
    section_dropdown = ttk.Combobox(frame, textvariable=section_var, state="readonly", font=("Helvetica", 14), width=30)
    section_dropdown.pack(pady=(0, 10))

    # ✅ Subject Dropdown (Initially Disabled)
    ttk.Label(frame, text="Select Subject:", font=("Helvetica", 16)).pack(pady=(10, 5))
    subject_var = tk.StringVar()
    subject_dropdown = ttk.Combobox(frame, textvariable=subject_var, state="disabled", font=("Helvetica", 14), width=30)
    subject_dropdown.pack(pady=(0, 20))

    # ✅ Populate Sections Dropdown
    sections = list(set([sec for sec, _ in assigned_data]))
    section_dropdown["values"] = sections

    def update_subject_dropdown(event=None):
        """Updates the subject dropdown based on the selected section."""
        selected_section = section_var.get()
        if selected_section:
            subject_dropdown["state"] = "readonly"
            subjects = [sub for sec, sub in assigned_data if sec == selected_section]
            subject_dropdown["values"] = subjects
        else:
            subject_dropdown["state"] = "disabled"
            subject_var.set("")

    section_dropdown.bind("<<ComboboxSelected>>", update_subject_dropdown)

    # ✅ Proceed Button
    def proceed():
        selected_section = section_var.get()
        selected_subject = subject_var.get()
        if selected_section and selected_subject:
            frame.destroy()  # Removes the frame and buttons before proceeding
            show_teacher_options(teacher, selected_section, selected_subject)
        else:
            messagebox.showerror("Error", "Please select a section and subject!")

    ttk.Button(frame, text="Proceed", style="Orange.TButton", command=proceed).pack(pady=10)

    # ✅ Back Button
    ttk.Button(frame, text="Back", style="Orange.TButton", command=show_teacher_selection).pack()


# Hover effects: Change button state to 'active' (light violet with dark font)
def on_enter(button):
    button.state(['active'])


# Remove hover effects: Revert to normal state (pink with white font)
def on_leave(button):
    button.state(['!active'])


def handle_option(option):
    messagebox.showinfo("Option Selected", f"You selected: {option}")


def get_subjects_for_teacher(teacher, section):
    """Fetch subjects assigned to a teacher for a specific section."""
    conn = sqlite3.connect(os.path.join(os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "..")), 'teacher_management.db'))
    c = conn.cursor()
    c.execute("SELECT subject FROM teacher_subjects WHERE instructor = ? AND section = ?", (teacher, section))
    subjects = [row[0] for row in c.fetchall()]
    conn.close()
    return subjects


def add_subject(teacher, section, subject):
    """Adds a subject for a teacher in a specific section."""
    conn = sqlite3.connect(os.path.join(os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "..")), 'teacher_management.db'))
    c = conn.cursor()

    # Ensure the table exists
    c.execute('''CREATE TABLE IF NOT EXISTS teacher_subjects
                 (
                     id
                     INTEGER
                     PRIMARY
                     KEY
                     AUTOINCREMENT,
                     instructor
                     TEXT,
                     section
                     TEXT,
                     subject
                     TEXT
                 )''')

    # Prevent duplicates
    c.execute("SELECT COUNT(*) FROM teacher_subjects WHERE instructor = ? AND section = ? AND subject = ?",
              (teacher, section, subject))
    if c.fetchone()[0] == 0:
        c.execute("INSERT INTO teacher_subjects (instructor, section, subject) VALUES (?, ?, ?)",
                  (teacher, section, subject))
        conn.commit()
        messagebox.showinfo("Success", f"Subject '{subject}' added successfully!")
    else:
        messagebox.showwarning("Warning", f"Subject '{subject}' already exists for {teacher} in {section}.")

    conn.close()


def remove_subject(teacher, section, subject):
    """Remove a subject for a teacher from a section."""
    conn = sqlite3.connect(os.path.join(os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "..")), 'teacher_management.db'))
    c = conn.cursor()
    c.execute("DELETE FROM teacher_subjects WHERE instructor = ? AND section = ? AND subject = ?",
              (teacher, section, subject))
    conn.commit()
    conn.close()


def get_subjects_for_teacher(teacher, section):
    conn = sqlite3.connect(os.path.join(os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "..")), 'teacher_management.db'))
    c = conn.cursor()
    c.execute("SELECT subject FROM teacher_subjects WHERE instructor = ? AND section = ?", (teacher, section))
    subjects = [row[0] for row in c.fetchall()]
    conn.close()
    return subjects if subjects else ["No subjects found"]


def show_subjects(teacher, section):
    """Displays available subjects as a dropdown with an option to add new subjects."""
    clear_grid()

    # ✅ Title Labels (Improved UI)
    title_label = ttk.Label(window, text=f"Welcome Professor: {teacher}", font=("Helvetica", 20, "bold"))
    title_label.place(relx=0.5, rely=0.05, anchor='center')

    section_label = ttk.Label(window, text=f"You're in Section: {section} | Select Subject",
                              font=("Helvetica", 16))
    section_label.place(relx=0.5, rely=0.1, anchor='center')

    subjects = get_subjects_for_teacher(teacher, section)

    # ✅ Dropdown for subject selection
    subject_var = tk.StringVar()
    subject_dropdown = ttk.Combobox(square_frame, textvariable=subject_var, values=subjects, state="readonly",
                                    font=("Helvetica", 14), width=30)  # ⬅️ Wider dropdown
    subject_dropdown.pack(pady=10)

    # ✅ Define `proceed()` **inside** `show_subjects()`
    def proceed():
        selected_subject = subject_var.get().strip()
        if selected_subject:
            global current_section, current_subject
            current_section = section  # Assign selected section
            current_subject = selected_subject  # Assign selected subject
            print(f"✅ Section and Subject Set: {current_section} | {current_subject}")  # Debugging
            show_teacher_options(teacher, current_section, current_subject)
        else:
            messagebox.showerror("Error", "Please select a subject!")

    # ✅ Create uniform button style
    button_style = {"style": "Transparent.TButton", "width": 25, "padding": (10, 5)}

    # ✅ "Proceed" button to go to `show_teacher_options()`
    proceed_btn = ttk.Button(square_frame, text="Proceed", **button_style, command=proceed)
    proceed_btn.pack(pady=10)

    # ✅ "Add Subject" button (Opens Pop-up)
    add_btn = ttk.Button(square_frame, text="Add Subject", **button_style,
                         command=lambda: open_add_subject_popup(teacher, section))
    add_btn.pack(pady=10)

    # ✅ "Remove Subject" button
    remove_btn = ttk.Button(square_frame, text="Remove Subject", **button_style,
                            command=lambda: remove_subject_and_refresh(teacher, section, subject_var.get()))
    remove_btn.pack(pady=10)

    # ✅ Back Button
    back_btn = ttk.Button(square_frame, text="Back", **button_style,
                          command=lambda: show_sections(teacher))
    back_btn.pack(pady=20)


def open_add_subject_popup(teacher, section):
    """Opens a pop-up window to add a subject."""
    popup = tk.Toplevel(window)
    popup.title("Add Subject")
    popup.geometry("400x200")
    popup.resizable(False, False)

    ttk.Label(popup, text="Enter Subject Name:", font=("Helvetica", 14)).pack(pady=10)

    new_subject_var = tk.StringVar()
    new_subject_entry = ttk.Entry(popup, textvariable=new_subject_var, font=("Helvetica", 14))
    new_subject_entry.pack(pady=5)

    def add_subject_and_close():
        subject = new_subject_var.get().strip()
        if subject:
            add_subject(teacher, section, subject)
            show_subjects(teacher, section)  # Refresh subjects
        popup.destroy()  # Close pop-up

    submit_btn = ttk.Button(popup, text="Add", style="Transparent.TButton",
                            command=add_subject_and_close)
    submit_btn.pack(pady=10)

    close_btn = ttk.Button(popup, text="Cancel", style="Transparent.TButton",
                           command=popup.destroy)
    close_btn.pack(pady=5)


def add_subject_and_refresh(teacher, section, subject):
    """Adds a subject and refreshes the UI."""
    if subject.strip():  # Ensure subject is not empty
        add_subject(teacher, section, subject.strip())  # ✅ Pass all three arguments correctly
    show_subjects(teacher, section)  # Refresh the UI


def remove_subject_and_refresh(teacher, section, subject):
    """Removes a subject and refreshes the UI."""
    remove_subject(teacher, section, subject)
    show_subjects(teacher, section)


def ensure_activity_scores_columns():
    conn = sqlite3.connect(os.path.join(os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "..")), 'school.db'))
    c = conn.cursor()

    # Step 1: Ensure table exists
    c.execute("""
              CREATE TABLE IF NOT EXISTS activity_scores
              (
                  id
                  INTEGER
                  PRIMARY
                  KEY
                  AUTOINCREMENT,
                  student_name
                  TEXT,
                  activity_number
                  INTEGER,
                  activity_name
                  TEXT,
                  score
                  REAL,
                  date_taken
                  TEXT
              )
              """)

    # Step 2: Check for missing columns
    c.execute("PRAGMA table_info(activity_scores)")
    existing_cols = [col[1] for col in c.fetchall()]

    # Step 3: Add any missing ones
    if "instructor" not in existing_cols:
        c.execute("ALTER TABLE activity_scores ADD COLUMN instructor TEXT")
    if "section" not in existing_cols:
        c.execute("ALTER TABLE activity_scores ADD COLUMN section TEXT")
    if "subject" not in existing_cols:
        c.execute("ALTER TABLE activity_scores ADD COLUMN subject TEXT")

    conn.commit()
    conn.close()


def show_teacher_options(teacher, section, subject):
    global current_instructor, current_section, current_subject
    current_instructor = teacher
    current_section = section
    current_subject = subject
    clear_grid()
    window.title("Instructor Options")

    # ✅ Instructor Info Label
    title_label = ttk.Label(
        window,
        text=f"Instructor: {teacher} | Section: {section} | Subject: {subject}",
        font=("Helvetica", 16, "bold"),
        foreground="#8A2BE2",
        background="white"
    )
    title_label.place(relx=0.5, rely=0.05, anchor="center")

    # ✅ Welcome Message
    welcome_label = ttk.Label(
        window,
        text=f"Hello Professor {teacher},\nWhat would you like to do?",
        font=("Helvetica", 20, "bold"),
        foreground="#8A2BE2",
        background="white"
    )
    welcome_label.place(relx=0.5, rely=0.13, anchor="center")

    # ✅ Main Options Frame
    option_frame = ttk.Frame(window, padding=20)
    option_frame.place(relx=0.5, rely=0.55, anchor="center")

    options = [
        ("Start a Timer for Students", lambda: show_timer_options(teacher, section, subject)),
        ("Start an Activity/Laboratory for Students", lambda: show_activity_options(teacher, section, subject)),
        ("Open a Computer Program", lambda: show_program_selection(teacher, section, subject)),
        ("View Recorded Scores of Students Laboratory/Activities",
         lambda: view_activity_scores("View Recorded Scores of Students Laboratory/Activities")),
        ("Start the RFID Scanner Feature", lambda: start_gui(teacher, section, subject)),
        ("Start a Quiz for them", lambda: show_quiz_upload(teacher, section, subject)),
        ("View Student Quiz Scores", lambda: view_recored_scores_quiz(teacher, section, subject)),
        ("Add/Remove Student on your Section", lambda: show_add_remove_student(teacher, section, subject))
    ]

    # ✅ Render option buttons
    for i, (label, command) in enumerate(options):
        btn = ttk.Button(option_frame, text=label, style="Transparent.TButton", width=50, command=command)
        btn.grid(row=i, column=0, pady=8)

        # Hover effects
        btn.bind("<Enter>", lambda e, b=btn: b.configure(style="TButton"))
        btn.bind("<Leave>", lambda e, b=btn: b.configure(style="Transparent.TButton"))

    # ✅ Back Button
    back_btn = ttk.Button(option_frame, text="Back", style="Secondary.TButton",
                          command=lambda: show_sections(teacher))
    back_btn.grid(row=len(options), column=0, pady=20)


rfid_scanner_active = False
rfid_scanner_mode = False
current_mode = 'attendance'  # Modes: 'sign_in' or 'attendance'
shutdown_time = None  # To store the shutdown time for clients


def get_student_by_rfid(rfid):
    try:
        conn = sqlite3.connect(os.path.join(os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "..")), 'teacher_management.db'))
        c = conn.cursor()

        c.execute("""
                  SELECT s.name, s.student_number, s.rfid_code, a.section, a.subject, a.instructor
                  FROM students s
                           JOIN student_assignments a ON s.student_number = a.student_number
                  WHERE s.rfid_code = ?
                    AND a.instructor = ?
                    AND a.section = ?
                    AND a.subject = ?
                  """, (rfid, current_instructor, current_section, current_subject))

        row = c.fetchone()
        conn.close()

        if row:
            return {
                "name": row[0],
                "student_number": row[1],
                "rfid_code": row[2],
                "section": row[3],
                "subject": row[4],
                "instructor": row[5],
            }
        return None

    except Exception as e:
        print(f"⚠️ Error fetching student by RFID: {e}")
        return None


def save_attendance_record(date, student_name, section, time, instructor, subject, mode="TIME_IN", pc_number="Unknown"):
    conn = sqlite3.connect(os.path.join(os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "..")), 'school.db'))
    c = conn.cursor()

    # Make sure the table exists
    c.execute('''
              CREATE TABLE IF NOT EXISTS attendance
              (
                  id INTEGER PRIMARY KEY AUTOINCREMENT,
                  date TEXT,
                  student_name TEXT,
                  section TEXT,
                  time_in TEXT,
                  instructor TEXT,
                  subject TEXT,
                  time_out TEXT,
                  pc_number TEXT DEFAULT "Unknown"
              )
              ''')
              
    # Ensure time_out column exists for older DBs
    try:
        c.execute("ALTER TABLE attendance ADD COLUMN time_out TEXT")
    except sqlite3.OperationalError:
        pass # Column already exists

    if mode == "TIME_IN":
        # Check for duplicate attendance
        c.execute('''
                  SELECT id
                  FROM attendance
                  WHERE date = ? AND student_name = ? AND section = ? AND instructor = ? AND subject = ?
                  ''', (date, student_name, section, instructor, subject))

        already_exists = c.fetchone()
        if not already_exists:
            # Insert attendance
            c.execute('''
                      INSERT INTO attendance (date, student_name, section, time_in, instructor, subject, time_out, pc_number)
                      VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                      ''', (date, student_name, section, time, instructor, subject, "", pc_number))
            conn.commit()
            print(f"🟢 Recorded Time In: {student_name} - {date} {time}")
            messagebox.showinfo("Time In Success", f"{student_name} timed in at {time}")
        else:
            if pc_number != "Unknown":
                c.execute("UPDATE attendance SET pc_number = ? WHERE id = ?", (pc_number, already_exists[0]))
                conn.commit()
            print(f"🟡 Skipped duplicate: {student_name} - {date}")
            # Suppress popup for duplicate if this was a network sign-in to avoid annoying the instructor
            if pc_number == "Unknown":
                messagebox.showwarning("Already Timed In", f"{student_name} has already timed in today.")
            
    elif mode == "TIME_OUT":
        # Find their record for today
        c.execute('''
                  SELECT id, time_out
                  FROM attendance
                  WHERE date = ? AND student_name = ? AND section = ? AND instructor = ? AND subject = ?
                  ''', (date, student_name, section, instructor, subject))
        record = c.fetchone()
        
        if record:
            record_id, existing_time_out = record
            if existing_time_out:
                messagebox.showwarning("Already Timed Out", f"{student_name} has already timed out at {existing_time_out}.")
            else:
                c.execute("UPDATE attendance SET time_out = ? WHERE id = ?", (time, record_id))
                conn.commit()
                print(f"🔴 Recorded Time Out: {student_name} - {date} {time}")
                messagebox.showinfo("Time Out Success", f"{student_name} timed out at {time}")
        else:
            messagebox.showerror("Time Out Error", f"No Time In record found for {student_name} today.")

    conn.close()

def process_local_rfid(rfid_code):
    global rfid_scanner_mode, current_instructor, current_section, current_subject
    
    if not rfid_scanner_mode:
        messagebox.showwarning("Scanner Off", "Please select Time In or Time Out mode first.")
        return
        
    student = get_student_by_rfid(rfid_code)
    if student:
        current_time = datetime.now().strftime('%H:%M:%S')
        current_date = datetime.now().strftime('%Y-%m-%d')
        
        save_attendance_record(
            current_date,
            student['name'],
            student['section'],
            current_time,
            current_instructor,
            current_subject,
            mode=rfid_scanner_mode
        )
    else:
        messagebox.showerror("Error", "RFID not recognized or student not found in this section/subject.")



def load_attendance_records():
    global attendance_records
    attendance_records = {}

    conn = sqlite3.connect(os.path.join(os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "..")), 'school.db'))
    c = conn.cursor()
    
    # Ensure time_out column exists
    c.execute('''
              CREATE TABLE IF NOT EXISTS attendance
              (
                  id INTEGER PRIMARY KEY AUTOINCREMENT,
                  date TEXT,
                  student_name TEXT,
                  section TEXT,
                  time_in TEXT,
                  instructor TEXT,
                  subject TEXT,
                  time_out TEXT,
                  pc_number TEXT DEFAULT "Unknown"
              )
              ''')
              
    try:
        c.execute("ALTER TABLE attendance ADD COLUMN time_out TEXT")
    except sqlite3.OperationalError:
        pass # Column already exists
    
    try:
        c.execute("ALTER TABLE attendance ADD COLUMN pc_number TEXT DEFAULT 'Unknown'")
    except sqlite3.OperationalError:
        pass # Column already exists

    # 🔹 Filter only current instructor's data
    c.execute('''
              SELECT date, student_name, section, time_in, time_out, instructor, subject, pc_number
              FROM attendance
              WHERE instructor = ? AND section = ? AND subject = ?
              ''', (current_instructor, current_section, current_subject))

    rows = c.fetchall()
    conn.close()

    for row in rows:
        date, name, section, time_in, time_out, instructor, subject, pc_number = row
        if date not in attendance_records:
            attendance_records[date] = []
        attendance_records[date].append({
            'name': name,
            'section': section,
            'time_in': time_in,
            'time_out': time_out if time_out else "N/A",
            'date': date,
            'instructor': instructor,
            'subject': subject
        })


def process_rfid(client_socket, rfid_message):
    global rfid_scanner_active

    context, rfid = rfid_message.split(":")
    student = get_student_by_rfid(rfid)

    if student:
        current_time = datetime.now().strftime('%H:%M:%S')
        current_date = datetime.now().strftime('%Y-%m-%d')

        if context == 'signin':
            response = {"status": "success",
                        "message": f"Student {student['name']}, {student['section']} signed in successfully"}
        elif context == 'attendance':
            if rfid_scanner_active:  # Only allow attendance recording when active
                save_attendance_record(
                    current_date,
                    student['name'],
                    student['section'],
                    current_time,
                    current_instructor,
                    current_subject
                )

                response = {"status": "success",
                            "message": f"Attendance recorded for {student['name']} at {current_time} on {current_date}"}
            else:
                response = {"status": "error",
                            "message": "Attendance recording is disabled. Please wait for the instructor to start."}
        else:
            response = {"status": "error", "message": "Unknown action."}

    else:
        response = {"status": "error", "message": "RFID not recognized or student not found."}

    client_socket.sendall(json.dumps(response).encode())
import sqlite3

def store_quiz_score(student_name, quiz_number, score, section, subject, date_taken):
    try:
        conn = sqlite3.connect(os.path.join(os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "..")), 'school.db'))
        c = conn.cursor()

        print(
            f"📌 Saving to DB: {student_name}, Quiz {quiz_number}, Score: {score}, Section: {section}, Subject: {subject}, Date: {date_taken}")

        c.execute("""
                  INSERT INTO quiz_scores (student_name, quiz_number, score, section, subject, date_taken)
                  VALUES (?, ?, ?, ?, ?, ?)
                  """, (student_name, quiz_number, score, section, subject, date_taken))

        conn.commit()
        conn.close()
        print("✅ Quiz Score Successfully Saved!")

    except Exception as e:
        print(f"⚠️ Database Error: {e}")


def view_activity_scores(title):
    ensure_activity_scores_columns()
    clear_grid()
    window.title(title)

    selected_activity_filter = tk.StringVar()

    # Outer frame (card style)
    main_frame = tk.Frame(window, bg="white", bd=2, relief="groove")
    main_frame.place(relx=0.5, rely=0.5, anchor="center", width=1150, height=760)

    # Header label
    header_label = ttk.Label(
        main_frame,
        text=f"Instructor: {current_instructor}  |  Section: {current_section}  |  Subject: {current_subject}",
        font=("Helvetica", 20, "bold"),
        foreground="#5D3FD3",
        background="white"
    )
    header_label.pack(pady=20)

    # Filter area
    filter_frame = tk.Frame(main_frame, bg="white")
    filter_frame.pack(pady=(0, 10))

    tk.Label(filter_frame, text="Filter by Activity:", font=("Helvetica", 13), bg="white").pack(side="left")
    activity_dropdown = ttk.Combobox(
        filter_frame, textvariable=selected_activity_filter,
        state="readonly", width=60, font=("Helvetica", 12)
    )
    activity_dropdown.pack(side="left", padx=10)

    # Table container
    table_container = tk.Frame(main_frame, bg="white")
    table_container.pack(padx=20, fill="both", expand=True)

    columns = ("Name", "Score", "Date")
    tree = ttk.Treeview(table_container, columns=columns, show="tree headings", height=18)

    # Treeview style
    style = ttk.Style()
    style.configure("Treeview", font=("Helvetica", 13), rowheight=34)
    style.configure("Treeview.Heading", font=("Helvetica", 13, "bold"))

    tree.heading("#0", text="Activity", anchor="w")
    tree.column("#0", width=340, anchor="w")

    for col in columns:
        tree.heading(col, text=col, anchor="center")
        tree.column(col, anchor="center", width=230)

    # Scrollbar
    scrollbar_y = ttk.Scrollbar(table_container, orient="vertical", command=tree.yview)
    scrollbar_y.pack(side="right", fill="y")
    tree.configure(yscrollcommand=scrollbar_y.set)
    tree.pack(fill="both", expand=True)

    def load_scores(filter_activity=None):
        tree.delete(*tree.get_children())
        try:
            conn = sqlite3.connect(os.path.join(os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "..")), 'school.db'))
            c = conn.cursor()
            c.execute("""
                      SELECT student_name, activity_number, activity_name, score, date_taken
                      FROM activity_scores
                      WHERE instructor = ? AND section =? AND subject=?
                      ORDER BY activity_number, activity_name, student_name
                      """, (current_instructor, current_section, current_subject))
            rows = c.fetchall()
            conn.close()

            from collections import defaultdict
            grouped = defaultdict(list)
            for name, num, act_name, score, date in rows:
                key = f"Activity {num}: {act_name}"
                grouped[key].append((f"🧑 {name}", score, date))

            activity_dropdown['values'] = ["All"] + list(grouped.keys())
            if not selected_activity_filter.get():
                selected_activity_filter.set("All")

            for activity_key, students in grouped.items():
                if filter_activity and activity_key != filter_activity:
                    continue
                parent = tree.insert("", "end", text=activity_key)
                for student in students:
                    tree.insert(parent, "end", values=student)

        except Exception as e:
            messagebox.showerror("Database Error", f"Failed to load scores: {e}")

    def on_filter_selected(event):
        value = selected_activity_filter.get()
        load_scores(value if value != "All" else None)

    activity_dropdown.bind("<<ComboboxSelected>>", on_filter_selected)

    # Delete a score
    def remove_score():
        selected_item = tree.focus()
        if not selected_item or not tree.parent(selected_item):
            messagebox.showwarning("No Selection", "Please select a student score to remove.")
            return

        parent = tree.item(tree.parent(selected_item))['text']
        name, score, date = tree.item(selected_item)['values']
        act_number = parent.split(":")[0].split()[-1].strip()
        act_name = parent.split(":")[1].strip()

        confirm = messagebox.askyesno("Confirm", f"Remove score for {name} on {parent}?")
        if confirm:
            try:
                conn = sqlite3.connect(os.path.join(os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "..")), 'school.db'))
                c = conn.cursor()
                c.execute("""
                          DELETE
                          FROM activity_scores
                          WHERE student_name = ?
                            AND activity_number = ?
                            AND activity_name = ?
                            AND instructor = ?
                            AND section =?
                            AND subject=?
                          """, (name.replace("🧑 ", ""), act_number, act_name,
                                current_instructor, current_section, current_subject))
                conn.commit()
                conn.close()
                load_scores(selected_activity_filter.get() if selected_activity_filter.get() != "All" else None)
                messagebox.showinfo("Removed", f"{name}'s score removed.")
            except Exception as e:
                messagebox.showerror("Error", f"Failed to remove: {e}")

    # Export to Excel
    # Export to Excel
    def export_scores():
        try:
            import os
            import subprocess
            from datetime import datetime
            from openpyxl import load_workbook
            from openpyxl.styles import Font, Alignment

            conn = sqlite3.connect(os.path.join(os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "..")), 'school.db'))
            df = pd.read_sql_query("""
                                   SELECT student_name    AS Name,
                                          activity_number AS ActivityNumber,
                                          activity_name   AS ActivityName,
                                          score           AS Score,
                                          date_taken AS Date
                                   FROM activity_scores
                                   WHERE instructor=? AND section =? AND subject=?
                                   ORDER BY activity_number, student_name
                                   """, conn, params=(current_instructor, current_section, current_subject))
            conn.close()

            if df.empty:
                messagebox.showinfo("No Data", "No activity scores to export.")
                return

            # 📂 Desktop paths
            desktop = os.path.join(os.path.expanduser("~"), "Desktop")
            base_folder = os.path.join(desktop, "activity_scores_instructor")
            instructor_folder = os.path.join(base_folder, current_instructor)
            os.makedirs(instructor_folder, exist_ok=True)

            # 🕒 Date
            now = datetime.now()
            date_str = now.strftime("%Y_%m_%d")

            filename = f"{current_instructor}_{current_section}_{current_subject}_{date_str}.xlsx"
            file_path = os.path.join(instructor_folder, filename)

            # ✍ Write Excel with space for header
            with pd.ExcelWriter(file_path, engine="openpyxl") as writer:
                df.to_excel(writer, index=False, startrow=6, sheet_name="Activity Scores")

            # 🎨 Style header
            wb = load_workbook(file_path)
            ws = wb.active

            ws.merge_cells("A1:E1")
            ws.merge_cells("A2:E2")
            ws.merge_cells("A3:E3")
            ws.merge_cells("A4:E4")

            ws["A1"] = "ACTIVITY SCORES REPORT"
            ws["A2"] = f"Instructor: {current_instructor}"
            ws["A3"] = f"Section: {current_section} | Subject: {current_subject}"
            ws["A4"] = f"Date Exported: {now.strftime('%B %d, %Y')}"

            ws["A1"].font = Font(size=16, bold=True)
            ws["A2"].font = Font(size=12, bold=True)
            ws["A3"].font = Font(size=12)
            ws["A4"].font = Font(size=11, italic=True)

            for r in range(1, 5):
                ws[f"A{r}"].alignment = Alignment(horizontal="center")

            wb.save(file_path)

            # 📂 Auto-open folder
            try:
                if os.name == "nt":
                    subprocess.Popen(f'explorer "{instructor_folder}"')
                else:
                    subprocess.Popen(["open", instructor_folder])
            except:
                pass

            messagebox.showinfo(
                "Export Successful",
                f"Activity scores exported successfully!\n\n"
                f"Instructor: {current_instructor}\n"
                f"Section: {current_section}\n"
                f"Subject: {current_subject}\n\n"
                f"File:\n{filename}"
            )

        except Exception as e:
            messagebox.showerror("Export Failed", str(e))

    # Button frame (⬅ this is what fixes the issue)
    # Bottom white button frame, centered horizontally
    bottom_button_frame = tk.Frame(window, bg="white", bd=2, relief="flat")
    bottom_button_frame.place(relx=0.5, rely=0.96, anchor="center", width=1100, height=60)

    # Internal container to center buttons inside the bottom frame
    inner_button_holder = tk.Frame(bottom_button_frame, bg="white")
    inner_button_holder.place(relx=0.5, rely=0.5, anchor="center")

    # Back Button
    ttk.Button(inner_button_holder, text="🔙 Back", bootstyle="secondary",
               command=lambda: [main_frame.destroy(), bottom_button_frame.destroy(),
                                show_teacher_options(current_instructor, current_section, current_subject)]).pack(
        side="left", padx=20, ipadx=12, ipady=6)

    # Remove Score Button
    ttk.Button(inner_button_holder, text="🗑️ Remove Score", bootstyle="danger", command=remove_score).pack(
        side="left", padx=20, ipadx=12, ipady=6)

    # Export to Excel Button
    ttk.Button(inner_button_holder, text="📤 Export to Excel", bootstyle="success", command=export_scores).pack(
        side="left", padx=20, ipadx=12, ipady=6)

    load_scores()


def handle_client(client_socket):
    global client_sockets, current_instructor, current_section, current_subject

    decoder = json.JSONDecoder()
    buffer = ""

    try:
        while True:
            chunk = client_socket.recv(4096).decode("utf-8")
            if not chunk:
                break

            buffer += chunk

            # ✅ Parse and handle ALL complete JSON objects currently in buffer
            while True:
                buffer = buffer.lstrip()
                if not buffer:
                    break

                try:
                    request, idx = decoder.raw_decode(buffer)
                except json.JSONDecodeError:
                    # incomplete JSON, wait for more data
                    break

                buffer = buffer[idx:]
                action = request.get("action")
                if action == "REQUEST_INSTRUCTOR_INFO":
                    conn = sqlite3.connect(os.path.join(os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "..")), 'teacher_management.db'))
                    c = conn.cursor()
                    c.execute(
                        "SELECT section, subject FROM teacher_subjects WHERE instructor=?",
                        (current_instructor,)
                    )
                    assigned_data = c.fetchone()
                    conn.close()

                    if assigned_data:
                        current_section, current_subject = assigned_data
                    else:
                        current_section, current_subject = "N/A", "N/A"

                    instructor_info = {
                        "action": "INSTRUCTOR_INFO",
                        "instructor": current_instructor if current_instructor else "N/A",
                        "section": current_section,
                        "subject": current_subject
                    }
                    client_socket.sendall(json.dumps(instructor_info).encode("utf-8"))
                    print("✅ Sent Instructor Info:", instructor_info)

                elif action == "HELLO_PC":
                    pc_name = request.get("pc_name") or request.get("hostname") or "Unknown-PC"
                    mac_address = request.get("mac_address", "Unknown-MAC")
                    
                    assigned_pc_number = "Unknown"
                    try:
                        conn_pc = sqlite3.connect(os.path.join(os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "..")), 'teacher_management.db'))
                        c_pc = conn_pc.cursor()
                        c_pc.execute("SELECT pc_number FROM pc_assignments WHERE mac_address=?", (mac_address,))
                        res = c_pc.fetchone()
                        if res:
                            assigned_pc_number = res[0]
                        else:
                            c_pc.execute("INSERT INTO pc_assignments (mac_address, pc_number, hostname) VALUES (?, ?, ?)", (mac_address, "New-PC", pc_name))
                            conn_pc.commit()
                            assigned_pc_number = "New-PC"
                        conn_pc.close()
                    except Exception as e:
                        print("Error handling pc_assignments:", e)

                    with clients_lock:
                        cid = client_id_by_sock.get(client_socket)
                        if cid is not None and cid in clients_by_id:
                            clients_by_id[cid]["pc"] = pc_name
                            clients_by_id[cid]["pc_number"] = assigned_pc_number

                    reply = {"action": "WELCOME_PC", "client_id": cid if cid is not None else 0, "label": assigned_pc_number}
                    client_socket.sendall(json.dumps(reply).encode("utf-8"))
                    print(f"✅ HELLO_PC from {pc_name} ({mac_address}) -> assigned {assigned_pc_number}")

                elif action == "signin":
                    rfid = request.get("rfid")
                    try:
                        conn = sqlite3.connect(os.path.join(os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "..")), 'teacher_management.db'))
                        c = conn.cursor()
                        c.execute("""
                            SELECT s.name, a.section, a.subject, a.instructor
                            FROM students s
                            JOIN student_assignments a ON s.student_number = a.student_number
                            WHERE s.rfid_code = ?
                        """, (rfid,))
                        student = c.fetchone()
                        conn.close()

                        if student:
                            name, section, subject, instructor = student
                            response = {
                                "status": "success",
                                "message": f"{name} signed in successfully",
                                "name": name,
                                "section": section,
                                "subject": subject
                            }
                            
                            with clients_lock:
                                cid = client_id_by_sock.get(client_socket)
                                pc_num = clients_by_id[cid].get("pc_number", "Unknown") if cid and cid in clients_by_id else "Unknown"
                                
                            if rfid_scanner_active:
                                cur_time = datetime.now().strftime('%H:%M:%S')
                                cur_date = datetime.now().strftime('%Y-%m-%d')
                                save_attendance_record(cur_date, name, section, cur_time, instructor, subject, mode="TIME_IN", pc_number=pc_num)
                            
                        else:
                            response = {"status": "error", "message": "Student not found or not assigned."}

                        client_socket.sendall(json.dumps(response).encode("utf-8"))

                    except Exception as e:
                        client_socket.sendall(json.dumps({
                            "status": "error",
                            "message": f"Sign-in error: {e}"
                        }).encode("utf-8"))

                elif action == "attendance":
                    rfid = request.get("rfid")

                    # ── Two-step student lookup ─────────────────────────────────────
                    # Step 1: strict match (section + subject + instructor)
                    student = get_student_by_rfid(rfid)

                    # Step 2: fallback — RFID only (catches newly-added students
                    #          whose assignment may not yet match the active filter)
                    if not student:
                        try:
                            conn_fb = sqlite3.connect(os.path.join(os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "..")), 'teacher_management.db'))
                            c_fb = conn_fb.cursor()
                            c_fb.execute("""
                                SELECT s.name, a.section, a.subject, a.instructor
                                FROM students s
                                LEFT JOIN student_assignments a ON s.student_number = a.student_number
                                WHERE s.rfid_code = ?
                                LIMIT 1
                            """, (rfid,))
                            row_fb = c_fb.fetchone()
                            conn_fb.close()
                            if row_fb:
                                student = {
                                    "name": row_fb[0],
                                    "section": row_fb[1] or current_section,
                                    "subject": row_fb[2] or current_subject,
                                    "instructor": row_fb[3] or current_instructor,
                                    "rfid_code": rfid
                                }
                                print(f"⚠️ Fallback RFID lookup used for attendance: {student['name']}")
                        except Exception as fb_err:
                            print(f"⚠️ Fallback lookup error: {fb_err}")

                    if student:
                        if not rfid_scanner_active:
                            response = {
                                "message": "Attendance recording is disabled. Please wait for the instructor to start.",
                                "status": "error"
                            }
                        else:
                            current_time = time.strftime('%H:%M:%S')
                            current_date = time.strftime('%Y-%m-%d')

                            conn = sqlite3.connect(os.path.join(os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "..")), 'school.db'))
                            c = conn.cursor()
                            # Ensure table exists (first run)
                            c.execute('''
                                      CREATE TABLE IF NOT EXISTS attendance
                                      (id INTEGER PRIMARY KEY AUTOINCREMENT,
                                       date TEXT, student_name TEXT, section TEXT,
                                       time_in TEXT, instructor TEXT, subject TEXT)
                                      ''')
                            c.execute('''
                                      SELECT COUNT(*)
                                      FROM attendance
                                      WHERE date = ?
                                        AND student_name = ?
                                        AND instructor = ?
                                        AND section = ?
                                        AND subject = ?
                                      ''', (current_date, student['name'], current_instructor, student['section'],
                                            current_subject))
                            exists = c.fetchone()[0]

                            if exists:
                                response = {
                                    "message": f"{student['name']} already has attendance recorded today.",
                                    "status": "duplicate"
                                }
                            else:
                                with clients_lock:
                                    cid = client_id_by_sock.get(client_socket)
                                    pc_num = clients_by_id[cid].get("pc_number", "Unknown") if cid and cid in clients_by_id else "Unknown"
                                    
                                # ✅ Insert into database
                                c.execute('''
                                          INSERT INTO attendance (date, student_name, section, time_in, instructor, subject, time_out, pc_number)
                                          VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                                          ''', (current_date, student['name'], student['section'], current_time,
                                                current_instructor, current_subject, "", pc_num))
                                conn.commit()

                                # ✅ Add to local memory log
                                if current_date not in attendance_records:
                                    attendance_records[current_date] = []

                                attendance_records[current_date].append({
                                    "name": student["name"],
                                    "section": student["section"],
                                    "time": current_time
                                })

                                # ✅ Trigger refresh on UI thread
                                if 'window' in globals():
                                    try:
                                        window.after(0, refresh_attendance_log)
                                    except Exception as ui_err:
                                        print(f"⚠️ Failed to refresh UI: {ui_err}")

                                response = {
                                    "message": f"Attendance recorded for {student['name']} at {current_time} on {current_date}",
                                    "status": "success"
                                }

                            conn.close()
                    else:
                        response = {"message": "RFID not recognized or student not found.", "status": "error"}

                    client_socket.sendall(json.dumps(response).encode('utf-8'))



                elif action == "get_activities":

                    try:

                        req_instructor = str(request.get("instructor", "")).strip()

                        req_section = str(request.get("section", "")).strip()

                        req_subject = str(request.get("subject", "")).strip()

                        # If student didn’t send these, return empty (safer)

                        if not req_instructor or not req_section or not req_subject or req_instructor.upper() == "N/A":
                            client_socket.sendall(json.dumps({}).encode("utf-8"))

                            continue

                        filtered = {}

                        for key, details in activities.items():

                            if not isinstance(details, dict):
                                continue

                            if (str(details.get("instructor", "")).strip() == req_instructor and

                                    str(details.get("section", "")).strip() == req_section and

                                    str(details.get("subject", "")).strip() == req_subject):
                                filtered[key] = {

                                    "status": "Accepting" if details.get("accepting", False) else "Closed",

                                    "subject": details.get("subject", "Unknown"),

                                    "accepting": details.get("accepting", False),

                                    "instruction_text": details.get("instruction_text", None),

                                    "instruction_image": details.get("instruction_image", None),

                                    "time_limit": details.get("time_limit", None),

                                    "language": details.get("language", "Python"),

                                    "test_input": details.get("test_input", "")

                                }

                        client_socket.sendall(json.dumps(filtered).encode("utf-8"))


                    except Exception as e:

                        client_socket.sendall(json.dumps({"error": str(e)}).encode("utf-8"))



                elif action == "get_scores":
                    student_name = request.get("student_name")
                    section = request.get("section")
                    subject = request.get("subject")

                    try:
                        conn = sqlite3.connect(os.path.join(os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "..")), 'school.db'))
                        c = conn.cursor()
                        c.execute("""
                            SELECT activity_number, activity_name, score
                            FROM activity_scores
                            WHERE student_name = ? AND section = ? AND subject = ?
                            ORDER BY date_taken DESC
                        """, (student_name, section, subject))
                        rows = c.fetchall()
                        conn.close()

                        if rows:
                            scores_dict = {f"Activity {num}: {name}": f"{score}%" for num, name, score in rows}
                            print(f"📤 Sending DB activity scores to {student_name}: {len(scores_dict)}")
                            send_json_response(client_socket, scores_dict)
                        else:
                            send_json_response(client_socket, {})
                    except Exception as e:
                        print(f"❌ Failed to fetch activity scores from DB: {e}")
                        send_json_response(client_socket, {"error": "Failed to retrieve activity scores."})



                elif action == "submit_activity":

                    try:

                        load_activities()  # ✅ always load latest activities.json before grading

                        student_name = request.get("student_name", "Unknown")

                        activity_key = request.get("activity_key")

                        student_code = request.get("code", "")

                        instructor = request.get("instructor", "N/A")

                        section = request.get("section", "N/A")

                        subject = request.get("subject", "N/A")

                        language = (request.get("language") or "Python").strip()

                        if not activity_key:
                            send_json_response(client_socket, {

                                "action": "activity_result", "ok": False,

                                "error": "Missing activity_key", "similarity": 0.0

                            })

                            continue

                        if activity_key not in activities:
                            send_json_response(client_socket, {

                                "action": "activity_result", "ok": False,

                                "error": "Activity not found", "similarity": 0.0

                            })

                            continue

                        if not activities[activity_key].get("accepting", False):
                            send_json_response(client_socket, {

                                "action": "activity_result", "ok": False,

                                "error": "Submissions are closed", "similarity": 0.0

                            })

                            continue

                        teacher_file_path = activities[activity_key]["path"]

                        test_input = activities[activity_key].get("test_input", "")

                        # ✅ DEBUG: confirm server sees the stored test input

                        print("✅ [DEBUG submit_activity] activity_key:", activity_key)

                        print("✅ [DEBUG submit_activity] test_input repr:", repr(test_input))

                        print("✅ [DEBUG submit_activity] test_input lines:", test_input.splitlines())

                        similarity = compare_output(

                            teacher_file_path,

                            student_code,

                            language,

                            test_input=test_input

                        )

                        similarity = float(similarity)

                        # ✅ Save score to DB (same as you already do)

                        conn = sqlite3.connect(os.path.join(os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "..")), 'school.db'))

                        c = conn.cursor()

                        c.execute("""

                            CREATE TABLE IF NOT EXISTS activity_scores (

                                id INTEGER PRIMARY KEY AUTOINCREMENT,

                                student_name TEXT,

                                activity_number INTEGER,

                                activity_name TEXT,

                                score REAL,

                                date_taken TEXT,

                                instructor TEXT,

                                section TEXT,

                                subject TEXT

                            )

                        """)

                        activity_number = list(activities).index(activity_key) + 1

                        date_taken = time.strftime("%Y-%m-%d")

                        c.execute("""

                            INSERT INTO activity_scores (

                                student_name, activity_number, activity_name,

                                score, date_taken, instructor, section, subject

                            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?)

                        """, (student_name, activity_number, activity_key, similarity, date_taken, instructor, section,
                              subject))

                        conn.commit()

                        conn.close()

                        send_json_response(client_socket, {

                            "action": "activity_result",

                            "ok": True,

                            "similarity": similarity

                        })


                    except Exception as e:

                        send_json_response(client_socket, {

                            "action": "activity_result",

                            "ok": False,

                            "error": str(e),

                            "similarity": 0.0

                        })

                elif action == "get_activity_code":

                    activity_key = request.get("activity_key")

                    try:

                        if activity_key in activities:

                            with open(activities[activity_key]["path"], "r", encoding="utf-8", errors="ignore") as f:

                                instructor_code = f.read()

                            send_json_response(client_socket, {"ok": True, "code": instructor_code})

                        else:

                            send_json_response(client_socket, {"ok": False, "error": "Activity not found", "code": ""})

                    except Exception as e:

                        send_json_response(client_socket, {"ok": False, "error": str(e), "code": ""})
                elif action == "submit_quiz_score":
                    student_name = request.get("student_name")
                    quiz_number = int(request.get("quiz_number", 0) or 0)
                    score = float(request.get("score", 0) or 0)
                    section = request.get("section", "")
                    subject = request.get("subject", "")
                    instructor = request.get("instructor", "")
                    date_taken = request.get("date_taken", "")

                    try:
                        conn = sqlite3.connect(os.path.join(os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "..")), 'school.db'))
                        c = conn.cursor()

                        # ✅ correct table schema (includes instructor)
                        c.execute("""
                                  CREATE TABLE IF NOT EXISTS quiz_scores
                                  (
                                      id
                                      INTEGER
                                      PRIMARY
                                      KEY
                                      AUTOINCREMENT,
                                      student_name
                                      TEXT
                                      NOT
                                      NULL,
                                      instructor
                                      TEXT
                                      NOT
                                      NULL,
                                      section
                                      TEXT
                                      NOT
                                      NULL,
                                      subject
                                      TEXT
                                      NOT
                                      NULL,
                                      quiz_number
                                      INTEGER
                                      NOT
                                      NULL,
                                      score
                                      REAL
                                      NOT
                                      NULL,
                                      date_taken
                                      TEXT
                                      NOT
                                      NULL
                                  )
                                  """)

                        # ✅ optional: block duplicates at DB level
                        c.execute("""
                                  SELECT 1
                                  FROM quiz_scores
                                  WHERE student_name = ?
                                    AND instructor = ?
                                    AND section =?
                                    AND subject=?
                                    AND quiz_number=?
                                      LIMIT 1
                                  """, (student_name, instructor, section, subject, quiz_number))
                        if c.fetchone():
                            conn.close()
                            client_socket.sendall(
                                json.dumps({"status": "ok", "message": "Already recorded"}).encode("utf-8"))
                            continue

                        # ✅ insert
                        c.execute("""
                                  INSERT INTO quiz_scores (student_name, instructor, section, subject, quiz_number,
                                                           score, date_taken)
                                  VALUES (?, ?, ?, ?, ?, ?, ?)
                                  """, (student_name, instructor, section, subject, quiz_number, score, date_taken))

                        conn.commit()
                        conn.close()

                        client_socket.sendall(json.dumps({"status": "ok", "message": "Score saved"}).encode("utf-8"))
                        print("✅ Quiz score saved to school.db (server).")

                    except Exception as e:
                        print("❌ submit_quiz_score error:", e)
                        client_socket.sendall(json.dumps({"status": "error", "message": str(e)}).encode("utf-8"))



                    client_socket.send(json.dumps(response).encode())
                elif action == "get_quizzes":
                    try:
                        # ✅ Normalize / strip inputs
                        instructor = str(request.get("instructor", "")).strip()
                        section = str(request.get("section", "")).strip()
                        subject = str(request.get("subject", "")).strip()

                        print("📥 get_quizzes request:",
                              "instructor=", repr(instructor),
                              "section=", repr(section),
                              "subject=", repr(subject))

                        # ✅ If missing fields, DO NOT silently return ok:[]
                        if (not instructor) or (not section) or (not subject) or (instructor.upper() == "N/A"):
                            client_socket.sendall(json.dumps({
                                "status": "error",
                                "message": "Missing/invalid instructor/section/subject",
                                "quizzes": []
                            }).encode("utf-8"))
                            continue

                        base_dir = get_quiz_base_dir(instructor, section, subject)
                        quiz_list = []

                        if os.path.exists(base_dir):
                            for filename in os.listdir(base_dir):
                                if filename.endswith(".json"):
                                    path = os.path.join(base_dir, filename)
                                    try:
                                        with open(path, "r", encoding="utf-8") as f:
                                            data = json.load(f)

                                        quiz_list.append({
                                            "quiz_number": data.get("quiz_number"),
                                            "subject": data.get("subject"),
                                            "status": data.get("status", "Closed"),
                                            "timestamp": data.get("timestamp", "Unknown Date"),
                                            "quiz_title": data.get("quiz_title", "Untitled"),
                                            "time_limit": data.get("time_limit", 0),
                                            "instructor": data.get("instructor"),
                                            "section": data.get("section"),
                                        })
                                    except Exception as e:
                                        print(f"⚠️ Skipping quiz file {path}: {e}")
                        else:
                            print("📁 base_dir not found:", base_dir)

                        quiz_list.sort(key=lambda x: int(x.get("quiz_number", 999999) or 999999))

                        client_socket.sendall(json.dumps({
                            "status": "ok",
                            "quizzes": quiz_list
                        }).encode("utf-8"))

                        print(f"📤 Sent {len(quiz_list)} quizzes from:", base_dir)

                    except Exception as e:
                        print(f"❌ get_quizzes error: {e}")
                        client_socket.sendall(json.dumps({
                            "status": "error",
                            "message": str(e),
                            "quizzes": []
                        }).encode("utf-8"))

                elif action == "get_quiz_file":

                    try:
                        quiz_number = int(request.get("quiz_number", 0))
                        instructor = request.get("instructor")
                        section = request.get("section")
                        subject = request.get("subject")

                        if quiz_number <= 0:
                            client_socket.sendall(
                                json.dumps({"status": "error", "message": "Invalid quiz number."}).encode("utf-8"))
                            continue

                        if not instructor or not section or not subject or str(instructor).strip() == "N/A":
                            client_socket.sendall(json.dumps(
                                {"status": "error", "message": "Missing instructor/section/subject."}).encode("utf-8"))
                            continue

                        base_dir = get_quiz_base_dir(instructor, section, subject)
                        quiz_file_path = os.path.join(base_dir, f"quiz_{quiz_number}.json")

                        if not os.path.exists(quiz_file_path):
                            client_socket.sendall(json.dumps({
                                "status": "error",
                                "message": f"No quiz found for Quiz {quiz_number} under {instructor} | {section} | {subject}"
                            }).encode("utf-8"))
                            continue
                        with open(quiz_file_path, "r", encoding="utf-8") as f:
                            quiz_info = json.load(f)
                        client_socket.sendall(json.dumps({"status": "ok", "quiz_info": quiz_info}).encode("utf-8"))
                    except Exception as e:
                        client_socket.sendall(
                            json.dumps({"status": "error", "message": f"Failed to read quiz: {e}"}).encode("utf-8"))


                elif action == "check_quiz_taken":

                    try:

                        ensure_quiz_scores_schema()

                        student = str(request.get("student_name", "")).strip()

                        instructor = str(request.get("instructor", "")).strip()

                        section = str(request.get("section", "")).strip()

                        subject = str(request.get("subject", "")).strip()

                        quiz_number = int(request.get("quiz_number", 0) or 0)

                        if not student or not instructor or not section or not subject or quiz_number <= 0:
                            client_socket.sendall(
                                json.dumps({"taken": False, "error": "missing fields"}).encode("utf-8"))

                            continue

                        conn = sqlite3.connect(os.path.join(os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "..")), 'school.db'))

                        c = conn.cursor()

                        c.execute("""

                                  SELECT 1

                                  FROM quiz_scores

                                  WHERE student_name = ?

                                    AND instructor = ?

                                    AND section = ?

                                    AND subject = ?

                                    AND quiz_number = ?

                                      LIMIT 1

                                  """, (student, instructor, section, subject, quiz_number))

                        taken = c.fetchone() is not None

                        conn.close()

                        client_socket.sendall(json.dumps({"taken": taken}).encode("utf-8"))


                    except Exception as e:

                        print(f"❌ check_quiz_taken error: {e}")

                        client_socket.sendall(json.dumps({"taken": False, "error": str(e)}).encode("utf-8"))


                # ✅ Handle Activity Submission

                elif action == "get_attendance_logs":
                    req_student = str(request.get("student_name", "")).strip()
                    req_section = str(request.get("section", "")).strip()
                    req_subject = str(request.get("subject", "")).strip()

                    try:
                        conn = sqlite3.connect(os.path.join(os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "..")), 'school.db'))
                        c = conn.cursor()

                        # Ensure table exists in case it hasn't been created yet
                        c.execute('''
                            CREATE TABLE IF NOT EXISTS attendance (
                                id INTEGER PRIMARY KEY AUTOINCREMENT,
                                date TEXT, student_name TEXT, section TEXT,
                                time_in TEXT, instructor TEXT, subject TEXT,
                                time_out TEXT
                            )
                        ''')
                        
                        try:
                            c.execute("ALTER TABLE attendance ADD COLUMN time_out TEXT")
                        except sqlite3.OperationalError:
                            pass

                        c.execute("""
                            SELECT date, time_in, time_out, instructor, subject, section
                            FROM attendance
                            WHERE student_name = ?
                            ORDER BY date DESC, time_in DESC
                        """, (req_student,))

                        rows = c.fetchall()
                        conn.close()

                        logs = [
                            {
                                "date":       row[0],
                                "time_in":    row[1],
                                "time_out":   row[2] if row[2] else "",
                                "instructor": row[3],
                                "subject":    row[4],
                                "section":    row[5]
                            }
                            for row in rows
                        ]

                        print(f"📤 Sending {len(logs)} attendance log(s) for {req_student}")
                        client_socket.sendall(json.dumps({
                            "status": "ok",
                            "logs": logs
                        }).encode("utf-8"))

                    except Exception as e:
                        print(f"❌ get_attendance_logs error: {e}")
                        client_socket.sendall(json.dumps({
                            "status": "error",
                            "message": str(e),
                            "logs": []
                        }).encode("utf-8"))

                else:
                    print(f"❌ Unrecognized action: {action}")
                    client_socket.sendall(json.dumps({"message": "Unrecognized command."}).encode('utf-8'))



    except Exception as e:
        print(f"❌ Error handling client: {e}")



    finally:
        try:
            if client_socket in client_sockets:
                client_sockets.remove(client_socket)
        except:
            pass
        with clients_lock:
            cid = client_id_by_sock.pop(client_socket, None)
            if cid is not None:
                clients_by_id.pop(cid, None)
                pc_timers.pop(cid, None)
        try:
            client_socket.close()
        except:
            pass
        print("❌ Client disconnected.")



ACTIVITIES_FILE = "../../activities.json"
activities = {}
submissions = {}



def save_activities():
    try:
        data = {"activities": activities, "submissions": submissions}
        with open(ACTIVITIES_FILE, "w") as file:
            json.dump(data, file)
        print("Activities and submissions saved successfully.")
    except Exception as e:
        print(f"Error saving activities: {e}")

def load_activities():
    global activities, submissions
    if os.path.exists(ACTIVITIES_FILE):
        try:
            with open(ACTIVITIES_FILE, "r") as file:
                data = json.load(file)
                activities = data.get("activities", {})
                submissions = data.get("submissions", {})
                print("Activities and submissions loaded successfully.")
        except Exception as e:
            print(f"Error loading activities: {e}")

def get_saved_subjects():
    conn = sqlite3.connect(os.path.join(os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "..")), 'teacher_management.db'))
    c = conn.cursor()
    c.execute("SELECT DISTINCT subject_name FROM subjects")
    subjects = [row[0] for row in c.fetchall()]
    conn.close()
    return subjects
def remove_subject(subject):
    conn = sqlite3.connect(os.path.join(os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "..")), 'teacher_management.db'))
    c = conn.cursor()
    c.execute("DELETE FROM subjects WHERE subject_name = ?", (subject,))
    conn.commit()
    conn.close()
    messagebox.showinfo("Success", f"Subject '{subject}' removed successfully!")
def evaluate_java_code(student_code, expected_output_path):
    try:
        with tempfile.TemporaryDirectory() as tmpdir:
            java_file = os.path.join(tmpdir, "Main.java")
            with open(java_file, "w") as f:
                f.write(student_code)
            compile_result = subprocess.run(["javac", "Main.java"], cwd=tmpdir, capture_output=True, text=True)
            if compile_result.returncode != 0:
                print("❌ Java compilation failed:", compile_result.stderr)
                return 0.0
            run_result = subprocess.run(["java", "Main"], cwd=tmpdir, capture_output=True, text=True, timeout=5)
            student_output = run_result.stdout.strip()
            if expected_output_path.endswith(".py"):
                expected_output = subprocess.run(["python", expected_output_path], capture_output=True,
                                                 text=True).stdout.strip()
            elif expected_output_path.endswith(".java"):
                with open(expected_output_path, "r") as f:
                    instructor_java_code = f.read()
                expected_output = run_java_code_with_input(instructor_java_code, "")
            else:
                with open(expected_output_path, "r") as f:
                    expected_output = f.read().strip()

            return 100.0 if student_output == expected_output else 0.0
    except subprocess.TimeoutExpired:
        print("⏰ Java code execution timed out.")
        return 0.0
    except Exception as e:
        print(f"⚠️ Java eval error: {e}")
        return 0.0
def upload_activity(teacher, section, subject):
    def submit_details():
        activity_name = name_entry.get().strip()
        activity_number = number_entry.get().strip()
        selected_subject = subject_var.get().strip()
        instruction_text = instructions_textbox.get("1.0", tk.END).strip()
        test_input = test_input_textbox.get("1.0", tk.END).strip()   # ✅ now exists
        time_limit = time_entry.get().strip()

        if not activity_name or not activity_number or not selected_subject:
            messagebox.showerror("Error", "Fill in all required fields.")
            return

        try:
            selected_language = (language_var.get() or "").strip()
            if selected_language == "Python":
                filetypes = [("Python Files", "*.py")]
            elif selected_language == "Java":
                filetypes = [("Java Files", "*.java")]
            else:
                messagebox.showerror("Error", "Please select a programming language.")
                return

            activity_path = filedialog.askopenfilename(
                title="Select Activity File", filetypes=filetypes
            )
            if not activity_path:
                messagebox.showerror("Error", "Activity file not selected.")
                return

            instruction_image = filedialog.askopenfilename(
                title="Optional: Upload Instruction Image",
                filetypes=[("Image Files", "*.png;*.jpg;*.jpeg")]
            )

            activity_key = f"{activity_name} (#{activity_number})"
            print(f"[DEBUG] Activity Key Format: {activity_key}")

            if activity_key in activities:
                messagebox.showwarning("Duplicate", f"Activity '{activity_key}' already exists.")
                return

            activities[activity_key] = {
                "path": activity_path,
                "subject": selected_subject,
                "instructor": teacher,
                "section": section,
                "accepting": True,
                "date_uploaded": datetime.now().strftime("%Y-%m-%d %H:%M"),

                # ✅ NEW: store test input for input()-based activities
                "test_input": test_input,  # multiline allowed

                "instruction_text": instruction_text,
                "instruction_image": instruction_image or "",
                "language": selected_language,
                "time_limit": int(time_limit) if time_limit.isdigit() else None
            }

            submissions[activity_key] = {}
            save_activities()
            update_activity_list()
            messagebox.showinfo("Success", f"Uploaded: {activity_key}")
            upload_window.destroy()

        except Exception as e:
            messagebox.showerror("Error", str(e))

    upload_window = tk.Toplevel()
    upload_window.title("Upload Activity")
    upload_window.geometry("420x650")  # ✅ taller to fit test input
    upload_window.configure(bg="#D1BAFF")

    tk.Label(upload_window, text="Activity Name:", bg="#D1BAFF", font=("Arial", 11)).pack(pady=4)
    name_entry = tk.Entry(upload_window, font=("Arial", 11))
    name_entry.pack()

    tk.Label(upload_window, text="Activity Number:", bg="#D1BAFF", font=("Arial", 11)).pack(pady=4)
    number_entry = tk.Entry(upload_window, font=("Arial", 11))
    number_entry.pack()

    tk.Label(upload_window, text="Time Limit (mins):", bg="#D1BAFF", font=("Arial", 11)).pack(pady=4)
    time_entry = tk.Entry(upload_window, font=("Arial", 11))
    time_entry.pack()

    tk.Label(upload_window, text="Test Input (for input()):", bg="#D1BAFF", font=("Arial", 11)).pack(pady=4)
    test_input_textbox = tk.Text(upload_window, height=4, width=40, font=("Arial", 10))
    test_input_textbox.pack()

    tk.Label(upload_window, text="Instruction (Text):", bg="#D1BAFF", font=("Arial", 11)).pack(pady=4)
    instructions_textbox = tk.Text(upload_window, height=4, width=40, font=("Arial", 10))
    instructions_textbox.pack()

    tk.Label(upload_window, text="Subject:", bg="#D1BAFF", font=("Arial", 11)).pack(pady=4)
    subject_var = tk.StringVar()
    subject_dropdown = ttk.Combobox(
        upload_window, textvariable=subject_var, values=[subject],
        font=("Arial", 11), state="readonly"
    )
    subject_dropdown.set(subject)
    subject_dropdown.pack()

    tk.Label(upload_window, text="Language:", bg="#D1BAFF", font=("Arial", 11)).pack(pady=4)
    language_var = tk.StringVar()
    language_dropdown = ttk.Combobox(
        upload_window, textvariable=language_var,
        values=["Python", "Java"], font=("Arial", 11), state="readonly"
    )
    language_dropdown.set("Python")
    language_dropdown.pack()

    tk.Button(
        upload_window, text="Upload Activity",
        font=("Arial", 11), bg="white", command=submit_details
    ).pack(pady=10)

def delete_activity():
    try:
        sel = activity_listbox.curselection()
        if not sel:
            messagebox.showerror("Error", "No activity selected.")
            return

        idx = sel[0]  # ✅ tuple -> int
        raw_item = activity_listbox.get(idx).strip()

        # ✅ prevent deleting the subject header row
        if raw_item.startswith("📚"):
            messagebox.showerror("Error", "Please select an activity, not a subject category.")
            return

        # remove UI markers
        selected_activity = re.sub(
            r"➜|\[#\d+\]|\[✔ Accepting\]|\[❌ Closed\]|\[\/ Accepting\]|\[\/ Closed\]",
            "",
            raw_item
        ).strip()

        print(f"[DELETE] raw_item='{raw_item}'")
        print(f"[DELETE] parsed selected_activity='{selected_activity}'")
        print(f"[DELETE] exists in activities? {selected_activity in activities}")
        print(f"[DELETE] exists in submissions? {selected_activity in submissions}")

        if not (selected_activity in activities):
            messagebox.showerror(
                "Error",
                f"Activity not found in dictionary.\n\nSelected: {selected_activity}"
            )
            return

        if not messagebox.askyesno("Confirm Delete", f"Delete '{selected_activity}'?"):
            return

        # ✅ delete activity + its submission bucket
        activities.pop(selected_activity, None)
        submissions.pop(selected_activity, None)

        # ✅ VERY IMPORTANT: persist to activities.json
        save_activities()

        # refresh list
        update_activity_list()

        messagebox.showinfo("Deleted", f"Activity '{selected_activity}' deleted successfully.")
        print(f"[DELETE] ✅ deleted '{selected_activity}' and saved.")

    except Exception as e:
        print(f"[DELETE] ❌ error: {e}")
        messagebox.showerror("Error", str(e))

def toggle_accepting_status():
    selected_activity = activity_listbox.get(tk.ACTIVE).strip()
    if selected_activity.startswith("📚"):
        messagebox.showerror("Error", "Please select an activity, not a subject category.")
        return

    selected_activity = re.sub(r"➜|\[#\d+\]|\[✔ Accepting\]|\[❌ Closed\]", "", selected_activity).strip()

    if selected_activity in activities:
        activities[selected_activity]["accepting"] = not activities[selected_activity]["accepting"]
        save_activities()
        update_activity_list()
        status = "Accepting submissions" if activities[selected_activity]["accepting"] else "Submissions closed"
        messagebox.showinfo("Status Updated", f"Activity '{selected_activity}' is now {status}.")
    else:
        messagebox.showerror("Error", "No activity selected or activity not found.")


def update_activity_list():
    activity_listbox.delete(0, tk.END)

    if not current_instructor or not current_section or not current_subject:
        print("⚠️ No instructor/section/subject set - Unable to update activities!")
        return

    subject_activities = {}
    load_activities()

    for activity, details in activities.items():
        if (details["instructor"] == current_instructor and
                details["section"] == current_section and
                details["subject"] == current_subject):

            subject = details.get("subject", "No Subject")
            if subject not in subject_activities:
                subject_activities[subject] = []
            subject_activities[subject].append(activity)
    for activity, details in activities.items():
        print(
            f"[DEBUG] Checking Activity: {activity} => Instr: {details['instructor']} | Sect: {details['section']} | Subj: {details['subject']}")

    print(f"[DEBUG] Current => Instr: {current_instructor} | Sect: {current_section} | Subj: {current_subject}")

    for subject, activity_list in subject_activities.items():
        activity_listbox.insert(tk.END, f"📚 {subject}")
        for activity in activity_list:
            details = activities[activity]
            status = "✔ Accepting" if details["accepting"] else "❌ Closed"
            date = details.get("date_uploaded", "No Date")

            activity_listbox.insert(
                tk.END,
                f"   ➜ {activity} (🕒 {date}) [{status}]"
            )

def show_activity_options(teacher, section, subject):
    global current_instructor, current_section, current_subject
    current_instructor = teacher
    current_section = section
    current_subject = subject
    clear_grid()
    style = ttk.Style()
    style.configure("Orange.TButton", font=("Helvetica", 12, "bold"), background="#FF8C00", foreground="white")
    style.map("Orange.TButton",
              background=[("active", "#E07B00")],
              foreground=[("active", "white")])

    title_label = ttk.Label(
        window,
        text=f"Activity Management Please Upload Your Activity\nProfessor: {teacher} | Section: {section} | Subject: {subject}",
        font=("Helvetica", 18, "bold"), background="white", foreground="#8A2BE2", anchor="center"
    )
    title_label.place(relx=0.5, rely=0.07, anchor="center")

    main_frame = ttk.Frame(window, bootstyle="light", padding=20)
    main_frame.place(relx=0.5, rely=0.5, anchor="center", width=1150, height=790)

    listbox_frame = ttk.Frame(main_frame)
    listbox_frame.pack(pady=10, fill=tk.BOTH, expand=True)

    global activity_listbox
    activity_listbox = tk.Listbox(listbox_frame,font=("Helvetica", 20), height=10, width=50)
    activity_listbox.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

    scrollbar = ttk.Scrollbar(listbox_frame, orient=tk.VERTICAL, command=activity_listbox.yview)
    scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
    activity_listbox.config(yscrollcommand=scrollbar.set)
    button_frame = ttk.Frame(main_frame)
    button_frame.pack(pady=10, fill=tk.X)

    button_frame.grid_columnconfigure(0, weight=1)
    button_frame.grid_columnconfigure(1, weight=1)
    button_frame.grid_columnconfigure(2, weight=1)
    upload_button = ttk.Button(button_frame, text="Upload Activity", style="Orange.TButton",
                               command=lambda: upload_activity(teacher, section, subject))
    upload_button.grid(row=0, column=0, padx=5, pady=5, sticky="ew")
    delete_button = ttk.Button(button_frame, text="Delete Activity", style="Orange.TButton", command=delete_activity)
    delete_button.grid(row=0, column=1, padx=5, pady=5, sticky="ew")
    preview_button = ttk.Button(button_frame, text="Preview Instructions", style="Orange.TButton",
                                command=preview_instructions)
    preview_button.grid(row=0, column=2, padx=5, pady=5, sticky="ew")
    toggle_button = ttk.Button(button_frame, text="Toggle Accepting Status", style="Orange.TButton",
                               command=toggle_accepting_status)
    toggle_button.grid(row=1, column=0, columnspan=3, padx=10, pady=5, sticky="ew")
    back_button = ttk.Button(main_frame, text="Back", style="Orange.TButton",
                             command=lambda: show_teacher_options(teacher, section, subject))
    back_button.pack(pady=10)
    button_frame.grid_columnconfigure(0, weight=1)
    button_frame.grid_columnconfigure(1, weight=1)

    update_activity_list()

def preview_instructions():
    selected_activity = activity_listbox.get(tk.ACTIVE).strip()
    selected_activity = re.sub(r"➜|\[#\d+\]|\[✔ Accepting\]|\[❌ Closed\]", "", selected_activity).strip()

    if selected_activity in activities:
        instr = activities[selected_activity]
        instr_text = instr.get("instruction_text", "")
        instr_img_path = instr.get("instruction_image", "")

        preview_win = tk.Toplevel()
        preview_win.title("Edit Activity Instructions")
        preview_win.geometry("550x700")

        tk.Label(preview_win, text="Edit Instructions:", font=("Arial", 12, "bold")).pack(pady=5)
        text_box = tk.Text(preview_win, height=6, wrap=tk.WORD, font=("Arial", 10))
        text_box.insert(tk.END, instr_text)
        text_box.pack(pady=5)

        image_frame = ttk.Frame(preview_win)
        image_frame.pack(pady=5)

        image_label = tk.Label(image_frame)
        image_label.pack()

        def load_image(path):
            if os.path.exists(path):
                from PIL import Image, ImageTk
                img = Image.open(path)
                img.thumbnail((400, 300))
                tk_img = ImageTk.PhotoImage(img)
                image_label.configure(image=tk_img)
                image_label.image = tk_img
            else:
                image_label.configure(text="No instruction image uploaded.", font=("Arial", 10, "italic"))

        load_image(instr_img_path)

        def browse_image():
            new_path = filedialog.askopenfilename(title="Select New Instruction Image",
                                                  filetypes=[("Image Files", "*.png;*.jpg;*.jpeg;*.gif")])
            if new_path:
                instr["instruction_image"] = new_path
                load_image(new_path)

        browse_button = ttk.Button(preview_win, text="Browse Image", command=browse_image)
        browse_button.pack(pady=5)

        def save_changes():
            new_text = text_box.get("1.0", tk.END).strip()
            new_image_path = instr.get("instruction_image", "")
            activities[selected_activity]["instruction_text"] = new_text
            activities[selected_activity]["instruction_image"] = new_image_path
            messagebox.showinfo("Success", "Instructions updated successfully.")
            preview_win.destroy()

        ttk.Button(preview_win, text="Save Changes", style="Orange.TButton", command=save_changes).pack(pady=10)

    else:
        messagebox.showerror("Error", "No valid activity selected.")
import subprocess, tempfile, os, difflib
def execute_code(code, input_data=""):
    try:
        process = subprocess.run(
            ["python", "-c", code],
            capture_output=True,
            text=True,
            timeout=5,
            input=input_data if input_data is not None else ""
        )
        return process.stdout.strip() if process.returncode == 0 else process.stderr.strip()
    except subprocess.TimeoutExpired:
        return "Execution Timed Out ⚠️"
    except Exception as e:
        return f"Execution Error: {str(e)}"


def run_java_code_with_input(code, stdin_text=""):
    java_path = r"C:\Program Files\Java\jdk-24\bin\java.exe"
    javac_path = r"C:\Program Files\Java\jdk-24\bin\javac.exe"

    with tempfile.TemporaryDirectory() as tempdir:
        file_path = os.path.join(tempdir, "Main.java")
        with open(file_path, "w", encoding="utf-8", errors="ignore") as f:
            f.write(code)

        compile_result = subprocess.run(
            [javac_path, file_path],
            capture_output=True, text=True, timeout=8
        )
        if compile_result.returncode != 0:
            return compile_result.stderr.strip() or "Java compilation failed."

        run_result = subprocess.run(
            [java_path, "-cp", tempdir, "Main"],
            capture_output=True, text=True, timeout=8,
            input=stdin_text if stdin_text is not None else ""
        )
        out = run_result.stdout.strip()
        err = run_result.stderr.strip()
        return out if run_result.returncode == 0 else (err or out)


import difflib
import re

def _normalize_output(s: str) -> str:
    """Make output comparison more forgiving."""
    if s is None:
        return ""
    s = s.replace("\r\n", "\n").replace("\r", "\n")
    # Remove common prompt-ish patterns (optional but helps)
    # Example: "Enter a number:" lines won’t destroy similarity
    s = re.sub(r"(?i)enter\s+.*?:", "", s)
    # collapse extra whitespace
    s = "\n".join(line.strip() for line in s.splitlines() if line.strip() != "")
    return s.strip()

def _normalize_code_for_similarity(code: str) -> str:
    """Cheap code normalization for fallback similarity."""
    if not code:
        return ""
    # remove trailing spaces
    code = "\n".join(line.rstrip() for line in code.splitlines())
    # remove blank lines
    code = "\n".join(line for line in code.splitlines() if line.strip())
    return code.strip()

def compare_output(teacher_file_path, student_code, language="Python", test_input=""):
    """
    Primary: run both -> compare output similarity
    Fallback: if input-based causes EOF / runtime error -> compare CODE similarity instead
    """
    try:
        with open(teacher_file_path, "r", encoding="utf-8", errors="ignore") as f:
            instructor_code = f.read()

        # Run
        if language.lower() == "java":
            instructor_out = run_java_code_with_input(instructor_code, test_input)
            student_out = run_java_code_with_input(student_code, test_input)
        else:
            instructor_out = execute_code(instructor_code, input_data=test_input)
            student_out = execute_code(student_code, input_data=test_input)

        # Detect execution problems
        out_combined = (instructor_out or "") + "\n" + (student_out or "")
        if ("EOFError" in out_combined) or ("Traceback" in out_combined) or ("Error" in out_combined):
            # ✅ FALLBACK: grade by code similarity instead of hard 0
            a = _normalize_code_for_similarity(instructor_code)
            b = _normalize_code_for_similarity(student_code)
            sim = difflib.SequenceMatcher(None, a, b).ratio()
            return round(sim * 100, 2)

        # Normal output compare
        instructor_out = _normalize_output(instructor_out)
        student_out = _normalize_output(student_out)

        sim = difflib.SequenceMatcher(None, instructor_out, student_out).ratio()
        return round(sim * 100, 2)

    except Exception as e:
        print(f"⚠ compare_output() failed: {e}")
        # last-resort fallback: code similarity
        try:
            with open(teacher_file_path, "r", encoding="utf-8", errors="ignore") as f:
                instructor_code = f.read()
            a = _normalize_code_for_similarity(instructor_code)
            b = _normalize_code_for_similarity(student_code)
            sim = difflib.SequenceMatcher(None, a, b).ratio()
            return round(sim * 100, 2)
        except:
            return 0.0


import sqlite3

def ensure_quiz_scores_schema():
    """Make sure school.db.quiz_scores has instructor/section/subject columns (safe migration)."""
    conn = sqlite3.connect(os.path.join(os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "..")), 'school.db'))
    c = conn.cursor()

    # create table if it doesn't exist (latest schema)
    c.execute("""
        CREATE TABLE IF NOT EXISTS quiz_scores (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            student_name TEXT NOT NULL,
            instructor TEXT NOT NULL,
            section TEXT NOT NULL,
            subject TEXT NOT NULL,
            quiz_number INTEGER NOT NULL,
            score REAL NOT NULL,
            date_taken TEXT NOT NULL
        )
    """)

    # check existing columns
    c.execute("PRAGMA table_info(quiz_scores)")
    cols = {row[1] for row in c.fetchall()}

    # migrate older schemas
    if "instructor" not in cols:
        c.execute("ALTER TABLE quiz_scores ADD COLUMN instructor TEXT")
    if "section" not in cols:
        c.execute("ALTER TABLE quiz_scores ADD COLUMN section TEXT")
    if "subject" not in cols:
        c.execute("ALTER TABLE quiz_scores ADD COLUMN subject TEXT")

    # optional: backfill NULLs so your WHERE clauses behave
    c.execute("UPDATE quiz_scores SET instructor='N/A' WHERE instructor IS NULL")
    c.execute("UPDATE quiz_scores SET section='N/A' WHERE section IS NULL")
    c.execute("UPDATE quiz_scores SET subject='N/A' WHERE subject IS NULL")

    conn.commit()
    conn.close()

def start_discovery_service():
    """Starts a UDP listener to respond to client discovery broadcasts."""

    def discovery_handler():
        udp_socket = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
        udp_socket.setsockopt(socket.SOL_SOCKET, socket.SO_BROADCAST, 1)
        # Bind to 0.0.0.0 so we can receive broadcasts
        try:
            udp_socket.bind(('0.0.0.0', 9999))
            print("📡 Discovery Service Started on UDP 9999")
            while True:
                try:
                    data, addr = udp_socket.recvfrom(1024)
                    message = data.decode('utf-8').strip()
                    if message == "DISCOVER_SERVER":
                        # Respond to the client so they know our IP
                        # We send back "I_AM_SERVER"
                        udp_socket.sendto("I_AM_SERVER".encode('utf-8'), addr)
                except Exception as e:
                    print(f"⚠️ Discovery Error: {e}")
        except Exception as e:
            print(f"❌ Failed to bind Discovery Service: {e}")
        finally:
            udp_socket.close()

    threading.Thread(target=discovery_handler, daemon=True).start()


def start_server():
    global client_sockets, _next_client_id

    start_discovery_service()

    server_socket = socket.socket(socket.AF_INET, socket.SOCK_STREAM)

    # ✅ reuse address (helps after restart)
    server_socket.setsockopt(socket.SOL_SOCKET, socket.SO_REUSEADDR, 1)

    server_socket.bind(("0.0.0.0", 9999))

    # ✅ IMPORTANT: backlog bigger than 5 for comlab
    server_socket.listen(100)

    print("Server listening on port 9999 (0.0.0.0)")

    while True:
        client_socket, addr = server_socket.accept()
        print(f"Connection from {addr}")

        client_sockets.append(client_socket)

        with clients_lock:
            cid = _next_client_id
            _next_client_id += 1
            clients_by_id[cid] = {"sock": client_socket, "addr": addr, "pc": None}
            client_id_by_sock[client_socket] = cid

        threading.Thread(target=handle_client, args=(client_socket,), daemon=True).start()

def view_attendance():
    # Make sure we load the latest records from DB before viewing
    load_attendance_records()
    
    def display_records(records, parent_window):
        record_window = tk.Toplevel(parent_window)
        record_window.title(f"Attendance Records for {selected_date}")
        record_window.geometry("750x450")
        record_window.configure(bg='#D1BAFF')

        columns = ('Name', 'PC Number', 'Time In', 'Time Out', 'Section', 'Date', 'Subject')
        tree = ttk.Treeview(record_window, columns=columns, show='headings')
        for col in columns:
            tree.heading(col, text=col)
            tree.column(col, width=110)

        for record in records:
            tree.insert('', tk.END, values=(
                record['name'], record.get('pc_number', 'Unknown'), record['time_in'], record['time_out'], record['section'],
                record['date'], record['subject']
            ))

        tree.pack(fill=tk.BOTH, expand=True)

    selected_date = date_var.get()
    if selected_date in attendance_records:
        records = attendance_records[selected_date]
        display_records(records, window)
    else:
        messagebox.showinfo("Attendance Records", f"No records found for {selected_date}.")


def update_date_dropdown():
    dates = sorted(list(attendance_records.keys()), reverse=True)
    date_dropdown['values'] = dates
    if dates:
        date_var.set(dates[0])


def delete_attendance(teacher, section, subject):
    selected_date = date_var.get()
    if not selected_date:
        messagebox.showwarning("No Date Selected", "Please select a date to delete.")
        return

    if selected_date not in attendance_records:
        messagebox.showinfo("No Records", f"No attendance found for {selected_date}.")
        return

    confirm = messagebox.askyesno(
        "Confirm Deletion",
        f"Are you sure you want to delete all attendance records for {selected_date}?"
    )
    if not confirm:
        return

    # ✅ Correct table and columns
    conn = sqlite3.connect(os.path.join(os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "..")), 'school.db'))
    c = conn.cursor()
    c.execute('''
              DELETE
              FROM attendance
              WHERE date = ? AND instructor = ? AND section = ? AND subject = ?
              ''', (selected_date, teacher, section, subject))
    conn.commit()
    conn.close()

    # ✅ Remove from in-memory dict and refresh
    del attendance_records[selected_date]
    update_date_dropdown()
    refresh_attendance_log()

    messagebox.showinfo("Deleted", f"Attendance for {selected_date} has been deleted.")


def export_to_excel(teacher, section, subject):
    # Ensure latest data is loaded before exporting
    load_attendance_records()
    selected_date = date_var.get()

    if selected_date not in attendance_records or not attendance_records[selected_date]:
        messagebox.showinfo("Export to Excel", f"No records found for {selected_date}.")
        return

    try:
        import os
        import subprocess
        from datetime import datetime
        from openpyxl import load_workbook
        from openpyxl.styles import Font, Alignment

        # 📊 Convert to DataFrame
        df = pd.DataFrame(attendance_records[selected_date])

        # 🖥 Desktop path
        desktop = os.path.join(os.path.expanduser("~"), "Desktop")

        # 📁 Base folder
        base_folder = os.path.join(desktop, "Attendance_Reports")
        os.makedirs(base_folder, exist_ok=True)

        # 🕒 Date formatting
        date_str = datetime.strptime(selected_date, "%Y-%m-%d").strftime("%Y_%m_%d")

        # 📄 File name
        filename = f"{teacher}_{section}_{subject}_{date_str}.xlsx"
        file_path = os.path.join(base_folder, filename)

        # ✍ Write Excel with space for headers
        with pd.ExcelWriter(file_path, engine="openpyxl") as writer:
            df.to_excel(writer, index=False, startrow=6, sheet_name="Attendance")

        # 🎨 Styling
        wb = load_workbook(file_path)
        ws = wb.active

        ws.merge_cells("A1:E1")
        ws.merge_cells("A2:E2")
        ws.merge_cells("A3:E3")
        ws.merge_cells("A4:E4")

        ws["A1"] = "ATTENDANCE REPORT"
        ws["A2"] = f"Instructor: {teacher}"
        ws["A3"] = f"Section: {section} | Subject: {subject}"
        ws["A4"] = f"Date: {selected_date}"

        ws["A1"].font = Font(size=16, bold=True)
        ws["A2"].font = Font(size=12, bold=True)
        ws["A3"].font = Font(size=12)
        ws["A4"].font = Font(size=11, italic=True)

        for r in range(1, 5):
            ws[f"A{r}"].alignment = Alignment(horizontal="center")

        wb.save(file_path)

        # 📂 Auto-open folder
        try:
            if os.name == "nt":  # Windows
                subprocess.Popen(f'explorer "{base_folder}"')
            else:
                subprocess.Popen(["open", base_folder])
        except Exception as e:
            print("⚠ Could not open folder:", e)

        # ✅ Confirmation
        messagebox.showinfo(
            "Export Successful",
            f"Attendance exported successfully!\n\n"
            f"Instructor: {teacher}\n"
            f"Section: {section}\n"
            f"Subject: {subject}\n"
            f"Date: {selected_date}\n\n"
            f"File:\n{filename}"
        )

    except Exception as e:
        messagebox.showerror("Export Failed", str(e))


def toggle_rfid_scanner(mode, status_label):
    global rfid_scanner_mode
    rfid_scanner_mode = mode

    current_date = datetime.now().strftime('%Y-%m-%d')
    if current_date not in attendance_records:
        attendance_records[current_date] = []

    if mode == "TIME_IN":
        status_text = "Status: TIME IN MODE"
        status_color = "blue"
    elif mode == "TIME_OUT":
        status_text = "Status: TIME OUT MODE"
        status_color = "orange"
    else:
        status_text = "Status: OFF"
        status_color = "red"

    status_label.config(
        text=status_text,
        foreground=status_color
    )

    print(f"RFID scanner mode set to {mode} on {current_date}")
    update_date_dropdown()


def ensure_quiz_score_columns():
    conn = sqlite3.connect(os.path.join(os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "..")), 'school.db'))

    c = conn.cursor()
    c.execute("PRAGMA table_info(quiz_scores)")
    columns = [col[1] for col in c.fetchall()]
    if "section" not in columns:
        c.execute("ALTER TABLE quiz_scores ADD COLUMN section TEXT")
    if "subject" not in columns:
        c.execute("ALTER TABLE quiz_scores ADD COLUMN subject TEXT")
    conn.commit()
    conn.close()


def normalize_all_quiz_scores():
    conn = sqlite3.connect(os.path.join(os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "..")), 'student_quiz_scores.db'))
    c = conn.cursor()
    c.execute("""
              UPDATE quiz_scores
              SET section      = LOWER(TRIM(section)),
                  subject      = LOWER(TRIM(subject)),
                  student_name = TRIM(student_name)
              WHERE section IS NOT NULL AND subject IS NOT NULL
              """)
    conn.commit()
    conn.close()
    print("✅ Quiz scores normalized.")


def view_recored_scores_quiz(teacher, section, subject):
    clear_grid()
    ensure_quiz_score_columns()
    normalize_all_quiz_scores()
    show_all_quiz_scores()
    scores = fetch_quiz_scores(section, subject)
    print("🧪 From fetch_quiz_scores():")
    import os
    print("📂 Using DB:", os.path.abspath("../../school.db"))

    for s in scores:
        print("→", s)

    def list_all_quiz_score_names():
        conn = sqlite3.connect(os.path.join(os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "..")), 'school.db'))  # ✅ not student_quiz_scores.db
        c = conn.cursor()
        c.execute("SELECT DISTINCT student_name FROM quiz_scores")
        names = c.fetchall()
        conn.close()
        print("📋 All student_name entries in quiz_scores (school.db):")
        for name in names:
            print(f"→ '{name[0]}'")

    list_all_quiz_score_names()

    main_frame = tk.Frame(window, bg="white", bd=2, relief="solid")
    main_frame.place(relx=0.5, rely=0.5, anchor="center", relwidth=0.9, relheight=0.9)

    title_label = ttk.Label(main_frame, text=f"Quiz Scores - {section} - {subject}  |  Teacher: {teacher}",
                            font=("Helvetica", 20, "bold"), foreground="#5D3FD3", background="white")
    title_label.pack(pady=15)

    scores = fetch_quiz_scores(section, subject)
    quiz_numbers = sorted(set(score[1] for score in scores))
    quiz_numbers_display = ["Show All"] + [f"Quiz {qn}" for qn in quiz_numbers]

    selected_quiz = tk.StringVar()
    filter_dropdown = ttk.Combobox(main_frame, textvariable=selected_quiz,
                                   values=quiz_numbers_display, state="readonly", width=20)
    filter_dropdown.set("Show All")
    filter_dropdown.pack()

    # Treeview Frame
    tree_container = tk.Frame(main_frame, bg="white")
    tree_container.pack(fill="both", expand=True, padx=20, pady=(10, 5))

    columns = ("Student Name", "Quiz Number", "Score", "Date Taken")
    tree = ttk.Treeview(tree_container, columns=columns, show="headings")

    for col in columns:
        width = 300 if col == "Student Name" else 220
        tree.heading(col, text=col, anchor="center", command=lambda _col=col: sort_column(_col, False))
        tree.column(col, width=width, anchor="center")

    scrollbar = ttk.Scrollbar(tree_container, orient="vertical", command=tree.yview)
    tree.configure(yscrollcommand=scrollbar.set)
    scrollbar.pack(side="right", fill="y")
    tree.pack(side="left", fill="both", expand=True)

    style = ttk.Style()
    style.configure("Treeview", rowheight=35, font=("Helvetica", 12))
    style.configure("Treeview.Heading", font=("Helvetica", 13, "bold"))
    tree.tag_configure("evenrow", background="#f7f7ff")
    tree.tag_configure("oddrow", background="#ffffff")

    def sort_column(col, reverse):
        data = [(tree.set(child, col), child) for child in tree.get_children('')]
        if col in ("Quiz Number", "Score"):
            data.sort(key=lambda t: int(t[0]) if str(t[0]).isdigit() else 0, reverse=reverse)
        else:
            data.sort(reverse=reverse)
        for index, (val, child) in enumerate(data):
            tree.move(child, '', index)
        tree.heading(col, command=lambda: sort_column(col, not reverse))

    def refresh_table(event=None):
        selected = selected_quiz.get()
        tree.delete(*tree.get_children())
        filtered = scores if selected == "Show All" else [
            s for s in scores if str(s[1]) == selected.split(" ")[1]
        ]
        for i, score in enumerate(filtered):
            row_tags = ("evenrow" if i % 2 == 0 else "oddrow",)
            tree.insert("", "end", values=score, tags=row_tags)

    filter_dropdown.bind("<<ComboboxSelected>>", refresh_table)
    refresh_table()

    # Buttons
    button_frame = ttk.Frame(main_frame, style="TFrame")
    button_frame.pack(pady=10)

    ttk.Button(button_frame, text="🗑 Delete Selected", bootstyle="danger",
               command=lambda: delete_selected_score(tree, section, subject)).grid(row=0, column=0, padx=15, ipadx=12)

    ttk.Button(button_frame, text="📥 Export to Excel", bootstyle="success",
               command=lambda: export_scores_to_excel(scores, section, subject)).grid(row=0, column=1, padx=15,
                                                                                      ipadx=12)

    ttk.Button(button_frame, text="🔙 Back", bootstyle="secondary",
               command=lambda: [main_frame.destroy(), show_teacher_options(teacher, section, subject)]).grid(row=0,
                                                                                                             column=2,
                                                                                                             padx=15,
                                                                                                             ipadx=12)


def show_all_quiz_scores():
    conn = sqlite3.connect(os.path.join(os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "..")), 'school.db'))  # ✅ use the same DB as the instructor/server
    c = conn.cursor()
    c.execute("SELECT student_name, quiz_number, section, subject, score, date_taken FROM quiz_scores")
    rows = c.fetchall()
    conn.close()

    print("📊 All Rows in school.db.quiz_scores:")
    for row in rows:
        print("→", row)


def delete_selected_score(tree, section, subject):
    selected_item = tree.selection()
    if not selected_item:
        messagebox.showerror("Error", "No score selected!")
        return

    student_name, quiz_number, score, date_taken = tree.item(selected_item, "values")

    print("🔍 DELETE DEBUG:")
    print("Name:", student_name)
    print("Quiz:", quiz_number)
    print("Section:", section)
    print("Subject:", subject)

    conn = sqlite3.connect(os.path.join(os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "..")), 'school.db'))  # ✅ Make sure this is correct
    c = conn.cursor()

    # Show rows that exactly match
    c.execute("""
              SELECT *
              FROM quiz_scores
              WHERE TRIM(student_name) = TRIM(?)
                AND quiz_number = ?
                AND TRIM(section) = TRIM(?)
                AND TRIM(subject) = TRIM(?)
              """, (student_name, quiz_number, section, subject))

    matches = c.fetchall()
    print("📋 Matching rows found:", matches)

    if not matches:
        messagebox.showerror("Error", "No matching score found.")
        return

    confirm = messagebox.askyesno("Confirm Delete", f"Delete score for {student_name}, Quiz {quiz_number}?")
    if not confirm:
        return

    c.execute("""
              DELETE
              FROM quiz_scores
              WHERE TRIM(student_name) = TRIM(?)
                AND quiz_number = ?
                AND TRIM(section) = TRIM(?)
                AND TRIM(subject) = TRIM(?)
              """, (student_name, quiz_number, section, subject))

    conn.commit()
    conn.close()
    messagebox.showinfo("Deleted", f"Score for {student_name}, Quiz {quiz_number} deleted.")
    view_recored_scores_quiz(current_instructor, section, subject)


def export_scores_to_excel(scores, section, subject):
    if not scores:
        messagebox.showerror("Error", "No quiz scores to export!")
        return

    import os
    import subprocess
    from datetime import datetime
    from openpyxl import load_workbook
    from openpyxl.styles import Font, Alignment

    # Create DataFrame
    df = pd.DataFrame(scores, columns=[
        "Student Name", "Quiz Number", "Score", "Date Taken"
    ])

    # 📁 Desktop path
    desktop = os.path.join(os.path.expanduser("~"), "Desktop")

    # 📁 Base folder
    base_folder = os.path.join(desktop, "quiz_scores_instructor")

    # 📁 Instructor folder
    instructor_folder = os.path.join(base_folder, current_instructor)
    os.makedirs(instructor_folder, exist_ok=True)

    # 🕒 Date info
    now = datetime.now()
    date_str = now.strftime("%Y_%m_%d")

    # 📄 File name format
    filename = f"{current_instructor}_{section}_{subject}_{date_str}.xlsx"
    file_path = os.path.join(instructor_folder, filename)

    # ✍ Write Excel (start data lower to make space for title)
    with pd.ExcelWriter(file_path, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, startrow=6, sheet_name="Quiz Scores")

    # 🧾 Open workbook for styling
    wb = load_workbook(file_path)
    ws = wb.active

    # 🏷 Header content
    ws.merge_cells("A1:D1")
    ws.merge_cells("A2:D2")
    ws.merge_cells("A3:D3")
    ws.merge_cells("A4:D4")

    ws["A1"] = "QUIZ SCORES REPORT"
    ws["A2"] = f"Teacher: {current_instructor}"
    ws["A3"] = f"Section: {section} | Subject: {subject}"
    ws["A4"] = f"Date Exported: {now.strftime('%B %d, %Y')}"

    # 🎨 Styling
    ws["A1"].font = Font(size=16, bold=True)
    ws["A2"].font = Font(size=12, bold=True)
    ws["A3"].font = Font(size=12)
    ws["A4"].font = Font(size=11, italic=True)

    for row in range(1, 5):
        ws[f"A{row}"].alignment = Alignment(horizontal="center")

    # 💾 Save styled file
    wb.save(file_path)

    # 📂 Open folder automatically
    try:
        if os.name == "nt":  # Windows
            subprocess.Popen(f'explorer "{instructor_folder}"')
        else:
            subprocess.Popen(["open", instructor_folder])
    except Exception as e:
        print("⚠ Could not open folder:", e)

    # ✅ Confirmation
    messagebox.showinfo(
        "Export Successful",
        f"Quiz scores exported successfully!\n\n"
        f"Teacher: {current_instructor}\n"
        f"Section: {section}\n"
        f"Subject: {subject}\n\n"
        f"File:\n{filename}"
    )

import ttkbootstrap as ttk
from datetime import datetime
import threading

attendance_records = {}
last_log_date = datetime.now().strftime("%Y-%m-%d")
attendance_listbox = None  # Ensure defined at top

def check_date_rollover():
    global last_log_date
    current_date = datetime.now().strftime("%Y-%m-%d")
    if current_date != last_log_date:
        attendance_records.clear()
        last_log_date = current_date
        print("🕛 New day detected — attendance log cleared.")
    window.after(60000, check_date_rollover)  # check every 60 seconds


def refresh_attendance_log():
    """Reload today's attendance directly from DB so the live log is always accurate."""
    if not attendance_listbox:
        return
    today = datetime.now().strftime('%Y-%m-%d')
    try:
        conn = sqlite3.connect(os.path.join(os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "..")), 'school.db'))
        c = conn.cursor()
        c.execute("""
            SELECT time_in, time_out, student_name, section
            FROM attendance
            WHERE date = ? AND instructor = ? AND section = ? AND subject = ?
            ORDER BY time_in ASC
        """, (today, current_instructor, current_section, current_subject))
        rows = c.fetchall()
        conn.close()

        try:
            attendance_listbox.delete(0, tk.END)
            for time_in, time_out, name, section in rows:
                out_str = time_out if time_out else "Still IN"
                line = f"IN: {time_in} | OUT: {out_str} | {today} - {name} ({section})"
                attendance_listbox.insert(tk.END, line)
        except tk.TclError:
            pass  # widget was destroyed (navigated away)
    except Exception as e:
        print(f"⚠️ refresh_attendance_log error: {e}")


def start_gui(teacher, section, subject):
    clear_grid()
    global date_var, date_dropdown, attendance_listbox

    window.grid_rowconfigure(0, weight=1)
    window.grid_rowconfigure(1, weight=1)
    window.grid_rowconfigure(2, weight=3)
    window.grid_rowconfigure(3, weight=1)
    window.grid_columnconfigure(0, weight=1)
    window.grid_columnconfigure(1, weight=3)
    window.grid_columnconfigure(2, weight=1)

    print(f"🎓 Logged in as: {teacher}")

    title_frame = ttk.Frame(window, padding=10)
    title_frame.grid(row=0, column=1, pady=(10, 5), padx=40, sticky=tk.EW)
    ttk.Label(
        title_frame, text="RFID Scanner Feature - Attendance",
        font=("Helvetica", 18, "bold"), foreground="#5D3FD3", anchor="center"
    ).pack(fill=tk.BOTH, expand=True)

    btn_frame = ttk.Frame(window, padding=10, bootstyle="secondary")
    btn_frame.grid(row=2, column=1, pady=10, padx=40, sticky=tk.NSEW)

    ttk.Label(
        btn_frame, text=f"👤 {teacher} | 🏫 {section} | 📚 {subject}",
        font=("Helvetica", 12, "bold"), foreground="black", background="#D3D3D3",
        padding=5, anchor="center"
    ).pack(fill=tk.X, pady=(0, 10))

    ttk.Button(
        btn_frame, text="Time In Mode",
        style="info.Outline.TButton", width=16,
        command=lambda: [toggle_rfid_scanner("TIME_IN", attendance_status_label), rfid_entry.focus_set()]
    ).pack(pady=(5, 3))

    ttk.Button(
        btn_frame, text="Time Out Mode",
        style="warning.Outline.TButton", width=16,
        command=lambda: [toggle_rfid_scanner("TIME_OUT", attendance_status_label), rfid_entry.focus_set()]
    ).pack(pady=(3, 3))

    # ✅ Show current state when page loads (persists across Back navigation)
    if rfid_scanner_mode == "TIME_IN":
        _scanner_status_text = "Status: TIME IN MODE"
        _scanner_status_color = "blue"
    elif rfid_scanner_mode == "TIME_OUT":
        _scanner_status_text = "Status: TIME OUT MODE"
        _scanner_status_color = "orange"
    else:
        _scanner_status_text = "Status: OFF"
        _scanner_status_color = "red"

    attendance_status_label = ttk.Label(
        btn_frame,
        text=_scanner_status_text,
        font=("Arial", 12, "bold"),
        foreground=_scanner_status_color
    )
    attendance_status_label.pack(pady=5)

    # ✅ Enlarged and Visible RFID Entry Field
    ttk.Label(btn_frame, text="Manual RFID Entry:", font=("Arial", 10, "bold")).pack(pady=(5, 2))
    rfid_entry = ttk.Entry(btn_frame, font=("Helvetica", 14), width=25)
    rfid_entry.pack(pady=(0, 10))
    
    def on_rfid_scan(event):
        code = rfid_entry.get().strip()
        if code:
            process_local_rfid(code)
            rfid_entry.delete(0, tk.END)
            # small delay before refresh
            window.after(500, refresh_attendance_log)

    rfid_entry.bind("<Return>", on_rfid_scan)
    rfid_entry.focus_set()

    # ✅ Single Live Attendance Log
    log_frame = ttk.LabelFrame(btn_frame, text="📋 Live Attendance Log", padding=5)
    log_frame.pack(pady=5, fill=tk.BOTH, expand=False)

    attendance_listbox = tk.Listbox(log_frame, height=8, font=("Arial", 11))
    attendance_listbox.pack(fill=tk.BOTH, expand=True)

    ttk.Button(
        btn_frame, text="Stop RFID Scanner",
        style="danger.Outline.TButton", width=16,
        command=lambda: toggle_rfid_scanner(False, attendance_status_label)
    ).pack(pady=(3, 10))

    ttk.Label(
        btn_frame, text="Select Date:",
        font=("Arial", 12, "bold"), foreground="black"
    ).pack(pady=(5, 3))

    date_var = tk.StringVar()
    date_dropdown = ttk.Combobox(
        btn_frame, textvariable=date_var,
        font=("Arial", 12), bootstyle="info", width=14
    )
    date_dropdown.pack(pady=(0, 10))

    ttk.Button(
        btn_frame, text="View Attendance",
        style="primary.Outline.TButton", width=16,
        command=view_attendance
    ).pack(pady=(5, 3))

    ttk.Button(
        btn_frame, text="Delete Attendance",
        style="danger.Outline.TButton", width=16,
        command=lambda: delete_attendance(teacher, section, subject)
    ).pack(pady=(3, 3))

    ttk.Button(
        btn_frame, text="Export to Excel",
        style="warning.Outline.TButton", width=16,
        command=lambda: export_to_excel(teacher, section, subject)
    ).pack(pady=(3, 15))

    ttk.Button(
        btn_frame, text="🔙 Back",
        style="danger.TButton", width=12,
        command=lambda: show_teacher_options(teacher, section, subject)
    ).pack(pady=15)

    load_attendance_records()
    update_date_dropdown()

    # ✅ Load live log
    refresh_attendance_log()

    # ✅ Auto-refresh live log every 3 seconds (polls DB so new scans appear immediately)
    def auto_refresh_live_log():
        if not attendance_listbox:
            return  # stop polling if we left the page
        try:
            if not attendance_listbox.winfo_exists():
                return
        except tk.TclError:
            return
        refresh_attendance_log()
        update_date_dropdown()
        window.after(3000, auto_refresh_live_log)

    window.after(3000, auto_refresh_live_log)

    # ✅ Begin date rollover check
    check_date_rollover()

    # ✅ Start server
    threading.Thread(target=start_server, daemon=True).start()

import tkinter as tk
from tkinter import ttk, messagebox
def _set_timer_for_pc(cid: int, action: str, minutes: int):
    """Store end time so UI can show remaining."""
    if minutes and minutes > 0:
        pc_timers[cid] = {"action": action.lower(), "end": time.time() + (minutes * 60)}
    else:
        pc_timers.pop(cid, None)


def send_command_to_one(cid: int, action: str, minutes: int = 0):
    payload = {"action": action.lower(), "timer": int(minutes or 0)}
    msg = json.dumps(payload).encode("utf-8")

    with clients_lock:
        info = clients_by_id.get(cid)
        if not info or not info.get("sock"):
            messagebox.showwarning("Not Found", f"PC-{cid:02d} not connected.")
            return
        sock = info["sock"]

    try:
        sock.sendall(msg)
        with clients_lock:
            _set_timer_for_pc(cid, action, int(minutes or 0))

        messagebox.showinfo("Sent", f"✅ Sent {action.upper()} ({minutes} min) to PC-{cid:02d}")
    except Exception as e:
        messagebox.showerror("Send Error", f"Failed sending to PC-{cid:02d}: {e}")


def send_command_to_many(cids, action: str, minutes: int = 0):
    payload = {"action": action.lower(), "timer": int(minutes or 0)}
    msg = json.dumps(payload).encode("utf-8")

    sent = 0
    dead = []

    with clients_lock:
        targets = [(cid, clients_by_id.get(cid)) for cid in cids]

    for cid, info in targets:
        sock = info.get("sock") if info else None
        if not sock:
            dead.append(cid)
            continue
        try:
            sock.sendall(msg)
            sent += 1
            with clients_lock:
                _set_timer_for_pc(cid, action, int(minutes or 0))
        except:
            dead.append(cid)

    if dead:
        print("⚠️ Failed to send to:", dead)

    messagebox.showinfo("Sent", f"✅ Sent to {sent} PC(s). Failed: {len(dead)}")

def send_command_to_all(action: str, minutes: int = 0):
    with clients_lock:
        all_ids = list(clients_by_id.keys())
    send_command_to_many(all_ids, action, minutes)

def show_timer_options(teacher, section, subject):
    clear_grid()

    timer_options = {
        "30 Minutes": 30,
        "1 Hour": 60,
        "2 Hours": 120,
        "3 Hours": 180
    }

    timer_frame = ttk.Frame(window, padding=20, bootstyle="dark")
    timer_frame.place(relx=0.5, rely=0.5, anchor=tk.CENTER, width=820, height=680)

    ttk.Label(
        timer_frame,
        text="Set Timer for Shutdown or Sleep",
        font=("Helvetica", 16, "bold"),
        foreground="#5D3FD3"
    ).grid(row=0, column=0, columnspan=2, pady=(5, 10))

    ttk.Label(timer_frame, text="Choose Action:", font=("Arial", 12, "bold")).grid(
        row=1, column=0, pady=5, sticky=tk.E
    )

    action_var = tk.StringVar(value="Shutdown")
    ttk.Combobox(
        timer_frame,
        textvariable=action_var,
        values=["Shutdown", "Sleep"],
        state="readonly",
        font=("Arial", 12),
        width=20
    ).grid(row=1, column=1, pady=5, padx=10, sticky=tk.W)

    ttk.Label(
        timer_frame,
        text="Connected PCs (CTRL/SHIFT to select multiple):",
        font=("Arial", 12, "bold")
    ).grid(row=2, column=0, columnspan=2, pady=(10, 5))

    pc_list = tk.Listbox(
        timer_frame,
        height=10,
        font=("Arial", 11),
        selectmode=tk.EXTENDED
    )
    pc_list.grid(row=3, column=0, columnspan=2, padx=10, pady=(0, 10), sticky="nsew")

    # listbox row index -> cid
    row_to_cid = []
    refresh_job = {"id": None}   # store after() id so we can cancel on Back
    refresh_paused = {"value": False}  # optional: pause refresh while user is selecting

    def fmt_remaining(cid: int) -> str:
        with clients_lock:
            t = pc_timers.get(cid)

        if not t:
            return "—"

        remaining = int(t["end"] - time.time())
        if remaining <= 0:
            return "0:00"

        mins = remaining // 60
        secs = remaining % 60
        if mins >= 60:
            h = mins // 60
            m = mins % 60
            return f"{h}h {m:02d}m"
        return f"{mins}:{secs:02d}"

    def get_selected_cids():
        sels = pc_list.curselection()
        return [row_to_cid[i] for i in sels if 0 <= i < len(row_to_cid)]

    # --- OPTIONAL: pause refresh while mouse is held down (helps selection feel stable) ---
    def _pause_refresh(_evt=None):
        refresh_paused["value"] = True

    def _resume_refresh(_evt=None):
        refresh_paused["value"] = False

    pc_list.bind("<ButtonPress-1>", _pause_refresh)
    pc_list.bind("<ButtonRelease-1>", _resume_refresh)

    def refresh_pc_list():
        # if user is currently selecting, skip this tick
        if refresh_paused["value"]:
            refresh_job["id"] = window.after(300, refresh_pc_list)
            return

        # ✅ remember selected CIDs BEFORE rebuilding
        selected_cids = set(get_selected_cids())

        pc_list.delete(0, tk.END)
        row_to_cid.clear()

        with clients_lock:
            ids = sorted(clients_by_id.keys())
            snapshot = []
            for cid in ids:
                info = clients_by_id[cid]
                pc = info.get("pc") or "Unknown-PC"
                addr = info.get("addr")
                ip = addr[0] if addr else "?"
                snapshot.append((cid, pc, ip))

        # rebuild list
        for cid, pc, ip in snapshot:
            rem = fmt_remaining(cid)
            pc_list.insert(tk.END, f"PC-{cid:02d} | {pc} | {ip} | Remaining: {rem}")
            row_to_cid.append(cid)

        # ✅ restore selection AFTER rebuilding
        for i, cid in enumerate(row_to_cid):
            if cid in selected_cids:
                pc_list.selection_set(i)

        refresh_job["id"] = window.after(1000, refresh_pc_list)

    refresh_pc_list()

    # ---------- Buttons ----------
    ttk.Label(timer_frame, text="Send to Selected PC(s):", font=("Arial", 12, "bold")).grid(
        row=4, column=0, columnspan=2, pady=(5, 5), sticky="w", padx=10
    )

    r = 5
    for label, minutes in timer_options.items():
        ttk.Button(
            timer_frame,
            text=f"{label} (Selected)",
            style="Transparent.TButton",
            width=34,
            command=lambda m=minutes: (
                messagebox.showwarning("No Selection", "Select one or more PCs first.")
                if not get_selected_cids()
                else send_command_to_many(get_selected_cids(), action_var.get(), m)
            )
        ).grid(row=r, column=0, columnspan=2, padx=10, pady=3, sticky="nsew")
        r += 1

    ttk.Button(
        timer_frame,
        text="Run Now (Selected)",
        style="Transparent.TButton",
        width=34,
        command=lambda: (
            messagebox.showwarning("No Selection", "Select one or more PCs first.")
            if not get_selected_cids()
            else send_command_to_many(get_selected_cids(), action_var.get(), 0)
        )
    ).grid(row=r, column=0, columnspan=2, padx=10, pady=(8, 3), sticky="nsew")
    r += 1

    ttk.Separator(timer_frame, orient="horizontal").grid(row=r, column=0, columnspan=2, sticky="ew", pady=10)
    r += 1

    ttk.Label(timer_frame, text="Broadcast to ALL PCs:", font=("Arial", 12, "bold")).grid(
        row=r, column=0, columnspan=2, pady=(0, 5), sticky="w", padx=10
    )
    r += 1

    ttk.Button(
        timer_frame,
        text="Run Now (ALL)",
        style="Transparent.TButton",
        width=34,
        command=lambda: send_command_to_all(action_var.get(), 0)
    ).grid(row=r, column=0, columnspan=2, padx=10, pady=3, sticky="nsew")
    r += 1

    def go_back():
        # ✅ stop refresh loop cleanly
        if refresh_job["id"]:
            try:
                window.after_cancel(refresh_job["id"])
            except:
                pass
            refresh_job["id"] = None
        show_teacher_options(teacher, section, subject)

    ttk.Button(
        timer_frame,
        text="Back",
        style="Transparent.TButton",
        width=15,
        command=go_back
    ).grid(row=r, column=0, columnspan=2, pady=12)

    timer_frame.grid_columnconfigure(0, weight=1)
    timer_frame.grid_columnconfigure(1, weight=1)
    timer_frame.grid_rowconfigure(3, weight=1)

def send_command_to_clients(action):
    try:
        if not client_sockets:
            messagebox.showwarning("No Students Connected", "There are no students currently connected.")
            return

        for client_socket in client_sockets:
            command = {"action": action.lower(), "timer": 0}  # ⬅️ Fixed key here!
            client_socket.send(json.dumps(command).encode())

        messagebox.showinfo("Command Sent", f"✅ Sent '{action}' command to all students.")

    except Exception as e:
        messagebox.showerror("Error", f"Failed to send command: {e}")

def send_command_to_students(action):
    """Send shutdown or sleep command to all connected students."""
    try:
        if not client_sockets:
            messagebox.showwarning("No Students Connected", "There are no students currently connected.")
            return

        for client_socket in client_sockets:
            command = {"action": action}
            client_socket.send(json.dumps(command).encode())

        messagebox.showinfo("Command Sent", f"✅ Sent '{action}' command to all students.")
    except Exception as e:
        messagebox.showerror("Error", f"Failed to send command: {e}")

def safe_filename(name: str) -> str:
    return "".join(ch for ch in name if ch not in r'<>:"/\|?*').strip()
def open_folder(path: str):
    try:
        if os.name == "nt":
            subprocess.Popen(f'explorer "{path}"')
        else:
            subprocess.Popen(["open", path])
    except Exception as e:
        print("⚠ Could not open folder:", e)

def ask_quiz_upload_details(teacher, section, subject):
    win = tk.Toplevel(window)
    win.title("Quiz Upload Details")
    win.geometry("420x280")
    win.resizable(False, False)
    win.grab_set()
    win.configure(bg="white")

    result = {"quiz_number": None, "quiz_title": "", "time_limit": 0}

    tk.Label(
        win,
        text="Enter Quiz Details",
        font=("Helvetica", 16, "bold"),
        bg="white",
        fg="#4B0082"
    ).pack(pady=10)
    form = tk.Frame(win, bg="white")
    form.pack(pady=10, padx=20, fill="x")

    # ✅ auto-suggest next quiz number
    next_quiz_num = 1
    try:
        base_dir = get_quiz_base_dir(teacher, section, subject)
        if os.path.exists(base_dir):
            nums = []
            for filename in os.listdir(base_dir):
                if filename.startswith("quiz_") and filename.endswith(".json"):
                    try:
                        n = int(filename.replace("quiz_", "").replace(".json", ""))
                        nums.append(n)
                    except:
                        pass
            if nums:
                next_quiz_num = max(nums) + 1
    except:
        pass

    tk.Label(form, text="Quiz Number:", font=("Helvetica", 12), bg="white").grid(row=0, column=0, sticky="w", pady=5)
    quiz_num_var = tk.StringVar(value=str(next_quiz_num))
    quiz_num_entry = ttk.Entry(form, textvariable=quiz_num_var, width=25)
    quiz_num_entry.grid(row=0, column=1, pady=5, sticky="ew")

    tk.Label(form, text="Quiz Title:", font=("Helvetica", 12), bg="white").grid(row=1, column=0, sticky="w", pady=5)
    quiz_title_var = tk.StringVar()
    quiz_title_entry = ttk.Entry(form, textvariable=quiz_title_var, width=25)
    quiz_title_entry.grid(row=1, column=1, pady=5, sticky="ew")

    tk.Label(form, text="Time Limit (minutes):", font=("Helvetica", 12), bg="white").grid(row=2, column=0, sticky="w", pady=5)
    time_var = tk.StringVar(value="0")
    time_entry = ttk.Entry(form, textvariable=time_var, width=25)
    time_entry.grid(row=2, column=1, pady=5, sticky="ew")

    form.grid_columnconfigure(1, weight=1)

    def on_ok():
        try:
            qn = int(quiz_num_var.get().strip())
            if qn <= 0:
                raise ValueError
        except:
            messagebox.showerror("Invalid", "Quiz number must be a positive number.")
            return

        title = quiz_title_var.get().strip()
        if not title:
            messagebox.showerror("Invalid", "Quiz title cannot be empty.")
            return

        try:
            tl = int(time_var.get().strip())
            if tl < 0:
                raise ValueError
        except:
            messagebox.showerror("Invalid", "Time limit must be 0 or a positive number.")
            return

        confirm = messagebox.askyesno(
            "Confirm Quiz Upload",
            f"Upload this quiz?\n\n"
            f"Quiz #: {qn}\n"
            f"Title: {title}\n"
            f"Time Limit: {tl} minute(s)"
        )
        if not confirm:
            return

        result["quiz_number"] = qn
        result["quiz_title"] = title
        result["time_limit"] = tl
        win.destroy()

    def on_cancel():
        win.destroy()

    btns = tk.Frame(win, bg="white")
    btns.pack(pady=15)

    ttk.Button(btns, text="Continue", command=on_ok).pack(side="left", padx=10, ipadx=10)
    ttk.Button(btns, text="Cancel", command=on_cancel).pack(side="left", padx=10, ipadx=10)

    quiz_title_entry.focus_set()
    window.wait_window(win)

    if result["quiz_number"] is None:
        return None
    return result

def download_quiz_excel_template(teacher, section, subject):
    import os
    from datetime import datetime
    import pandas as pd

    desktop = os.path.join(os.path.expanduser("~"), "Desktop")
    base = os.path.join(desktop, "Quiz_Templates", safe_filename(teacher))
    os.makedirs(base, exist_ok=True)

    date_str = datetime.now().strftime("%Y_%m_%d")
    filename = f"{safe_filename(teacher)}_{safe_filename(section)}_{safe_filename(subject)}_{date_str}_QUIZ_TEMPLATE.xlsx"
    filepath = os.path.join(base, filename)

    # ✅ Generic quiz sheet with sample rows only
    quiz_df = pd.DataFrame([
        {
            "Type": "MCQ",
            "Question": "What is Python?",
            "A": "Programming Language",
            "B": "Snake",
            "C": "Operating System",
            "D": "Database",
            "Answer": "A"
        },
        {
            "Type": "ID",
            "Question": "Who created Python?",
            "A": "",
            "B": "",
            "C": "",
            "D": "",
            "Answer": "Guido van Rossum"
        },
        {
            "Type": "",
            "Question": "",
            "A": "",
            "B": "",
            "C": "",
            "D": "",
            "Answer": ""
        },
    ], columns=["Type", "Question", "A", "B", "C", "D", "Answer"])

    guide_df = pd.DataFrame([
        {"Guide": "Type must be MCQ or ID"},
        {"Guide": "For MCQ, fill A/B/C/D and set Answer to A/B/C/D"},
        {"Guide": "For ID, leave A/B/C/D blank and put the text answer in Answer"},
        {"Guide": "You can add as many rows as you want"},
        {"Guide": "Only the QUIZ sheet will be parsed"},
    ])

    with pd.ExcelWriter(filepath, engine="openpyxl") as writer:
        quiz_df.to_excel(writer, index=False, sheet_name="QUIZ")
        guide_df.to_excel(writer, index=False, sheet_name="GUIDE")

    messagebox.showinfo(
        "Template Created",
        f"Excel template created!\n\n"
        f"You can now add as many MCQ or ID questions as you want.\n\n"
        f"Saved to:\n{filepath}"
    )
    open_folder(base)

def download_quiz_txt_template(teacher, section, subject):
    from datetime import datetime
    import os

    desktop = os.path.join(os.path.expanduser("~"), "Desktop")
    base = os.path.join(desktop, "Quiz_Templates", safe_filename(teacher))
    os.makedirs(base, exist_ok=True)

    date_str = datetime.now().strftime("%Y_%m_%d")
    filename = f"{safe_filename(teacher)}_{safe_filename(section)}_{safe_filename(subject)}_{date_str}_QUIZ_TEMPLATE.txt"
    filepath = os.path.join(base, filename)

    lines = [
        "# QUIZ TEMPLATE (TXT)",
        "# Lines starting with # are ignored",
        "# You can add as many questions as you want",
        "# TYPE must be MCQ or ID",
        "",
        "TYPE: MCQ",
        "Question: What is Python?",
        "A. Programming Language",
        "B. Snake",
        "C. Operating System",
        "D. Database",
        "Correct Answer: A",
        "",
        "TYPE: ID",
        "Question: Who created Python?",
        "Correct Answer: Guido van Rossum",
        "",
        "# Add more questions below this line",
        "",
        "TYPE: MCQ",
        "Question: ",
        "A. ",
        "B. ",
        "C. ",
        "D. ",
        "Correct Answer: ",
        "",
        "TYPE: ID",
        "Question: ",
        "Correct Answer: ",
        ""
    ]

    with open(filepath, "w", encoding="utf-8") as f:
        f.write("\n".join(lines))

    messagebox.showinfo(
        "Template Created",
        f"TXT template created!\n\n"
        f"You can now add as many MCQ or ID questions as you want.\n\n"
        f"Saved to:\n{filepath}"
    )
    open_folder(base)


def view_selected_quiz(teacher, section, subject):
    selected_text = quiz_listbox.get(tk.ACTIVE)
    if not selected_text or "No quizzes available" in selected_text:
        messagebox.showerror("Error", "Please select a quiz first.")
        return

    quiz_number = extract_quiz_number(selected_text)
    if quiz_number is None:
        messagebox.showerror("Error", "Invalid quiz selection format!")
        return

    base_dir = get_quiz_base_dir(teacher, section, subject)
    quiz_file_path = os.path.join(base_dir, f"quiz_{quiz_number}.json")

    if not os.path.exists(quiz_file_path):
        messagebox.showerror("Error", "Quiz file not found for this teacher/section/subject.")
        return

    try:
        with open(quiz_file_path, "r", encoding="utf-8") as f:
            quiz_info = json.load(f)
    except Exception as e:
        messagebox.showerror("Error", f"Failed to read quiz:\n{e}")
        return

    # ---- Popup UI ----
    win = tk.Toplevel(window)
    win.title(f"View Quiz {quiz_number}")
    win.geometry("900x650")
    win.configure(bg="#F5F5F5")

    header = tk.Frame(win, bg="white", height=70)
    header.pack(fill="x")

    title = quiz_info.get("quiz_title", "Untitled")
    time_limit = quiz_info.get("time_limit", 0)
    tl_txt = f"{time_limit} minute(s)" if time_limit else "No Timer"

    tk.Label(
        header,
        text=f"Quiz {quiz_number}: {title}",
        font=("Helvetica", 18, "bold"),
        bg="white", fg="#4B0082"
    ).pack(side="left", padx=20, pady=15)

    tk.Label(
        header,
        text=f"⏱ {tl_txt}",
        font=("Helvetica", 14, "bold"),
        bg="white", fg="red"
    ).pack(side="right", padx=20)

    meta = tk.Frame(win, bg="#F5F5F5")
    meta.pack(fill="x", padx=20, pady=(10, 5))

    tk.Label(
        meta,
        text=f"Instructor: {quiz_info.get('instructor', '')} | Section: {quiz_info.get('section', '')} | Subject: {quiz_info.get('subject', '')}",
        font=("Helvetica", 12),
        bg="#F5F5F5"
    ).pack(anchor="w")

    tk.Label(
        meta,
        text=f"Uploaded: {quiz_info.get('timestamp', '')}",
        font=("Helvetica", 12),
        bg="#F5F5F5"
    ).pack(anchor="w")

    # Scroll area
    body = tk.Frame(win, bg="#F5F5F5")
    body.pack(fill="both", expand=True, padx=20, pady=10)

    canvas = tk.Canvas(body, bg="#F5F5F5", highlightthickness=0)
    scrollbar = ttk.Scrollbar(body, orient="vertical", command=canvas.yview)
    scroll_frame = tk.Frame(canvas, bg="#F5F5F5")

    scroll_frame.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
    canvas.create_window((0, 0), window=scroll_frame, anchor="nw")
    canvas.configure(yscrollcommand=scrollbar.set)

    canvas.pack(side="left", fill="both", expand=True)
    scrollbar.pack(side="right", fill="y")

    # Optional: mousewheel scroll
    def _on_mousewheel(event):
        canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")

    win.bind_all("<MouseWheel>", _on_mousewheel)

    questions = quiz_info.get("questions", [])
    if not questions:
        tk.Label(scroll_frame, text="No questions found.", bg="#F5F5F5", font=("Helvetica", 14)).pack(pady=20)
    else:
        for i, q in enumerate(questions, start=1):
            qtype = q.get("type", "MCQ")
            card = tk.Frame(scroll_frame, bg="white", bd=2, relief="ridge")
            card.pack(fill="x", padx=5, pady=8)

            tk.Label(
                card,
                text=f"{i}. [{qtype}] {q.get('question', '')}",
                font=("Helvetica", 14, "bold"),
                bg="white",
                wraplength=780,
                justify="left"
            ).pack(anchor="w", padx=15, pady=(10, 5))

            if qtype == "MCQ":
                for ch in q.get("choices", []):
                    tk.Label(card, text=ch, font=("Helvetica", 12), bg="white").pack(anchor="w", padx=35)
                tk.Label(
                    card,
                    text=f"Answer: {q.get('answer', '')}",
                    font=("Helvetica", 12, "bold"),
                    bg="white", fg="green"
                ).pack(anchor="w", padx=15, pady=(5, 10))
            else:  # ID
                tk.Label(
                    card,
                    text=f"Answer: {q.get('answer', '')}",
                    font=("Helvetica", 12, "bold"),
                    bg="white", fg="green"
                ).pack(anchor="w", padx=15, pady=(5, 10))

    footer = tk.Frame(win, bg="white")
    footer.pack(fill="x")

    ttk.Button(footer, text="Close", command=win.destroy).pack(pady=10)


def show_quiz_upload(teacher, section, subject=None):
    global current_subject

    if not subject:
        subject = current_subject
        if not subject:
            messagebox.showerror("Error", "No subject selected! Please choose a subject first.")
            return

    print(f"✅ Uploading quiz for subject: {subject}")
    clear_grid()

    style = ttk.Style()

    # Orange style (your existing)
    style.configure("Orange.TButton",
                    font=("Helvetica", 12, "bold"),
                    background="#FF8C00",
                    foreground="white")
    style.map("Orange.TButton",
              background=[("active", "#E07B00")],
              foreground=[("active", "white")])

    # ✅ Red style for Delete button
    style.configure("Red.TButton",
                    font=("Helvetica", 12, "bold"),
                    background="#DC3545",
                    foreground="white")
    style.map("Red.TButton",
              background=[("active", "#B02A37")],
              foreground=[("active", "white")])

    title_label = ttk.Label(window, text="Quiz Management", font=("Helvetica", 18, "bold"),
                            background="white", foreground="#8A2BE2", anchor="center")
    title_label.place(relx=0.5, rely=0.05, anchor="center")

    subtitle_label = ttk.Label(window, text=f"Hello Professor: {teacher} | Section: {section} | Subject: {subject}",
                               font=("Helvetica", 14), background="white", foreground="black", anchor="center")
    subtitle_label.place(relx=0.5, rely=0.10, anchor="center")

    main_frame = ttk.Frame(window, bootstyle="light", padding=20)
    main_frame.place(relx=0.5, rely=0.5, anchor="center", width=700, height=520)

    border_frame = tk.Frame(main_frame, background="darkgreen", padx=5, pady=5)
    border_frame.pack(fill=tk.BOTH, expand=True)

    inner_frame = ttk.Frame(border_frame, bootstyle="light", padding=20)
    inner_frame.pack(fill=tk.BOTH, expand=True)

    listbox_frame = ttk.Frame(inner_frame)
    listbox_frame.pack(pady=10, fill=tk.BOTH, expand=True)

    global quiz_listbox
    quiz_listbox = tk.Listbox(listbox_frame, font=("Arial", 12), height=10, width=50)
    quiz_listbox.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

    scrollbar = ttk.Scrollbar(listbox_frame, orient=tk.VERTICAL, command=quiz_listbox.yview)
    scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
    quiz_listbox.config(yscrollcommand=scrollbar.set)

    button_frame = ttk.Frame(inner_frame)
    button_frame.pack(pady=10, fill=tk.X)

    # Row 0: Upload (full width)
    ttk.Button(
        button_frame, text="Upload Quiz", style="Orange.TButton",
        command=lambda: upload_quiz(teacher, section)
    ).grid(row=0, column=0, columnspan=2, padx=10, pady=5, sticky="ew")

    # Row 1: Delete + Toggle
    ttk.Button(
        button_frame, text="Delete Quiz", style="Red.TButton",
        command=lambda: delete_quiz(teacher, section, subject)
    ).grid(row=1, column=0, padx=10, pady=5, sticky="ew")

    ttk.Button(
        button_frame, text="Toggle Availability", style="Orange.TButton",
        command=lambda: toggle_quiz_status(teacher, section, subject)
    ).grid(row=1, column=1, padx=10, pady=5, sticky="ew")

    # Row 2: Templates
    ttk.Button(
        button_frame, text="⬇ Excel Template", style="Orange.TButton",
        command=lambda: download_quiz_excel_template(teacher, section, subject)
    ).grid(row=2, column=0, padx=10, pady=5, sticky="ew")

    ttk.Button(
        button_frame, text="⬇ TXT Template", style="Orange.TButton",
        command=lambda: download_quiz_txt_template(teacher, section, subject)
    ).grid(row=2, column=1, padx=10, pady=5, sticky="ew")

    # ✅ Row 3: View + Back (side by side)
    ttk.Button(
        button_frame, text="View Quiz", style="Orange.TButton",
        command=lambda: view_selected_quiz(teacher, section, subject)
    ).grid(row=3, column=0, padx=10, pady=(8, 0), sticky="ew")

    ttk.Button(
        button_frame, text="Back", style="Orange.TButton",
        command=lambda: show_teacher_options(teacher, section, subject)
    ).grid(row=3, column=1, padx=10, pady=(8, 0), sticky="ew")

    # Make columns expand evenly
    button_frame.grid_columnconfigure(0, weight=1)
    button_frame.grid_columnconfigure(1, weight=1)

    update_quiz_list(teacher, section, subject)


def parse_quiz_txt(file_path):
    questions = []
    with open(file_path, "r", encoding="utf-8") as f:
        lines = [line.strip() for line in f.readlines() if line.strip() and not line.strip().startswith("#")]

    i = 0
    while i < len(lines):
        if lines[i].upper().startswith("TYPE:"):
            qtype = lines[i].split(":", 1)[1].strip().upper()
            i += 1

            if i < len(lines) and lines[i].lower().startswith("question:"):
                question_text = lines[i].split(":", 1)[1].strip()
                i += 1
            else:
                raise ValueError("TXT format error: missing 'Question:'")

            if qtype == "MCQ":
                choices = [lines[i], lines[i + 1], lines[i + 2], lines[i + 3]]
                i += 4
                if not lines[i].lower().startswith("correct answer:"):
                    raise ValueError("TXT format error: missing 'Correct Answer:'")
                correct = lines[i].split(":", 1)[1].strip().upper()[-1]
                i += 1

                questions.append({
                    "type": "MCQ",
                    "question": question_text,
                    "choices": choices,
                    "answer": correct
                })

            elif qtype == "ID":
                if not lines[i].lower().startswith("correct answer:"):
                    raise ValueError("TXT format error: missing 'Correct Answer:' for ID")
                correct_text = lines[i].split(":", 1)[1].strip()
                i += 1

                questions.append({
                    "type": "ID",
                    "question": question_text,
                    "answer": correct_text
                })
            else:
                raise ValueError("TXT format error: TYPE must be MCQ or ID")
        else:
            i += 1

    return questions

def upload_quiz(teacher, section):
    global current_subject
    import os

    if not current_subject:
        messagebox.showerror("Error", "Please select a subject first!")
        return

    # ✅ pass the required arguments
    details = ask_quiz_upload_details(teacher, section, current_subject)
    if not details:
        return

    quiz_number = details["quiz_number"]
    quiz_title = details["quiz_title"]
    time_limit = details["time_limit"]

    file_path = filedialog.askopenfilename(
        title="Select Quiz File",
        filetypes=[
            ("Quiz Files", "*.txt *.xlsx"),
            ("Text Files", "*.txt"),
            ("Excel Files", "*.xlsx"),
        ]
    )
    if not file_path:
        messagebox.showerror("Error", "No file selected.")
        return

    ext = os.path.splitext(file_path)[1].lower()

    try:
        if ext == ".txt":
            parsed_questions = parse_quiz_txt(file_path)
        elif ext == ".xlsx":
            parsed_questions = parse_quiz_excel(file_path)
        else:
            messagebox.showerror("Error", "Unsupported file type. Use .txt or .xlsx only.")
            return
    except Exception as e:
        messagebox.showerror("Error", f"Failed to read quiz file:\n{e}")
        return

    if not parsed_questions:
        messagebox.showerror("Error", "No questions found in the file.")
        return

    base_dir = get_quiz_base_dir(teacher, section, current_subject)
    os.makedirs(base_dir, exist_ok=True)

    quiz_file_path = os.path.join(base_dir, f"quiz_{quiz_number}.json")

    if os.path.exists(quiz_file_path):
        replace = messagebox.askyesno(
            "Quiz Already Exists",
            f"Quiz {quiz_number} already exists for:\n\n"
            f"{teacher} | {section} | {current_subject}\n\n"
            f"Do you want to REPLACE it?"
        )
        if not replace:
            return

    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    quiz_data = {
        "quiz_number": quiz_number,
        "quiz_title": quiz_title,
        "time_limit": time_limit,
        "subject": current_subject,
        "instructor": teacher,
        "section": section,
        "timestamp": timestamp,
        "status": "Accepting",
        "questions": parsed_questions
    }

    with open(quiz_file_path, "w", encoding="utf-8") as file:
        json.dump(quiz_data, file, indent=4)

    messagebox.showinfo(
        "Success",
        f"Quiz uploaded successfully!\n\n"
        f"Quiz #: {quiz_number}\n"
        f"Title: {quiz_title}\n"
        f"Subject: {current_subject}\n"
        f"Time Limit: {time_limit} minute(s)\n"
        f"Questions: {len(parsed_questions)}\n\n"
        f"Saved to:\n{quiz_file_path}"
    )

    update_quiz_list(teacher, section, current_subject)



def parse_quiz_excel(file_path):
    df = pd.read_excel(file_path, sheet_name="QUIZ")

    required_cols = {"Type", "Question", "Answer"}
    if not required_cols.issubset(df.columns):
        raise ValueError("Excel must have at least: Type, Question, Answer")

    for col in ["A", "B", "C", "D"]:
        if col not in df.columns:
            df[col] = ""

    questions = []
    for _, row in df.iterrows():
        qtype = str(row["Type"]).strip().upper()
        question_text = str(row["Question"]).strip()

        if not question_text or question_text.lower() == "nan":
            continue

        if qtype == "MCQ":
            questions.append({
                "type": "MCQ",
                "question": question_text,
                "choices": [
                    f"A. {row['A']}",
                    f"B. {row['B']}",
                    f"C. {row['C']}",
                    f"D. {row['D']}",
                ],
                "answer": str(row["Answer"]).strip().upper()
            })
        elif qtype == "ID":
            questions.append({
                "type": "ID",
                "question": question_text,
                "answer": str(row["Answer"]).strip()
            })
        else:
            raise ValueError("Invalid Type found. Use only MCQ or ID.")

    return questions


def save_quiz_info(quiz_info):
    base_dir = get_quiz_base_dir(quiz_info["instructor"], quiz_info["section"], quiz_info["subject"])
    os.makedirs(base_dir, exist_ok=True)
    file_path = os.path.join(base_dir, f'quiz_{quiz_info["quiz_number"]}.json')
    with open(file_path, "w", encoding="utf-8") as f:
        json.dump(quiz_info, f, indent=4)


def update_quiz_list(teacher, section, subject):
    quiz_listbox.delete(0, tk.END)

    base_dir = get_quiz_base_dir(teacher, section, subject)
    if not os.path.exists(base_dir):
        quiz_listbox.insert(tk.END, "No quizzes available.")
        return

    quiz_files = [f for f in os.listdir(base_dir) if f.endswith(".json")]
    if not quiz_files:
        quiz_listbox.insert(tk.END, "No quizzes available.")
        return

    def quiz_sort_key(filename):
        try:
            with open(os.path.join(base_dir, filename), "r", encoding="utf-8") as file:
                data = json.load(file)
            return int(data.get("quiz_number", 999999))
        except:
            return 999999

    quiz_files.sort(key=quiz_sort_key)

    for quiz_file in quiz_files:
        path = os.path.join(base_dir, quiz_file)
        try:
            with open(path, "r", encoding="utf-8") as file:
                quiz_info = json.load(file)

            subject = quiz_info.get("subject", "No Subject")
            status = quiz_info.get("status", "Accepting")
            quiz_number = quiz_info.get("quiz_number", "Unknown")
            timestamp = quiz_info.get("timestamp", "Unknown Date")

            title = quiz_info.get("quiz_title", "Untitled")
            time_limit = quiz_info.get("time_limit", 0)
            tl_txt = f"{time_limit}m" if time_limit else "No Timer"

            formatted_entry = f"[{status}] {subject} - Quiz {quiz_number}: {title} ({tl_txt}) (Uploaded: {timestamp})"
            quiz_listbox.insert(tk.END, formatted_entry)

        except Exception as e:
            print(f"⚠️ Failed to read {quiz_file}: {e}")


import re


def extract_quiz_number(selected_text):
    match = re.search(r'Quiz (\d+)', selected_text)
    return int(match.group(1)) if match else None


def toggle_quiz_status(teacher, section, subject):
    selected_text = quiz_listbox.get(tk.ACTIVE)
    if not selected_text:
        messagebox.showerror("Error", "Please select a quiz to toggle.")
        return

    quiz_number = extract_quiz_number(selected_text)
    if quiz_number is None:
        messagebox.showerror("Error", "Invalid quiz selection format!")
        return

    base_dir = get_quiz_base_dir(teacher, section, subject)
    quiz_file_path = os.path.join(base_dir, f"quiz_{quiz_number}.json")

    if not os.path.exists(quiz_file_path):
        messagebox.showerror("Error", "Quiz file not found for this teacher/section/subject.")
        return

    try:
        with open(quiz_file_path, "r", encoding="utf-8") as file:
            quiz_info = json.load(file)

        current_status = quiz_info.get("status", "Accepting")
        new_status = "Closed" if current_status == "Accepting" else "Accepting"
        quiz_info["status"] = new_status

        with open(quiz_file_path, "w", encoding="utf-8") as file:
            json.dump(quiz_info, file, indent=4)

        messagebox.showinfo("Success", f"Quiz {quiz_number} is now {new_status}!")
        update_quiz_list(teacher, section, subject)

    except Exception as e:
        messagebox.showerror("Error", f"Failed to toggle quiz:\n{e}")


def delete_quiz(teacher, section, subject):
    selected = quiz_listbox.curselection()
    if not selected:
        messagebox.showerror("Error", "No quiz selected!")
        return

    selected_text = quiz_listbox.get(selected[0])
    quiz_number = extract_quiz_number(selected_text)
    if quiz_number is None:
        messagebox.showerror("Error", "Invalid quiz selection format!")
        return

    base_dir = get_quiz_base_dir(teacher, section, subject)
    quiz_file_path = os.path.join(base_dir, f"quiz_{quiz_number}.json")

    if not os.path.exists(quiz_file_path):
        messagebox.showerror("Error", "Quiz file not found for this teacher/section/subject.")
        return

    confirm = messagebox.askyesno(
        "Confirm Delete",
        f"Delete Quiz {quiz_number} for:\n\n{teacher} | {section} | {subject} ?"
    )
    if not confirm:
        return

    try:
        os.remove(quiz_file_path)
        update_quiz_list(teacher, section, subject)
        messagebox.showinfo("Success", f"Quiz {quiz_number} deleted successfully!")
    except Exception as e:
        messagebox.showerror("Error", f"Failed to delete quiz:\n{e}")


def show_quiz_summary(quiz_info):
    clear_grid()
    style = ttk.Style()
    style.configure("DarkViolet.TButton", font=("Arial", 12), background="#5D3FD3",
                    foreground="white")  # Increased font size for a larger button
    style.configure("DarkViolet.TButton", font=("Arial", 12), background="#5D3FD3",
                    foreground="white")  # Increased font size for a larger button
    style.map("DarkViolet.TButton",
              background=[("active", "#FF69B4")],
              foreground=[("active", "white")])

    main_frame = ttk.Frame(window, padding=10)
    main_frame.grid(row=0, column=0, sticky=tk.NSEW, padx=20, pady=20)

    lbl = ttk.Label(main_frame, text="Quiz Summary", font=("Helvetica", 16, "bold"), foreground="#8A2BE2")
    lbl.grid(row=0, column=0, columnspan=3, pady=10, sticky=tk.N)

    summary_frame = ttk.Frame(main_frame, padding=10, style="Gray.TFrame")  # Reduced padding for smaller frame
    summary_frame.grid(row=1, column=0, columnspan=3, padx=10, pady=10, sticky=tk.NSEW)

    style.configure("Gray.TFrame", background="#D3D3D3")  # Light gray background color

    lbl_quiz_number = ttk.Label(summary_frame, text=f"Quiz #: {quiz_info['quiz_number']}", font=("Helvetica", 14),
                                foreground="#8A2BE2", background="#D3D3D3")
    lbl_quiz_number.grid(row=0, column=0, pady=5, sticky=tk.W)

    lbl_num_questions = ttk.Label(summary_frame, text=f"Number of Questions: {quiz_info['num_questions']}",
                                  font=("Helvetica", 14), foreground="#8A2BE2", background="#D3D3D3")
    lbl_num_questions.grid(row=1, column=0, pady=5, sticky=tk.W)

    lbl_timestamp = ttk.Label(summary_frame, text=f"Date and Time: {quiz_info['timestamp']}", font=("Helvetica", 14),
                              foreground="#8A2BE2", background="#D3D3D3")
    lbl_timestamp.grid(row=2, column=0, pady=5, sticky=tk.W)

    def done_action():
        main_frame.destroy()
        show_teacher_options(quiz_info['teacher'], quiz_info['section'])

    done_btn = ttk.Button(main_frame, text="Done", style="DarkViolet.TButton", width=20, padding=(10, 5),
                          command=done_action)  # Increased button width and padding
    done_btn.grid(row=2, column=0, pady=20, sticky=tk.S)

    main_frame.grid_rowconfigure(1, weight=1)
    main_frame.grid_columnconfigure(0, weight=1)

    window.grid_rowconfigure(0, weight=1)
    window.grid_columnconfigure(0, weight=1)


def show_program_selection(teacher, section, subject):
    clear_grid()
    window.title("Program Selection")

    style = ttk.Style()
    style.configure("Program.TButton",
                    font=("Arial", 12, "bold"),
                    foreground="white",
                    background="#6f42c1",
                    padding=(10, 6))
    style.map("Program.TButton",
              background=[("active", "#5a32a3")])

    style.configure("Back.TButton",
                    font=("Arial", 11),
                    foreground="white",
                    background="#6c757d")
    style.map("Back.TButton",
              background=[("active", "#5a6268")])

    container = ttk.Frame(window, padding=30)
    container.place(relx=0.5, rely=0.5, anchor="center")

    ttk.Label(
        container,
        text="💻 Program Launcher",
        font=("Helvetica", 20, "bold"),
        foreground="#5D3FD3"
    ).pack(pady=(10, 5))

    ttk.Label(
        container,
        text=f"Instructor: {teacher} | Section: {section} | Subject: {subject}",
        font=("Helvetica", 12),
        foreground="gray"
    ).pack(pady=(0, 20))

    grid_frame = ttk.Frame(container, padding=10)
    grid_frame.pack()

    programs = [
        "XAMPP", "CHROME", "ECLIPSE", "PENCILDRAW", "VISUAL STUDIO",
        "MYSQL", "NETBEANS", "NODEJS", "PYTHONIDE", "C++"
    ]

    for idx, program in enumerate(programs):
        row, col = divmod(idx, 2)
        btn = ttk.Button(
            grid_frame, text=program, style="Program.TButton", width=25,
            command=lambda prog=program: send_program_command(prog, section)
        )
        btn.grid(row=row, column=col, padx=10, pady=10)

    back_btn = ttk.Button(container, text="🔙 Back", style="Back.TButton", width=20,
                          command=lambda: show_teacher_options(teacher, section, subject))
    back_btn.pack(pady=(20, 0))


def send_program_command(program, section):
    try:
        command = json.dumps({
            "action": "launch_program",
            "program": program,
            "section": section
        })
        print("⬆️ Sending:", command)

        for client_socket in client_sockets:
            try:
                client_socket.sendall(command.encode('utf-8'))
            except Exception as e:
                print(f"Failed to send command to a client: {e}")

        messagebox.showinfo("Program Command", f"Command to open {program} sent to section {section}")
    except Exception as e:
        messagebox.showerror("Command Error", f"Failed to send command: {e}")


conn = sqlite3.connect(os.path.join(os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "..")), 'teacher_management.db'))
c = conn.cursor()
c.execute("PRAGMA table_info(teacher_subjects)")
for row in c.fetchall():
    print(row)
conn.close()


def show_add_remove_student(teacher, section, subject):
    clear_grid()

    frame = ttk.Frame(window, bootstyle="dark", padding=20)
    frame.place(relx=0.5, rely=0.5, anchor="center", width=900, height=650)

    ttk.Label(frame, text="📚 Manage Students in Your Section",
              font=("Helvetica", 18, "bold"), foreground="#FF8C00").pack(pady=10)

    # Treeview
    columns = ("ID", "Name", "Student Number", "RFID")
    tree_frame = ttk.Frame(frame)
    tree_frame.pack(pady=10, fill="both", expand=True)

    tree = ttk.Treeview(tree_frame, columns=columns, show="headings", height=8)
    tree.pack(side="left", fill="both", expand=True)

    for col in columns:
        tree.heading(col, text=col)
        tree.column(col, anchor="center")

    scrollbar = ttk.Scrollbar(tree_frame, orient="vertical", command=tree.yview)
    tree.configure(yscroll=scrollbar.set)
    scrollbar.pack(side="right", fill="y")

    def update_student_list():
        tree.delete(*tree.get_children())
        students = get_students(section, subject, teacher)
        for student in students:
            tree.insert("", "end", values=(student[0], student[1], student[2], student[3]))

    def on_remove_student():
        selected_item = tree.selection()
        if not selected_item:
            messagebox.showwarning("Select Student", "Please select a student to remove.")
            return

        student_info = tree.item(selected_item, "values")
        student_number = student_info[2]  # Student Number column
        student_name = student_info[1]  # Student Name column

        confirm = messagebox.askyesno("Confirm Remove", f"Remove {student_name} from this subject?")
        if not confirm:
            return

        try:
            conn = sqlite3.connect(os.path.join(os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "..")), 'teacher_management.db'))
            c = conn.cursor()

            # ✅ Delete from student_assignments only
            c.execute("""
                      DELETE
                      FROM student_assignments
                      WHERE student_number = ? AND section = ? AND subject = ? AND instructor = ?
                      """, (student_number, section, subject, teacher))

            conn.commit()
            messagebox.showinfo("Success", f"{student_name} removed from this subject.")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to remove student: {e}")
        finally:
            conn.close()
            update_student_list()

    def create_excel_template():
        try:
            from openpyxl import Workbook
            import os
            import subprocess
            from datetime import datetime

            desktop = os.path.join(os.path.expanduser("~"), "Desktop")

            base_folder = os.path.join(desktop, "Student_Import_Templates")
            os.makedirs(base_folder, exist_ok=True)

            date_str = datetime.now().strftime("%Y_%m_%d")

            filename = f"Student_Import_Template_{section}_{subject}_{date_str}.xlsx"
            file_path = os.path.join(base_folder, filename)

            wb = Workbook()
            ws = wb.active
            ws.title = "Students"

            ws.append(["Name", "Student Number", "RFID Code"])

            ws.append(["Juan Dela Cruz", "2024-00001", "1234567890"])

            wb.save(file_path)

            try:
                if os.name == "nt":  # Windows
                    subprocess.Popen(f'explorer "{base_folder}"')
                else:
                    subprocess.Popen(["open", base_folder])
            except Exception as e:
                print("⚠ Could not open folder:", e)

            # ✅ Confirmation
            messagebox.showinfo(
                "Template Created",
                "Excel template created successfully!\n\n"
                "The folder has been opened for you.\n\n"
                "Fill up the file, then click 'Import from Excel'."
            )

        except Exception as e:
            messagebox.showerror("Error", f"Failed to create template:\n{e}")

    def add_students_from_excel():
        path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx *.xls")])
        if not path:
            return
        try:
            df = pd.read_excel(path)
            required = {"Name", "Student Number", "RFID Code"}
            if not required.issubset(df.columns):
                messagebox.showerror("Format Error", "Excel must include Name, Student Number, RFID Code.")
                return
            added = 0
            for _, row in df.iterrows():
                if pd.notna(row["Name"]) and pd.notna(row["Student Number"]) and pd.notna(row["RFID Code"]):
                    add_student(row["Name"], row["Student Number"], row["RFID Code"], section, subject, teacher)
                    added += 1
            messagebox.showinfo("Upload Complete", f"{added} students added.")
            update_student_list()
        except Exception as e:
            messagebox.showerror("Error", str(e))

    # Buttons
    style = ttk.Style()
    style.configure("Orange.TButton", font=("Helvetica", 12, "bold"), background="#FF8C00", foreground="white")
    style.map("Orange.TButton", background=[("active", "#E07B00")])

    button_frame = ttk.Frame(frame)
    button_frame.pack(pady=10)

    ttk.Button(button_frame, text="📄 Download Excel Template", style="Orange.TButton",
               command=create_excel_template).grid(row=0, column=0, padx=10, ipadx=10)

    ttk.Button(button_frame, text="📂 Import from Excel", style="Orange.TButton",
               command=add_students_from_excel).grid(row=0, column=1, padx=10, ipadx=10)

    ttk.Button(button_frame, text="❌ Remove Selected", style="Orange.TButton",
               command=on_remove_student).grid(row=0, column=2, padx=10, ipadx=10)

    # Search Area
    search_frame = ttk.LabelFrame(frame, text="🔍 Add Existing Student")
    search_frame.pack(pady=10, fill="x", padx=10)

    ttk.Label(search_frame, text="Search by Name or Student Number:").grid(row=0, column=0, padx=5, pady=5)
    name_var = tk.StringVar()
    search_entry = ttk.Combobox(search_frame, textvariable=name_var, width=40)
    search_entry.grid(row=0, column=1, padx=5)

    def get_all_student_names_and_numbers():
        conn = sqlite3.connect(os.path.join(os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "..")), 'teacher_management.db'))
        c = conn.cursor()
        c.execute("SELECT DISTINCT name, student_number FROM students")
        names = [f"{row[0]} | {row[1]}" for row in c.fetchall()]
        conn.close()
        return names

    def update_autocomplete(event=None):
        typed = name_var.get().lower()
        all_entries = get_all_student_names_and_numbers()
        filtered = [entry for entry in all_entries if typed in entry.lower()]
        search_entry["values"] = filtered

    search_entry.bind("<KeyRelease>", update_autocomplete)

    result_label = ttk.Label(search_frame, text="", font=("Arial", 10, "bold"), foreground="green")
    result_label.grid(row=1, column=0, columnspan=3)

    selected_student = {}

    def search_student():
        keyword = name_var.get().strip().split(" | ")[0]
        if not keyword:
            result_label.config(text="Please enter a name or student number.")
            return
        conn = sqlite3.connect(os.path.join(os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "..")), 'teacher_management.db'))
        c = conn.cursor()
        c.execute("SELECT name, student_number, rfid FROM students WHERE name LIKE ? OR student_number LIKE ?",
                  (f"%{keyword}%", f"%{keyword}%"))
        result = c.fetchone()
        conn.close()
        if result:
            name, number, rfid = result
            selected_student.clear()
            selected_student.update({"name": name, "student_number": number, "rfid": rfid})
            result_label.config(text=f"✅ Found: {name} | {number} | {rfid}")
        else:
            result_label.config(text="❌ No student found.")
            selected_student.clear()
    def assign_selected_student():
        if not selected_student:
            messagebox.showwarning("No Selection", "Please search for a student first.")
            return

        try:
            with sqlite3.connect(os.path.join(os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "..")), 'teacher_management.db')) as conn:
                c = conn.cursor()

                # Check for duplicates
                c.execute("""
                          SELECT 1
                          FROM student_assignments
                          WHERE student_number = ? AND section = ? AND subject = ? AND instructor = ?
                          """, (selected_student["student_number"], section, subject, teacher))
                already_exists = c.fetchone()

                if already_exists:
                    messagebox.showwarning("Already Assigned",
                                           f"{selected_student['name']} is already assigned to this subject and section.")
                    return

                # Add assignment
                c.execute("""
                          INSERT INTO student_assignments (student_number, section, subject, instructor)
                          VALUES (?, ?, ?, ?)
                          """, (selected_student["student_number"], section, subject, teacher))

                conn.commit()

            messagebox.showinfo("Assigned", f"{selected_student['name']} has been added to this section.")
            time.sleep(0.1)  # 🧠 Optional but helps release DB lock before refresh
            update_student_list()

        except sqlite3.OperationalError as e:
            messagebox.showerror("Database Error", f"Failed to assign student:\n{e}")
        finally:
            name_var.set("")
            result_label.config(text="")
            selected_student.clear()
    ttk.Button(search_frame, text="🔍 Search", style="Orange.TButton", command=search_student).grid(row=0, column=2,
                                                                                                   padx=5, pady=5)
    ttk.Button(search_frame, text="➕ Add This Student", style="Orange.TButton", command=assign_selected_student).grid(
        row=2, column=0, columnspan=3, pady=10)
    # Bottom button container for Back
    bottom_button_frame = ttk.Frame(frame)
    bottom_button_frame.pack(pady=10)
    bottom_button_frame = ttk.Frame(search_frame)
    bottom_button_frame.grid(row=2, column=0, columnspan=3, pady=10)

    ttk.Button(bottom_button_frame, text="➕ Add This Student", style="Orange.TButton",
               command=assign_selected_student).pack(side="left", padx=10)

    ttk.Button(bottom_button_frame, text="🔙 Back", style="Orange.TButton",
               command=lambda: show_teacher_options(teacher, section, subject)).pack(side="left", padx=20)

    update_student_list()


create_rfid_logs_table()
start_rfid_scanner()
show_teacher_selection()
window.mainloop()
