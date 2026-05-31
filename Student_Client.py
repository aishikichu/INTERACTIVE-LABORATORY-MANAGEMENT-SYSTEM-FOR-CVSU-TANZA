import tkinter as tk
from tkinter import messagebox, ttk
from PIL import Image, ImageTk
from datetime import datetime
import socket
import json
import os
import time
from ttkbootstrap import Style
import select
import threading
style = Style("cyborg")
window = style.master
window.title("Student Client")

window.minsize(900, 700)  # Prevent shrinking smaller than 900x700
window.state('zoomed')


client_pc_id = None
client_label = "Station: (connecting...)"
pc_id_label = None

def create_persistent_label():
    global pc_id_label
    pc_id_label = tk.Label(window, text=client_label, font=("Helvetica", 14, "bold"), bg="#5D3FD3", fg="white", padx=10, pady=5)
    pc_id_label.place(relx=1.0, rely=0.0, anchor="ne", x=-20, y=20)

def update_pc_label(label):
    global client_label, pc_id_label
    client_label = f"Station: {label}"
    if pc_id_label:
        pc_id_label.config(text=client_label)

decoder = json.JSONDecoder()
_recv_buf = ""
socket_lock = threading.Lock()
def recv_one_json(sock):
    global _recv_buf
    while True:
        _recv_buf = _recv_buf.lstrip()
        if _recv_buf:
            try:
                obj, idx = decoder.raw_decode(_recv_buf)
                _recv_buf = _recv_buf[idx:]
                return obj
            except json.JSONDecodeError:
                pass

        chunk = sock.recv(4096).decode("utf-8")
        if not chunk:
            raise ConnectionError("Server closed connection")
        _recv_buf += chunk

def toggle_fullscreen(event=None):
    is_fullscreen = window.attributes('-fullscreen')
    window.attributes('-fullscreen', not is_fullscreen)  # Toggle fullscreen


window.bind('<F11>', toggle_fullscreen)
window.grid_columnconfigure(0, weight=1)
window.grid_rowconfigure(0, weight=1)


def resize_image(event):
    new_width = max(event.width, 900)  # Prevent shrinking smaller than 900
    new_height = max(event.height, 700)  # Prevent shrinking smaller than 700
    resized_image = original_bg_image.resize((new_width, new_height), Image.Resampling.LANCZOS)

    global bg_image  # Keep reference to prevent garbage collection
    bg_image = ImageTk.PhotoImage(resized_image)
    canvas.create_image(0, 0, anchor=tk.NW, image=bg_image)
    canvas.image = bg_image  # Keep a reference


script_dir = os.path.dirname(os.path.abspath(__file__))
bg_path = os.path.join(script_dir, "bg.jpg")
if not os.path.exists(bg_path):
    print(f"⚠️ Warning: bg.jpg not found at {bg_path}")

original_bg_image = Image.open(bg_path)

canvas = tk.Canvas(window, width=900, height=700, highlightthickness=0)  # No border
canvas.grid(row=0, column=0, rowspan=4, columnspan=3, sticky="nsew")
canvas.bind("<Configure>", resize_image)
client_socket = None
student_name = None
student_section = None
quiz_number_var = tk.StringVar()
connected_server_ip = None


def discover_server():
    """Broadcasts a message to find the instructor server IP."""
    try:
        udp = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
        udp.setsockopt(socket.SOL_SOCKET, socket.SO_BROADCAST, 1)
        udp.settimeout(2)
        udp.sendto(b"DISCOVER_SERVER", ('<broadcast>', 9999))

        while True:
            data, addr = udp.recvfrom(1024)
            if data.decode().strip() == "I_AM_SERVER":
                return addr[0]
    except Exception as e:
        print(f"Discovery failed: {e}")
    return None


def connect_to_instructor():
    global client_socket, current_instructor, current_section, current_subject, connected_server_ip

    try:
        server_ip = discover_server()
        if not server_ip:
            messagebox.showerror("Connection Error", "Instructor Server not found!\nMake sure the server is running.")
            return

        connected_server_ip = server_ip

        client_socket = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
        client_socket.connect((server_ip, 9999))
        client_socket.settimeout(5)

        client_socket.sendall(json.dumps({"action": "REQUEST_INSTRUCTOR_INFO"}).encode("utf-8"))

        raw = client_socket.recv(4096).decode("utf-8").strip()
        instructor_info = json.loads(raw)

        current_instructor = instructor_info.get("instructor", "N/A")
        current_section = instructor_info.get("section", "N/A")
        current_subject = instructor_info.get("subject", "N/A")

        if current_instructor == "N/A" or current_section == "N/A" or current_subject == "N/A":
            messagebox.showwarning(
                "Instructor Not Ready",
                "Connected, but the instructor has not selected a section/subject yet.\n"
                "Ask your instructor to login and select section/subject, then try again."
            )
            client_socket.close()
            return

        print(f"✅ Connected to: {current_instructor} | {current_section} | {current_subject}")

        client_socket.settimeout(None)
        threading.Thread(target=listen_for_commands, daemon=True).start()
        show_sign_in_page()

    except Exception as e:
        print(f"❌ Failed to connect: {e}")
        messagebox.showerror("Connection Error", f"Error: {e}")



def show_landing_page():
    for widget in window.winfo_children():
        widget.destroy()

    global canvas
    canvas = tk.Canvas(window, width=900, height=700)
    canvas.grid(row=0, column=0, rowspan=4, columnspan=3, sticky="nsew")
    canvas.bind("<Configure>", resize_image)

    title_label = ttk.Label(window, text="Welcome to the Students Client of Computer Laboratory!",
                            style="Title.TLabel", anchor="center", background="white")
    title_label.place(relx=0.5, rely=0.05, anchor="n")  # Positioned at the top

    subtitle_label = ttk.Label(window, text="Empowering Future Educators",
                               font=("Helvetica", 18), background='#D1BAFF', foreground='white')
    subtitle_label.place(relx=0.5, rely=0.12, anchor="n")  # Below the title

    description_label = ttk.Label(window, text="Teacher's client is not yet open...\n\n"
                                               "Ask your teacher to open the Teacher Client server\n"
                                               "and sign in to the designated section.\n",
                                  style="Description.TLabel", anchor="center", justify="center", background="white")
    description_label.place(relx=0.5, rely=0.25, anchor="n")  # Below the subtitle

    connect_button = ttk.Button(window, text="Connect to Instructor Client", style="Curved.TButton",
                                command=connect_to_instructor)
    connect_button.place(relx=0.5, rely=0.40, anchor="n")  # Properly spaced below description


style.configure("Title.TLabel", font=("Helvetica", 24, "bold"), foreground='#6A329D')
style.configure("Description.TLabel", font=("Helvetica", 16), foreground='#6A329D')

style.configure("Curved.TButton",
                font=("Helvetica", 14, "bold"),
                background="white",
                foreground="black",
                borderwidth=2,
                relief="flat",
                padding=10,
                bordercolor="white",
                borderradius=15)

style.map("Curved.TButton",
          foreground=[('hover', 'white')],
          background=[('hover', '#3BB143')],  # Pink on hover
          bordercolor=[('hover', '#3BB143')])

show_landing_page()
timer = None
shutdown_popup = None


def show_shutdown_timer(minutes, action):
    global shutdown_timer, shutdown_popup

    shutdown_timer = minutes * 60
    shutdown_popup = tk.Toplevel()
    shutdown_popup.title(f"{action} Countdown")
    shutdown_popup.geometry("300x150")
    shutdown_popup.configure(bg="black")

    label = tk.Label(shutdown_popup, text=f"{action} in {minutes} minutes!", font=("Helvetica", 14, "bold"), fg="red",
                     bg="black")
    label.pack(pady=10)

    countdown_label = tk.Label(shutdown_popup, text=f"Time left: {minutes}:00", font=("Helvetica", 16), fg="white",
                               bg="black")
    countdown_label.pack(pady=10)

    def update_timer():
        global shutdown_timer
        if shutdown_timer <= 0:
            print(f"⏳ Countdown done. Proceeding with: {action}")
            shutdown_popup.destroy()
            if action == "Shutdown":
                print("🛑 Now calling shutdown_system()")
                shutdown_system()
            elif action == "Sleep":
                print("💤 Now calling sleep_system()")
                sleep_system()
        else:
            minutes_left = shutdown_timer // 60
            seconds_left = shutdown_timer % 60
            countdown_label.config(text=f"Time left: {minutes_left}:{seconds_left:02d}")
            shutdown_timer -= 1
            shutdown_popup.after(1000, update_timer)

    update_timer()


def shutdown_system():
    os.system('shutdown /s /t 1')


def sleep_system():
    os.system('rundll32.exe powrprof.dll,SetSuspendState 0,1,0')


def show_shutdown_timer(minutes, action):
    global shutdown_timer, shutdown_popup

    shutdown_timer = minutes * 60
    shutdown_popup = tk.Toplevel()
    shutdown_popup.title(f"{action} Timer")
    shutdown_popup.geometry("300x150")
    shutdown_popup.configure(bg="black")

    label = tk.Label(shutdown_popup, text=f"{action} in {minutes} minutes!", font=("Helvetica", 14), fg="red",
                     bg="black")
    label.pack(pady=10)

    countdown_label = tk.Label(shutdown_popup, text="", font=("Helvetica", 16), fg="white", bg="black")
    countdown_label.pack(pady=10)

    def update_timer():
        global shutdown_timer
        if shutdown_timer <= 0:
            shutdown_popup.destroy()
            if action.lower() == "shutdown":
                shutdown_system()
            elif action.lower() == "sleep":
                sleep_system()
        else:
            mins, secs = divmod(shutdown_timer, 60)
            countdown_label.config(text=f"Time left: {mins}:{secs:02d}")
            shutdown_timer -= 1
            shutdown_popup.after(1000, update_timer)

    update_timer()


def receive_json_with_retry(sock, retries=3, timeout=5):
    """Attempts to receive and decode JSON from the server, retrying on timeout."""
    for attempt in range(1, retries + 1):
        try:
            sock.settimeout(timeout)
            response = sock.recv(4096).decode()
            sock.settimeout(None)
            return json.loads(response)
        except socket.timeout:
            print(f"⚠️ Timeout attempt {attempt}/{retries}. Retrying...")
        except json.JSONDecodeError:
            print("❌ Failed to decode JSON response.")
            break
        except Exception as e:
            print(f"❌ Error during receive: {e}")
            break

    messagebox.showerror("Timeout", f"Failed to get response from server after {retries} attempts.")
    return None


def open_program(program):
    messagebox.showinfo("Start Activity", "Activity functionality is not yet implemented.")
    try:
        if program.upper() == "CHROME":
            chrome_path = "C:\\Program Files\\Google\\Chrome\\Application\\chrome.exe"
            if os.path.exists(chrome_path):
                subprocess.Popen([chrome_path])
            else:
                print("Chrome executable not found.")
        elif program.upper() == "ECLIPSE":
            eclipse_path = "C:\\path\\to\\eclipse.exe"
            subprocess.Popen([eclipse_path])
        elif program.upper() == "PYTHONIDE":
            python_ide_path = "C:\\path\\to\\pythonide.exe"
            subprocess.Popen([python_ide_path])
        else:
            print(f"No action defined for program: {program}")
    except Exception as e:
        print(f"Failed to open {program}: {e}")
import tkinter as tk
from tkinter import filedialog, messagebox, ttk

import socket
import json
import difflib
import matplotlib.pyplot as plt
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
submitted_activities = set()
timer_running = {"stop": False}
uploaded_file_path = {"path": None}
activity_started = {"status": False, "current_activity": None}
dropdown_mapping = {}
available_activities = {}   # ✅ add this

def send_request(payload: dict, timeout=8, bufsize=65536):
    target_ip = connected_server_ip or discover_server()
    if not target_ip:
        raise ConnectionError("Could not find server.")

    s = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
    s.settimeout(timeout)
    try:
        s.connect((target_ip, 9999))
        s.sendall(json.dumps(payload).encode("utf-8"))
        s.shutdown(socket.SHUT_WR)  # ✅ tell server we're done sending

        chunks = []
        while True:
            try:
                part = s.recv(bufsize)
            except socket.timeout:
                break
            if not part:
                break
            chunks.append(part)

        data = b"".join(chunks).decode("utf-8", errors="replace").strip()
        if not data:
            return {}

        return json.loads(data)

    finally:
        try:
            s.close()
        except:
            pass


def start_activity():
    global available_activities, timer_thread, activity_started
    activity_started["status"] = False
    activity_started["current_activity"] = None
    global start_button
    global pycharm_button, visualstudio_button, devcpp_button
    upload_button_is_visible = False
    submit_button_is_visible = False

    def prettify_runtime_output(out: str) -> str:
        if not out:
            return "(no output)"

        # If program asks for input() with no stdin, Python raises EOFError
        if "EOFError: EOF when reading a line" in out:
            return (
                "⚠️ This activity uses input().\n"
                "No test input was provided during grading, so Python stopped early (EOFError).\n"
                "Tip: Use the 'Run with Sample Input' box below to prove your code works."
            )

        # Hide huge traceback → show only last error line
        if "Traceback (most recent call last):" in out:
            lines = [ln.strip() for ln in out.splitlines() if ln.strip()]
            last = lines[-1] if lines else "Runtime error."
            return f"⚠️ Runtime error: {last}"

        return out

    def do_submit(selected_activity, student_code, language):
        resp = send_request({
            "action": "submit_activity",
            "student_name": student_name,
            "activity_key": selected_activity,
            "code": student_code,
            "instructor": current_instructor,
            "section": current_section,
            "subject": current_subject,
            "language": language
        })

        if isinstance(resp, dict) and resp.get("ok") is False:
            messagebox.showerror("Submission Failed", resp.get("error", "Unknown error"))
            return None

        return resp
    for widget in window.winfo_children():
        widget.destroy()

    window.title("Activity Management")

    # Load original background only once
    script_dir = os.path.dirname(os.path.abspath(__file__))
    bg_path = os.path.join(script_dir, "bg.jpg")
    original_bg_image = Image.open(bg_path)

    def resize_bg(event=None):
        new_width = window.winfo_width()
        new_height = window.winfo_height()

        resized_bg = original_bg_image.resize((new_width, new_height), Image.LANCZOS)
        bg_image = ImageTk.PhotoImage(resized_bg)

        background_label.config(image=bg_image)
        background_label.image = bg_image  # keep reference to prevent garbage collection

    background_label = tk.Label(window, bg="white")
    background_label.place(x=0, y=0, relwidth=1, relheight=1)
    resize_bg()
    window.bind("<Configure>", resize_bg)
    main_frame = tk.Frame(window, bg="white", bd=4, relief="ridge")
    main_frame.place(relx=0.5, rely=0.5, anchor="center", width=950, height=750)
    title = tk.Label(main_frame, text="Activity Management", font=("Helvetica", 16, "bold"), bg="white", fg="#5D3FD3")
    title.pack(pady=10)

    tk.Label(main_frame, text="Available Activities:", font=("Arial", 11), bg="white").pack()
    activity_dropdown = ttk.Combobox(main_frame, font=("Arial", 12), state="readonly", width=55)
    activity_dropdown.pack(pady=5)
    show_instructions_button = ttk.Button(
        main_frame,
        text="📘 Show Instructions",
        style="Curved.TButton",
        width=25,
        command=lambda: show_instructions(dropdown_mapping.get(activity_dropdown.get()))
    )
    show_instructions_button.pack(pady=5)

    instructions_frame = tk.Frame(main_frame, bg="white", relief="groove", bd=2)
    instructions_frame.pack(pady=10, padx=10, fill="both", expand=False)
    instruction_label = tk.Label(instructions_frame, text="📘 Instructions will appear here.",
                                 font=("Arial", 11), bg="white", wraplength=560, justify="left")
    instruction_label.pack(pady=5)

    instruction_image_label = tk.Label(instructions_frame, bg="white")
    instruction_image_label.pack()
    timer_display = tk.Label(main_frame, text="", bg="white", fg="green", font=("Arial", 11, "bold"))
    timer_display.pack(pady=(0, 5))

    def on_activity_select(event):
        selected = dropdown_mapping.get(activity_dropdown.get())
        print(f"[DEBUG] Selected activity: {selected}")
        show_instructions(selected)

        if selected in submitted_activities:
            start_button.pack_forget()
        else:
            if not start_button.winfo_ismapped():
                start_button.pack(pady=5)

    activity_dropdown.bind("<<ComboboxSelected>>", on_activity_select)
    from tkinter import messagebox

    def open_pycharm():
        try:
            subprocess.Popen([r"E:\PyCharm 2025.1\bin\pycharm64.exe"])
        except Exception as e:
            messagebox.showerror("Error", f"Could not open PyCharm:\n{e}")

    def open_visualstudio():
        try:
            subprocess.Popen([r"C:\Program Files\Microsoft Visual Studio\2022\Community\Common7\IDE\devenv.exe"])
        except Exception as e:
            messagebox.showerror("Error", f"Could not open Visual Studio:\n{e}")

    def open_java():
        try:
            subprocess.Popen([r"C:\Program Files\JetBrains\IntelliJ IDEA 2023.2\bin\idea64.exe"])

        except Exception as e:
            messagebox.showerror("Error", f"Could not open Java environment:\n{e}")

    def load_activities():
        def fetch():
            global available_activities
            try:
                resp = send_request({
                    "action": "get_activities",
                    "instructor": current_instructor,
                    "section": current_section,
                    "subject": current_subject
                })
                available_activities = resp if isinstance(resp, dict) else {}
                window.after(0, update_dropdown)
            except Exception as e:
                messagebox.showerror("Error", f"Activity fetch failed: {e}")

        threading.Thread(target=fetch, daemon=True).start()

    def update_dropdown():
        dropdown_mapping.clear()
        options = []

        if not isinstance(available_activities, dict):
            available_activities.clear()

        for key, info in available_activities.items():
            status = "✅ Done - You Submitted Already" if key in submitted_activities else \
                "✔ Accepting" if info.get("accepting") else "❌ Closed"

            display = f"{key} - {info.get('subject', '')} ({status})"
            dropdown_mapping[display] = key
            options.append(display)

        if not options:
            activity_dropdown["values"] = ["No activities available"]
            activity_dropdown.set("No activities available")
            start_button.pack_forget()
            return

        activity_dropdown["values"] = options
        activity_dropdown.set(options[0])

        selected_key = dropdown_mapping.get(options[0])
        if selected_key in submitted_activities:
            start_button.pack_forget()
        else:
            if not start_button.winfo_ismapped():
                start_button.pack(pady=5)
    activity_dropdown.bind("<<ComboboxSelected>>", on_activity_select)

    import base64
    from io import BytesIO

    def show_instructions(activity_key):
        global available_activities  # ✅ Ensure it's global to access activity info

        print(f"[DEBUG] show_instructions() called for activity: {activity_key}")
        activity_info = available_activities.get(activity_key)

        if not activity_info:
            print("[DEBUG] No activity info found.")
            return

        instruction_text = activity_info.get("instruction_text")
        instruction_image_path = activity_info.get("instruction_image")  # ✅ Now expecting file path
        time_limit = activity_info.get("time_limit")

        popup = tk.Toplevel(window)
        popup.title(f"Instructions - {activity_key}")
        popup.geometry("520x500")
        popup.configure(bg="white")

        # 🔹 Instruction Text
        if instruction_text:
            text_label = tk.Label(popup, text=f"📘 Instructions:\n{instruction_text}",
                                  font=("Arial", 11), bg="white", wraplength=480, justify="left")
            text_label.pack(pady=10, padx=10)
        else:
            text_label = tk.Label(popup, text="📘 No instruction text provided.",
                                  font=("Arial", 11), bg="white", wraplength=480, justify="left")
            text_label.pack(pady=10, padx=10)

        # 🔹 Instruction Image (File Path)
        if instruction_image_path and os.path.exists(instruction_image_path):
            try:
                img = Image.open(instruction_image_path)
                img.thumbnail((460, 250))
                img_tk = ImageTk.PhotoImage(img)
                img_label = tk.Label(popup, image=img_tk, bg="white")
                img_label.image = img_tk  # keep reference to prevent GC
                img_label.pack(pady=10)
            except Exception as e:
                err_label = tk.Label(popup, text=f"❌ Failed to load image.\n{e}", bg="white", fg="red")
                err_label.pack()
        elif instruction_image_path:
            err_label = tk.Label(popup, text="❌ Image file not found.", bg="white", fg="red")
            err_label.pack()

        # 🔹 Time Limit
        if time_limit:
            try:
                minutes = int(time_limit)
                tk.Label(popup, text=f"⏳ Time Limit: {minutes} minute(s)",
                         bg="white", fg="darkgreen", font=("Arial", 10, "italic")).pack()
                tk.Label(popup, text="(Timer starts when you confirm to begin)",
                         bg="white", fg="gray", font=("Arial", 9, "italic")).pack(pady=(0, 10))
            except ValueError:
                tk.Label(popup, text=f"⏳ Time Limit: {time_limit}",
                         bg="white", fg="darkgreen", font=("Arial", 10, "italic")).pack()

        # 🔹 Close Button
        ttk.Button(popup, text="Close", command=popup.destroy).pack(pady=10)

    def start_timer(seconds):
        timer_running["stop"] = False  # reset

        def safe_set(text):
            try:
                if timer_display.winfo_exists():
                    timer_display.config(text=text)
            except:
                pass

        def countdown():
            nonlocal seconds
            while seconds > 0 and not timer_running["stop"]:
                mins, secs = divmod(seconds, 60)
                window.after(0, safe_set, f"⏳ Time Left: {mins:02d}:{secs:02d}")
                time.sleep(1)
                seconds -= 1

            if timer_running["stop"]:
                window.after(0, safe_set, "✅ Activity submitted. Timer stopped.")
                return

            window.after(0, safe_set, "⏰ Time's up! Auto-submitting...")
            window.after(0, lambda: upload_code(auto=True))

        threading.Thread(target=countdown, daemon=True).start()

    def confirm_start_activity():
        selected = dropdown_mapping.get(activity_dropdown.get())

        if selected in submitted_activities:
            messagebox.showinfo("Already Submitted", f"You've already submitted '{selected}'.")
            return

        if not selected or selected == "No activities available":
            messagebox.showerror("Error", "Please select an activity.")
            return

        if activity_started["status"]:
            messagebox.showwarning("Already Started",
                                   f"Activity '{activity_started['current_activity']}' is already in progress.")
            return

        confirm = messagebox.askyesno("Confirm Start", f"Start activity '{selected}' now?")
        if confirm:
            activity_started["status"] = True
            activity_started["current_activity"] = selected

            info = available_activities.get(selected)
            if info and info.get("time_limit"):
                try:
                    minutes = int(info["time_limit"])
                    seconds = minutes * 60
                    start_timer(seconds)
                except:
                    timer_display.config(text="⚠ Invalid time limit.")

            else:
                timer_display.config(text="⏳ No timer for this activity.")

            # ✅ Show buttons after starting
            upload_button.pack(pady=5)
            submit_button.pack(pady=(5, 10))
            pycharm_button.pack(side="left", padx=4)
            visualstudio_button.pack(side="left", padx=4)
            devcpp_button.pack(side="left", padx=4)

            messagebox.showinfo("Activity Started", f"✅ You may now upload your code for: {selected}")

    def normalize_code(code):
        """Normalize code by stripping extra spaces and unifying indentation."""
        return "\n".join(line.strip() for line in code.splitlines() if line.strip())

    def get_instructor_code(activity_key):
        try:
            resp = send_request({"action": "get_activity_code", "activity_key": activity_key})
            return resp.get("code", "") if isinstance(resp, dict) else ""
        except Exception as e:
            print(f"❌ Error fetching instructor code: {e}")
            return ""

    def upload_code(auto=False, file_path=None):
        if not activity_started["status"]:
            messagebox.showwarning("Warning", "Please start the activity first.")
            return

        selected_activity = dropdown_mapping.get(activity_dropdown.get())
        if not selected_activity:
            messagebox.showerror("Error", "Please select a valid activity.")
            return

        info = available_activities.get(selected_activity, {})
        language = info.get("language", "Python") if isinstance(info, dict) else "Python"

        # ----------------------------
        # 1) Manual file pick (no submit yet)
        # ----------------------------
        if not auto:
            if language.lower() == "java":
                filetypes = [("Java Files", "*.java")]
            elif language.lower() in ["c", "c++"]:
                filetypes = [("C/C++ Files", "*.c *.cpp")]
            elif language.lower() in ["text", "plaintext"]:
                filetypes = [("Text Files", "*.txt")]
            else:
                filetypes = [
                    ("All Supported Code Files", "*.py *.java *.cpp *.c *.txt"),
                    ("Python Files", "*.py"),
                    ("Java Files", "*.java"),
                    ("C/C++ Files", "*.cpp *.c"),
                    ("Text Files", "*.txt"),
                    ("All Files", "*.*")
                ]

            file_path = filedialog.askopenfilename(title="Select Your Code File", filetypes=filetypes)
            if not file_path:
                messagebox.showerror("Error", "File not selected.")
                return

            uploaded_file_path["path"] = file_path
            uploaded_path_label.config(text=f"📂 File Selected: {file_path}")

            if not submit_button.winfo_ismapped():
                submit_button.pack(pady=(5, 10))
            return  # wait for manual submit button

        # ----------------------------
        # 2) Auto-submit (timer) OR you called upload_code(auto=True)
        # ----------------------------
        if selected_activity in submitted_activities:
            messagebox.showwarning("Already Submitted", "You have already submitted this activity.")
            return

        if not file_path:
            file_path = uploaded_file_path.get("path")
            if not file_path:
                messagebox.showerror("Error", "No file selected.")
                return

        try:
            with open(file_path, "r", encoding="utf-8", errors="ignore") as f:
                student_code = f.read()

            import re
            student_code = re.sub(r"[\u2028\u2029]", "", student_code)

            timer_running["stop"] = True

            # --- SUBMIT (STATELESS) ---
            response_data = do_submit(selected_activity, student_code, language)
            if not response_data:
                return

            if isinstance(response_data, dict) and response_data.get("ok") is False:
                messagebox.showerror("Submission Failed", response_data.get("error", "Unknown error"))
                return

            similarity_score = float(response_data.get("similarity", 0.0))

            submitted_activities.add(selected_activity)
            activity_started["status"] = False
            activity_started["current_activity"] = None

            # Hide buttons
            start_button.pack_forget()
            start_button.config(state="disabled")
            upload_button.pack_forget()
            upload_button.config(state="disabled")
            submit_button.pack_forget()
            pycharm_button.pack_forget()
            visualstudio_button.pack_forget()
            devcpp_button.pack_forget()

            update_dropdown()

            instructor_code = get_instructor_code(selected_activity)
            show_detailed_score(similarity_score, selected_activity, student_code, instructor_code, language)

        except Exception as e:
            messagebox.showerror("Error", f"Failed to submit activity: {e}")
    def submit_activity():
        # Do NOT reset activity_started until submission is confirmed successful
        selected_activity = dropdown_mapping.get(activity_dropdown.get())
        if not selected_activity or selected_activity == "No activities available":
            messagebox.showerror("Error", "Please select an activity.")
            return

        if selected_activity in submitted_activities:
            messagebox.showwarning("Already Submitted", "You have already submitted this activity.")
            return

        file_path = uploaded_file_path.get("path")
        if not file_path:
            messagebox.showerror("Error", "No file uploaded.")
            return

        try:
            with open(file_path, "r", encoding="utf-8", errors="ignore") as f:
                student_code = f.read()

            # sanitize unicode that can break JSON
            import re
            student_code = re.sub(r"[\u2028\u2029]", "", student_code)

            # stop timer
            timer_running["stop"] = True

            # detect language based on activity info
            info = available_activities.get(selected_activity, {})
            language = info.get("language", "Python") if isinstance(info, dict) else "Python"

            # --- SUBMIT (STATELESS) ---
            response_data = do_submit(selected_activity, student_code, language)
            if not response_data:
                return  # do_submit already showed error popup

            # If server returns ok:false
            if isinstance(response_data, dict) and response_data.get("ok") is False:
                messagebox.showerror("Submission Failed", response_data.get("error", "Unknown error"))
                return

            # --- SUCCESS PATH ---
            similarity_score = float(response_data.get("similarity", 0.0))

            submitted_activities.add(selected_activity)
            activity_started["status"] = False
            activity_started["current_activity"] = None

            # Hide buttons after submission
            start_button.pack_forget()
            start_button.config(state="disabled")
            upload_button.pack_forget()
            submit_button.pack_forget()
            pycharm_button.pack_forget()
            visualstudio_button.pack_forget()
            devcpp_button.pack_forget()

            update_dropdown()

            # Show detailed scoring window (optional)
            instructor_code = get_instructor_code(selected_activity)
            show_detailed_score(similarity_score, selected_activity, student_code, instructor_code, language)

        except Exception as e:
            messagebox.showerror("Error", f"Failed to submit activity: {e}")

    def execute_code(code, input_data=""):
        try:
            process = subprocess.run(
                ["python", "-c", code],
                capture_output=True, text=True,
                timeout=3,
                input=input_data if input_data is not None else ""
            )
            return process.stdout.strip() if process.returncode == 0 else process.stderr.strip()
        except subprocess.TimeoutExpired:
            return "Execution Timed Out ⚠️"
        except Exception as e:
            return f"Execution Error: {str(e)}"

    def show_detailed_score(similarity_score, activity_name, student_code, instructor_code, language):
        try:
            similarity_score = float(similarity_score)
        except:
            similarity_score = 0.0

        # helper: badge
        def badge_for(score):
            if score >= 90: return ("✅ EXCELLENT", "green")
            if score >= 70: return ("🟡 GOOD", "#c58f00")
            if score >= 50: return ("🟠 FAIR", "#d46b08")
            return ("❌ NEEDS IMPROVEMENT", "red")

        # pull test_input
        info = available_activities.get(activity_name, {}) if isinstance(available_activities, dict) else {}
        test_input = info.get("test_input", "") if isinstance(info, dict) else ""

        explanation, student_output, grading_mode = analyze_output_differences(
            student_code=student_code,
            instructor_code=instructor_code,
            match_percentage=similarity_score,
            language=language,
            test_input=test_input
        )

        badge_text, badge_color = badge_for(similarity_score)

        score_window = tk.Toplevel(window)
        score_window.title(f"Score Report - {activity_name}")
        score_window.geometry("920x780")
        score_window.configure(bg="white")

        # ✅ ROOT GRID (CENTER EVERYTHING)
        score_window.grid_rowconfigure(0, weight=1)
        score_window.grid_columnconfigure(0, weight=1)

        # ✅ Center container (all content lives here)
        outer = tk.Frame(score_window, bg="white")
        outer.grid(row=0, column=0, sticky="nsew")

        outer.grid_rowconfigure(0, weight=0)  # header
        outer.grid_rowconfigure(1, weight=0)  # chart
        outer.grid_rowconfigure(2, weight=1)  # notebook expands
        outer.grid_rowconfigure(3, weight=0)  # footer buttons
        outer.grid_columnconfigure(0, weight=1)

        # ===== HEADER (CENTERED) =====
        header = tk.Frame(outer, bg="white")
        header.grid(row=0, column=0, sticky="ew", padx=18, pady=(16, 8))
        header.grid_columnconfigure(0, weight=1)

        tk.Label(
            header,
            text=f"📄 Score Report: {activity_name}",
            bg="white",
            font=("Segoe UI", 36, "bold")
        ).grid(row=0, column=0, sticky="n", pady=(0, 8))

        # Score row centered
        score_row = tk.Frame(header, bg="white")
        score_row.grid(row=1, column=0)

        tk.Label(
            score_row, text=f"{similarity_score:.2f}%",
            bg="white", fg="#16a34a",
            font=("Segoe UI", 32, "bold")
        ).pack(side="left", padx=(0, 12))

        tk.Label(
            score_row, text=badge_text,
            bg="white", fg=badge_color,
            font=("Segoe UI", 13, "bold")
        ).pack(side="left")

        tk.Label(
            header,
            text=grading_mode,
            bg="white", fg="gray",
            font=("Segoe UI", 9, "italic"),
            wraplength=860, justify="center"
        ).grid(row=2, column=0, pady=(6, 0))

        # ===== CHART (CENTERED) =====
        chart_frame = tk.Frame(outer, bg="white")
        chart_frame.grid(row=1, column=0, pady=(0, 10))
        chart_frame.grid_columnconfigure(0, weight=1)

        fig, ax = plt.subplots(figsize=(4.0, 3.0))
        fig.patch.set_facecolor("white")
        values = [similarity_score, max(0.0, 100.0 - similarity_score)]
        ax.pie(values, labels=["Matched", "Unmatched"], autopct="%1.0f%%", startangle=90)
        ax.set_title("Output Similarity")

        canvas = FigureCanvasTkAgg(fig, master=chart_frame)
        canvas.draw()
        canvas.get_tk_widget().grid(row=0, column=0)  # ✅ grid centers naturally

        # ===== NOTEBOOK (CENTERED + EXPANDS) =====
        notebook = ttk.Notebook(outer)
        notebook.grid(row=2, column=0, sticky="nsew", padx=18, pady=(0, 12))

        tab_output = tk.Frame(notebook, bg="white")
        tab_run = tk.Frame(notebook, bg="white")
        tab_notes = tk.Frame(notebook, bg="white")

        for t in (tab_output, tab_run, tab_notes):
            t.grid_columnconfigure(0, weight=1)

        notebook.add(tab_output, text="Output Preview")
        notebook.add(tab_run, text="Run with Sample Input")
        notebook.add(tab_notes, text="Notes")

        # --- Output tab ---
        tk.Label(tab_output, text="Your Program Output (preview):",
                 bg="white", font=("Segoe UI", 12, "bold")).grid(row=0, column=0, sticky="w", padx=12, pady=(12, 6))

        out_box = tk.Text(tab_output, height=12, font=("Consolas", 10))
        out_box.grid(row=1, column=0, sticky="nsew", padx=12, pady=(0, 12))
        tab_output.grid_rowconfigure(1, weight=1)
        out_box.insert("1.0", student_output or "(no output)")
        out_box.config(state="disabled")

        # --- Run tab ---
        tk.Label(tab_run, text="Paste sample input here (one line per input()):",
                 bg="white", font=("Segoe UI", 11, "bold")).grid(row=0, column=0, sticky="w", padx=12, pady=(12, 6))

        sample_box = tk.Text(tab_run, height=6, font=("Consolas", 10))
        sample_box.grid(row=1, column=0, sticky="ew", padx=12, pady=(0, 8))
        sample_box.insert("1.0", test_input or "")

        tk.Label(tab_run, text="Run result:",
                 bg="white", font=("Segoe UI", 11, "bold")).grid(row=2, column=0, sticky="w", padx=12, pady=(0, 6))

        run_out = tk.Text(tab_run, height=10, font=("Consolas", 10))
        run_out.grid(row=3, column=0, sticky="nsew", padx=12, pady=(0, 10))
        tab_run.grid_rowconfigure(3, weight=1)

        def run_with_sample():
            sample = sample_box.get("1.0", "end").rstrip("\n")
            sample = (sample + "\n") if sample else ""
            out = execute_code(student_code, input_data=sample)
            out = prettify_runtime_output(out)
            run_out.delete("1.0", "end")
            run_out.insert("1.0", out)

        def copy_run_output():
            txt = run_out.get("1.0", "end").strip()
            score_window.clipboard_clear()
            score_window.clipboard_append(txt)
            messagebox.showinfo("Copied", "Run output copied to clipboard.")

        btn_row = tk.Frame(tab_run, bg="white")
        btn_row.grid(row=4, column=0, pady=(0, 12))
        ttk.Button(btn_row, text="▶ Run", style="Primary.TButton", command=run_with_sample).pack(side="left", padx=6)
        ttk.Button(btn_row, text="📋 Copy Output", style="Secondary.TButton", command=copy_run_output).pack(side="left",
                                                                                                           padx=6)

        # --- Notes tab ---
        tk.Label(tab_notes, text="Feedback:",
                 bg="white", font=("Segoe UI", 11, "bold")).grid(row=0, column=0, sticky="w", padx=12, pady=(12, 6))

        notes_lbl = tk.Label(tab_notes, text=explanation,
                             bg="white", justify="left", wraplength=860)
        notes_lbl.grid(row=1, column=0, sticky="nw", padx=12, pady=(0, 12))
        tab_notes.grid_rowconfigure(1, weight=1)

        # ===== FOOTER (CENTERED BUTTON) =====
        footer = tk.Frame(outer, bg="white")
        footer.grid(row=3, column=0, pady=(0, 16))
        ttk.Button(footer, text="Close", style="Primary.TButton", command=score_window.destroy).pack()

    import tempfile
    import subprocess

    def analyze_output_differences(student_code, instructor_code, match_percentage, language="Python", test_input=""):
        # normalize match_percentage
        try:
            mp = float(match_percentage or 0.0)
        except:
            mp = 0.0

        # 1) run student code (best-effort) using test_input (can be blank)
        raw_out = ""
        if str(language).lower() == "java":
            # If you don't support stdin for Java right now, just run without input
            # (You can extend later.)
            raw_out = "⚠️ Java runtime preview not enabled in student client yet."
        else:
            stdin_text = (test_input or "")
            stdin_text = stdin_text.rstrip("\n")
            stdin_text = (stdin_text + "\n") if stdin_text else ""
            raw_out = execute_code(student_code, input_data=stdin_text)

        student_output = prettify_runtime_output(raw_out)

        # 2) explanation text
        if mp >= 90:
            explanation = "✅ Excellent! Your output is very close to the expected result."
        elif mp >= 70:
            explanation = "🟡 Good! Your output is mostly correct but has minor differences."
        elif mp >= 50:
            explanation = "🔸 Fair. Some correct logic, but noticeable output differences."
        else:
            explanation = "❌ Needs Improvement. Your output differs significantly from the expected output."

        # 3) grading mode message (helps avoid confusion)
        if "EOFError" in raw_out:
            grading_mode = "⚠️ Fallback grading: input() detected but no test input was provided during grading."
            explanation += "\n\n⚠️ Your code uses input(). Use the Sample Input box to show it runs correctly."
        else:
            grading_mode = "✅ Output-based grading: ran your code to preview output."

        return explanation, student_output, grading_mode


    def get_student_score():
        if not student_name or not student_section or not student_subject:
            messagebox.showerror("Error", "You must be signed in to view your scores.")
            return

        try:
            scores_data = send_request({
                "action": "get_scores",
                "student_name": student_name,
                "section": student_section,
                "subject": student_subject
            })

            if isinstance(scores_data, dict) and "error" not in scores_data:
                scores_text = "\n".join([f"{activity}: {score}" for activity, score in scores_data.items()])
                messagebox.showinfo("My Scores",
                                    f"Your Activity Scores:\n\n{scores_text or 'No activity scores found.'}")
            else:
                messagebox.showerror("Error", scores_data.get("error", "Unexpected response."))

        except Exception as e:
            messagebox.showerror("Error", f"Failed to fetch scores: {e}")
    def back_to_main():
        """Go back to the student landing page."""
        show_student_landing_page(student_name, student_section)

    # **Buttons with Uniform Size & Hover Effects**
    button_width = 25
    start_button = ttk.Button(main_frame, text="Start Activity", style="Curved.TButton",
                              width=button_width, command=confirm_start_activity)

    upload_button = ttk.Button(main_frame, text="Upload Code", style="Curved.TButton",
                               width=25, command=upload_code)
    upload_button_is_visible = False  # track if already packed

    submit_button = ttk.Button(main_frame, text="Submit Activity", style="Curved.TButton",
                               width=25, command=submit_activity)

    uploaded_path_label = tk.Label(main_frame, text="", bg="white", font=("Arial", 10), wraplength=500)
    uploaded_path_label.pack()

    load_activities()

    score_button = ttk.Button(main_frame, text="View My Score", style="Curved.TButton",
                              width=button_width, command=get_student_score)
    score_button.pack(pady=5)

    back_button = ttk.Button(main_frame, text="Back", style="Curved.TButton",
                             command=lambda: show_student_landing_page(student_name, student_section, student_subject))
    back_button.pack(pady=10)
    # 🧠 IDE Button Row (Horizontal & Resized)
    ide_frame = tk.Frame(main_frame, bg="white")
    ide_frame.pack(pady=(10, 5))
    btn_width = 17

    pycharm_button = ttk.Button(ide_frame, text="Launch PyCharm", style="Curved.TButton",
                                width=btn_width, command=open_pycharm)
    pycharm_button.pack(side="left", padx=4)

    visualstudio_button = ttk.Button(ide_frame, text="Launch VisualStudio", style="Curved.TButton",
                                     width=btn_width, command=open_visualstudio)
    visualstudio_button.pack(side="left", padx=4)

    devcpp_button = ttk.Button(ide_frame, text="Launch Dev C++", style="Curved.TButton",
                               width=btn_width, command=open_java)
    devcpp_button.pack(side="left", padx=4)



import sqlite3

conn = sqlite3.connect(os.path.join(os.path.abspath(os.path.dirname(__file__)), 'teacher_management.db'))
c = conn.cursor()

# Check current table structure
c.execute("PRAGMA table_info(students)")
columns = c.fetchall()
conn.close()

for col in columns:
    print(col)

import sqlite3

conn = sqlite3.connect(os.path.join(os.path.abspath(os.path.dirname(__file__)), 'teacher_management.db'))
c = conn.cursor()

# Add the subject column if it doesn't exist
try:
    c.execute("ALTER TABLE students ADD COLUMN subject TEXT")
    conn.commit()
    print("✅ 'subject' column added successfully!")
except sqlite3.OperationalError:
    print("⚠️ Column 'subject' already exists!")

conn.close()


def tap_rfid(rfid):
    if rfid:
        send_rfid(rfid)
    else:
        messagebox.showerror("Input Error", "Please enter your RFID.")


def connect_to_server():
    global client_socket, connected_server_ip
    try:
        target_ip = connected_server_ip if connected_server_ip else discover_server()
        if not target_ip:
            print("Server not found")
            return

        client_socket = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
        client_socket.connect((target_ip, 9999))
        print(f"Connected to server at {target_ip}")
        
        import uuid
        mac = ':'.join(['{:02x}'.format((uuid.getnode() >> ele) & 0xff) for ele in range(0,8*6,8)][::-1])
        hello_req = {
            "action": "HELLO_PC",
            "hostname": socket.gethostname(),
            "mac_address": mac
        }
        client_socket.sendall(json.dumps(hello_req).encode('utf-8'))
    except Exception as e:
        print(f"Connection Error: {e}")
        messagebox.showerror("Connection Error", f"Failed to connect to server: {e}")


def sign_in_student(rfid):
    global student_name, student_section, student_subject
    global current_section, current_subject  # <-- use instructor’s selection

    try:
        request = {
            "action": "signin",
            "rfid": rfid
        }

        client_socket.sendall(json.dumps(request).encode('utf-8'))
        response_data = client_socket.recv(4096).decode('utf-8')
        response = json.loads(response_data)

        if response["status"] == "success":
            # Original DB student info
            student_name = response["name"]
            orig_section = response["section"]
            orig_subject = response["subject"]

            print(f"📘 DEBUG: DB Student -> {student_name}, Section: {orig_section}, Subject: {orig_subject}")

            # 🔥 FIX: Force section/subject to match instructor-selected values
            if current_section and current_subject:
                student_section = current_section
                student_subject = current_subject
                print(
                    f"🎯 FIX APPLIED: Using Instructor-selected -> Section={student_section}, Subject={student_subject}")
            else:
                # fallback (should not happen unless instructor not chosen)
                student_section = orig_section
                student_subject = orig_subject
                print("⚠ Instructor selection missing — using DB student values")

            # Proceed to student page
            show_student_landing_page(student_name, student_section, student_subject)

        else:
            messagebox.showerror("Access Denied", response["message"])

    except Exception as e:
        print(f"❌ Error during sign-in: {e}")
        messagebox.showerror("Connection Error", f"Failed to contact server.\n\n{e}")


session_data = {}


def student_sign_in():
    global student_name, student_section, student_subject  # Ensure global usage

    # Simulating a successful login
    student_name = "test"
    student_section = "BSIT 1-1"
    student_subject = "aaa"

    print(f"✅ DEBUG: Successfully signed in as {student_name}, Section: {student_section}, Subject: {student_subject}")


def show_attendance_logs():
    """Fetch and display attendance logs for the currently signed-in student."""
    global student_name, student_section, student_subject, client_socket, connected_server_ip

    if not student_name or not student_section:
        messagebox.showerror("Error", "You must sign in before viewing attendance logs.")
        return

    # ── Build the page ────────────────────────────────────────────────
    for widget in window.winfo_children():
        widget.destroy()

    script_dir = os.path.dirname(os.path.abspath(__file__))
    bg_path = os.path.join(script_dir, "bg.jpg")
    original_bg_image = Image.open(bg_path)

    bg_label = tk.Label(window, bg="white")
    bg_label.place(x=0, y=0, relwidth=1, relheight=1)

    def resize_bg(event=None):
        w, h = window.winfo_width(), window.winfo_height()
        img = original_bg_image.resize((w, h), Image.LANCZOS)
        photo = ImageTk.PhotoImage(img)
        bg_label.config(image=photo)
        bg_label.image = photo

    resize_bg()
    window.bind("<Configure>", resize_bg)

    # ── Outer card frame ──────────────────────────────────────────────
    card = tk.Frame(window, bg="white", bd=2, relief="groove")
    card.place(relx=0.5, rely=0.5, anchor="center", width=1100, height=700)

    tk.Label(card, text="My Attendance Logs",
             font=("Helvetica", 20, "bold"), fg="#5D3FD3", bg="white").pack(pady=(16, 4))
    tk.Label(card, text=f"Student: {student_name}  |  Section: {student_section}  |  Subject: {student_subject}",
             font=("Helvetica", 12), fg="#444444", bg="white").pack(pady=(0, 10))

    # ── Treeview ──────────────────────────────────────────────────────
    tree_frame = tk.Frame(card, bg="white")
    tree_frame.pack(fill="both", expand=True, padx=20, pady=(0, 8))

    cols = ("Date", "Time In", "Time Out", "Instructor", "Subject", "Section")
    atd_style = ttk.Style()
    atd_style.configure("Att.Treeview", font=("Helvetica", 13), rowheight=32)
    atd_style.configure("Att.Treeview.Heading", font=("Helvetica", 13, "bold"))
    atd_style.configure(
        "Att.Treeview",
        background="#111111",
        foreground="white",
        fieldbackground="#111111",
        rowheight=32,
        font=("Helvetica", 13)
    )


    tree = ttk.Treeview(tree_frame, columns=cols, show="headings", style="Att.Treeview")
    col_widths = {"Date": 110, "Time In": 100, "Time Out": 100, "Instructor": 200, "Subject": 160, "Section": 140}
    for col in cols:
        tree.heading(col, text=col, anchor="center")
        tree.column(col, anchor="center", width=col_widths.get(col, 120))

    sb = ttk.Scrollbar(tree_frame, orient="vertical", command=tree.yview)
    sb.pack(side="right", fill="y")
    tree.configure(yscrollcommand=sb.set)
    tree.pack(fill="both", expand=True)

    # ── Status label & Back button ─────────────────────────────────────
    status_lbl = tk.Label(card, text="", font=("Helvetica", 11), fg="#888888", bg="white")
    status_lbl.pack(pady=(0, 4))

    ttk.Button(card, text="Back", style="Orange.TButton",
               command=lambda: show_student_landing_page(
                   student_name, student_section, student_subject)).pack(pady=(0, 14))

    # ── Fetch logs from server ─────────────────────────────────────────
    def fetch_logs():
        try:
            target_ip = connected_server_ip if connected_server_ip else discover_server()
            if not target_ip:
                window.after(0, lambda: status_lbl.config(text="⚠ Could not find instructor server."))
                return

            tmp = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
            tmp.settimeout(8)
            tmp.connect((target_ip, 9999))
            req = json.dumps({
                "action": "get_attendance_logs",
                "student_name": student_name,
                "section": student_section,
                "subject": student_subject
            })
            tmp.sendall(req.encode("utf-8"))

            # Receive full response (may come in chunks)
            raw = b""
            while True:
                chunk = tmp.recv(65536)
                if not chunk:
                    break
                raw += chunk
                try:
                    json.loads(raw.decode("utf-8"))
                    break            # valid JSON received
                except json.JSONDecodeError:
                    continue        # wait for more data
            tmp.close()

            data = json.loads(raw.decode("utf-8"))

            def update_ui():
                tree.delete(*tree.get_children())
                if data.get("status") == "ok":
                    logs = data.get("logs", [])
                    if logs:
                        for i, entry in enumerate(logs):
                            tag = "even" if i % 2 == 0 else "odd"
                            tree.insert("", "end", values=(
                                entry.get("date", ""),
                                entry.get("time_in", ""),
                                entry.get("time_out", ""),
                                entry.get("instructor", ""),
                                entry.get("subject", ""),
                                entry.get("section", "")
                            ), tags=(tag,))
                        tree.tag_configure(
                            "even",
                            background="#111111",
                            foreground="white"
                        )

                        tree.tag_configure(
                            "odd",
                            background="#1A1A1A",
                            foreground="white"
                        )
                        status_lbl.config(text=f"✅ {len(logs)} attendance record(s) found.")
                    else:
                        status_lbl.config(text="No attendance records found for your account.")
                else:
                    status_lbl.config(text=f"⚠ {data.get('message', 'Unknown error')}")

            window.after(0, update_ui)

        except Exception as e:
            window.after(0, lambda: status_lbl.config(text=f"⚠ Failed to load logs: {e}"))

    status_lbl.config(text="Loading attendance logs…")
    threading.Thread(target=fetch_logs, daemon=True).start()



def show_sign_in_page():
    global current_instructor, current_section, current_subject

    for widget in window.winfo_children():
        widget.destroy()

    # ✅ Load the original image
    script_dir = os.path.dirname(os.path.abspath(__file__))
    bg_path = os.path.join(script_dir, "bg.jpg")
    original_image = Image.open(bg_path)

    # ✅ Function to Resize and Update Image when Window Resizes
    def resize_background(event=None):
        width = window.winfo_width()
        height = window.winfo_height()

        # ✅ Resize the image dynamically to match window size
        resized_image = original_image.resize((width, height), Image.LANCZOS)
        bg_image = ImageTk.PhotoImage(resized_image)

        # ✅ Update background label
        background_label.config(image=bg_image)
        background_label.image = bg_image  # Prevent garbage collection

    # ✅ Initial image setup
    global background_label
    bg_image = ImageTk.PhotoImage(original_image)
    background_label = tk.Label(window, image=bg_image, bg="white")

    background_label.image = bg_image
    background_label.place(x=0, y=0, relwidth=1, relheight=1)

    # ✅ Bind window resize event to update the background
    window.bind("<Configure>", resize_background)

    # ✅ Title Label
    title_label = ttk.Label(window, text="Student Log-In Page", style="Title.TLabel", anchor="center",
                            background="white")
    title_label.pack(pady=20)

    # ✅ Show Instructor Info
    instructor_info_label = ttk.Label(
        window,
        text=f"Instructor: {current_instructor}\nSection: {current_section}\nSubject: {current_subject}",
        font=("Helvetica", 14),
        background="white",
        foreground="#8A2BE2",
        anchor="center"
    )
    instructor_info_label.pack(pady=10)

    # ✅ Instruction Label
    sign_in_label = ttk.Label(window, text="Sign in by tapping your RFID.\n\n"
                                           "You can also enter manually and press Enter.",
                              style="SignIn.TLabel", anchor="center", background="white", justify="center")
    sign_in_label.pack(pady=20)

    # ✅ RFID Entry Field
    rfid_entry = ttk.Entry(window, font=("Helvetica", 14), width=30)
    rfid_entry.pack(pady=10)
    rfid_entry.focus_set()
    rfid_entry.bind("<Return>", lambda event: sign_in_student(rfid_entry.get()))

    # ✅ Sign-In Button with Bootstrap Styling
    sign_in_button = ttk.Button(window, text="Sign Student", style="Curved.TButton",
                                command=lambda: sign_in_student(rfid_entry.get().strip()))
    sign_in_button.pack(pady=10)


# ✅ Style Configuration for Labels & Buttons
style.configure("Title.TLabel", font=("Helvetica", 24, "bold"), foreground='#8A2BE2',
                background="white")  # Violet title
style.configure("SignIn.TLabel", font=("Helvetica", 14), foreground='#8A2BE2',
                background="white")  # Violet font for instruction
style.configure("Motto.TLabel", font=("Arial", 12), foreground='#8A2BE2', background="white")  # Violet font for motto

style.configure("Curved.TButton",
                font=("Helvetica", 14, "bold"),
                padding=10,
                relief="flat",
                background="white",
                borderwidth=2,
                bordercolor="white",
                borderradius=15)

style.map("Curved.TButton",
          foreground=[('hover', 'white')],
          background=[('hover', '#FF69B4')],  # Pink on hover
          bordercolor=[('hover', '#FF69B4')])

import threading


def send_rfid(rfid, context="signin"):
    global client_socket
    try:
        if not client_socket:
            connect_to_server()

        request = {"action": context, "rfid": rfid}
        client_socket.sendall(json.dumps(request).encode())

        response = client_socket.recv(1024).decode()
        response_data = json.loads(response)

        if "message" in response_data:
            messagebox.showinfo("Response", response_data["message"])
        else:
            messagebox.showerror("Error", "Unexpected response from server.")

    except Exception as e:
        messagebox.showerror("Error", f"Failed to send RFID: {str(e)}")


def reconnect_to_server():
    global client_socket
    try:
        client_socket.close()
    except:
        pass
    connect_to_server()


def logout():
    show_sign_in_page()


import sqlite3

def ensure_student_quiz_db():
    conn = sqlite3.connect(os.path.join(os.path.abspath(os.path.dirname(__file__)), 'student_quiz_scores.db'))
    c = conn.cursor()

    # if table doesn't exist, create the new one
    c.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='quiz_scores'")
    exists = c.fetchone() is not None

    if not exists:
        c.execute("""
                  CREATE TABLE quiz_scores
                  (
                      id           INTEGER PRIMARY KEY AUTOINCREMENT,
                      student_name TEXT    NOT NULL,
                      instructor   TEXT    NOT NULL,
                      section      TEXT    NOT NULL,
                      subject      TEXT    NOT NULL,
                      quiz_number  INTEGER NOT NULL,
                      score        REAL    NOT NULL,
                      date_taken   TEXT    NOT NULL
                  )
                  """)
        conn.commit()
        conn.close()
        return

    # table exists -> check columns
    c.execute("PRAGMA table_info(quiz_scores)")
    cols = [row[1] for row in c.fetchall()]

    needed = {"student_name", "instructor", "section", "subject", "quiz_number", "score", "date_taken"}

    # if missing new columns, migrate
    if not needed.issubset(set(cols)):
        # rename old table
        c.execute("ALTER TABLE quiz_scores RENAME TO quiz_scores_old")

        # create new correct table
        c.execute("""
                  CREATE TABLE quiz_scores
                  (
                      id           INTEGER PRIMARY KEY AUTOINCREMENT,
                      student_name TEXT    NOT NULL,
                      instructor   TEXT    NOT NULL,
                      section      TEXT    NOT NULL,
                      subject      TEXT    NOT NULL,
                      quiz_number  INTEGER NOT NULL,
                      score        REAL    NOT NULL,
                      date_taken   TEXT    NOT NULL
                  )
                  """)

        # copy what we can from old (best-effort)
        # old table probably has: student_name, quiz_number, score, subject, section, date_taken
        # instructor might not exist -> fill "N/A"
        try:
            c.execute("""
                      INSERT INTO quiz_scores (student_name, instructor, section, subject, quiz_number, score,
                                               date_taken)
                      SELECT student_name,
                             'N/A' as instructor,
                             COALESCE(section, 'N/A'),
                             COALESCE(subject, 'N/A'),
                             quiz_number,
                             score,
                             date_taken
                      FROM quiz_scores_old
                      """)
        except Exception as e:
            print("⚠️ Migration copy skipped:", e)

        # drop old
        c.execute("DROP TABLE quiz_scores_old")

        conn.commit()

    conn.close()


def has_taken_quiz(student_name, instructor, section, subject, quiz_number):
    try:
        ensure_student_quiz_db()
        conn = sqlite3.connect(os.path.join(os.path.abspath(os.path.dirname(__file__)), 'student_quiz_scores.db'))
        c = conn.cursor()
        c.execute("""
                  SELECT 1
                  FROM quiz_scores
                  WHERE student_name = ?
                    AND instructor = ?
                    AND section =?
                    AND subject=?
                    AND quiz_number=?
                      LIMIT 1
                  """, (student_name, instructor, section, subject, quiz_number))
        result = c.fetchone()
        conn.close()
        return result is not None
    except Exception as e:
        print(f"⚠️ DB Check Error: {e}")
        return False


def download_quiz(quiz_number):
    try:
        with open(f'quizzes/quiz_{quiz_number}.json', 'r') as file:
            quiz_info = json.load(file)
        with open(quiz_info['file_path'], 'r') as file:
            quiz_content = file.read()
        time_limit = quiz_info.get('time_limit', "No time limit specified")
        return quiz_content, time_limit
    except FileNotFoundError:
        messagebox.showerror("No Quiz Found", f"No quiz found for quiz number {quiz_number}.")
        return None, None


import re  # ✅ Required for this function to work


def extract_quiz_number(text):
    match = re.search(r'Quiz (\d+)', text)
    return int(match.group(1)) if match else None

def server_has_taken_quiz(student_name, instructor, section, subject, quiz_number):
    try:
        check = {
            "action": "check_quiz_taken",
            "student_name": student_name,
            "instructor": instructor,
            "section": section,
            "subject": subject,
            "quiz_number": int(quiz_number)
        }
        client_socket.sendall(json.dumps(check).encode("utf-8"))
        resp = json.loads(client_socket.recv(4096).decode("utf-8"))
        return bool(resp.get("taken"))
    except Exception as e:
        print("⚠️ server_has_taken_quiz failed:", e)
        return False


def start_quiz():
    global client_socket, quiz_dropdown_data
    global current_instructor, current_section, current_subject

    selected = quiz_number_var.get()
    if not selected:
        messagebox.showerror("Error", "Please select a quiz.")
        return

    quiz_meta = quiz_dropdown_data.get(selected)
    if not quiz_meta:
        messagebox.showerror("Error", "Quiz selection data not found. Please refresh.")
        return

    quiz_number = int(quiz_meta.get("quiz_number", 0) or 0)
    quiz_title = quiz_meta.get("quiz_title", "Untitled")

    if quiz_number <= 0:
        messagebox.showerror("Error", "Invalid quiz number.")
        return

    # ✅ Ask server if already taken (works across any PC)
    try:
        check = {
            "action": "check_quiz_taken",
            "student_name": student_name,
            "instructor": current_instructor,
            "section": current_section,
            "subject": current_subject,
            "quiz_number": quiz_number
        }
        client_socket.sendall(json.dumps(check).encode("utf-8"))
        resp = json.loads(client_socket.recv(2048).decode("utf-8"))

        if resp.get("taken"):
            messagebox.showinfo("Already Taken", "You have already completed this quiz.")
            return
    except Exception as e:
        print(f"❌ check_quiz_taken failed: {e}")
        messagebox.showerror("Error", "Failed to validate quiz status. Please try again.")
        return

    # ✅ Download quiz file
    try:
        request = {
            "action": "get_quiz_file",
            "quiz_number": quiz_number,
            "instructor": current_instructor,
            "section": current_section,
            "subject": current_subject
        }
        client_socket.sendall(json.dumps(request).encode("utf-8"))

        response = client_socket.recv(200000).decode("utf-8")
        quiz_data = json.loads(response)

        if quiz_data.get("status") != "ok":
            messagebox.showerror("Error", quiz_data.get("message", "Failed to load quiz."))
            return

        quiz_info = quiz_data.get("quiz_info", {})
        questions = quiz_info.get("questions", [])
        time_limit = int(quiz_info.get("time_limit", 0) or 0)
        quiz_title = quiz_info.get("quiz_title", quiz_title)

        if not questions:
            messagebox.showerror("Error", "No quiz questions found in this quiz.")
            return

        show_quiz_window(time_limit, quiz_number, quiz_title, questions)

    except Exception as e:
        print("❌ Failed to load quiz from server:", e)
        messagebox.showerror("Error", f"Failed to load quiz: {e}")


from datetime import datetime


def show_quiz_scores_page():
    global student_name, current_section, current_subject

    for widget in window.winfo_children():
        widget.destroy()

    # Background Setup
    script_dir = os.path.dirname(os.path.abspath(__file__))
    bg_path = os.path.join(script_dir, "bg.jpg")
    original_bg_image = Image.open(bg_path)

    def resize_bg(event=None):
        new_width = window.winfo_width()
        new_height = window.winfo_height()
        resized = original_bg_image.resize((new_width, new_height), Image.LANCZOS)
        bg = ImageTk.PhotoImage(resized)
        background_label.config(image=bg)
        background_label.image = bg

    background_label = tk.Label(window, bg="white")
    background_label.place(x=0, y=0, relwidth=1, relheight=1)
    resize_bg()
    window.bind("<Configure>", resize_bg)

    # Styles
    style = ttk.Style()
    style.configure("Treeview", font=("Helvetica", 12), rowheight=28)
    style.configure("Treeview.Heading", font=("Helvetica", 12, "bold"))
    style.configure("Curved.TButton", font=("Helvetica", 11))

    # Title
    ttk.Label(window, text=f"{student_name}'s Quiz Scores", font=("Helvetica", 18, "bold"),
              background="white", foreground="#5D3FD3").pack(pady=20)

    # Container Frame
    container = tk.Frame(window, bg="white")
    container.pack(pady=(0, 10), padx=20, fill=tk.BOTH, expand=True)

    # Treeview Frame
    tree_frame = tk.Frame(container, bg="white", bd=1, relief="solid")
    tree_frame.pack(fill=tk.BOTH, expand=True)

    columns = ("subject", "quiz", "score", "date")
    tree = ttk.Treeview(tree_frame, columns=columns, show="headings")
    sort_order = {col: True for col in columns}
    quiz_data = []

    def sort_column(col):
        reverse = sort_order[col]
        sort_order[col] = not reverse
        index = columns.index(col)
        sorted_rows = sorted(quiz_data, key=lambda x: x[index], reverse=reverse)
        tree.delete(*tree.get_children())
        for row in sorted_rows:
            tree.insert("", tk.END, values=row)

    for col in columns:
        tree.heading(col, text=col.capitalize(), command=lambda c=col: sort_column(c))
        tree.column(col, anchor="center", width=150)

    tree.column("quiz", width=80)
    tree.column("score", width=100)

    scrollbar = ttk.Scrollbar(tree_frame, orient="vertical", command=tree.yview)
    tree.configure(yscrollcommand=scrollbar.set)
    scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
    tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

    # Fetch and populate quiz scores
    try:
        conn = sqlite3.connect(os.path.join(os.path.abspath(os.path.dirname(__file__)), 'student_quiz_scores.db'))
        c = conn.cursor()
        c.execute("SELECT quiz_number, score, subject, date_taken FROM quiz_scores WHERE student_name = ?",
                  (student_name,))
        rows = c.fetchall()
        conn.close()

        if rows:
            for quiz_number, score, subject, date in rows:
                row = (subject, quiz_number, f"{score:.2f}%", date)
                quiz_data.append(row)
                tree.insert("", tk.END, values=row)
        else:
            tree.insert("", tk.END, values=("No scores found", "", "", ""))
    except Exception as e:
        print(f"⚠️ Error fetching scores: {e}")
        tree.insert("", tk.END, values=("Error retrieving scores", "", "", ""))

    button_frame = tk.Frame(window, bg="white")
    button_frame.pack(pady=(5, 20))

    def export_to_excel(student_name, section, subject):
        try:
            import subprocess
            from datetime import datetime
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment

            if not quiz_data:
                messagebox.showinfo("No Data", "No quiz scores to export.")
                return
            desktop = os.path.join(os.path.expanduser("~"), "Desktop")
            base_folder = os.path.join(desktop, "quiz_scores_students")
            os.makedirs(base_folder, exist_ok=True)
            student_folder = os.path.join(base_folder, student_name)
            os.makedirs(student_folder, exist_ok=True)

            now = datetime.now()
            date_str = now.strftime("%Y_%m_%d")

            filename = f"{student_name}_{section}_{subject}_{date_str}.xlsx"
            file_path = os.path.join(student_folder, filename)

            wb = Workbook()
            ws = wb.active
            ws.title = "Quiz Scores"

            ws.merge_cells("A1:D1")
            ws.merge_cells("A2:D2")
            ws.merge_cells("A3:D3")

            ws["A1"] = "QUIZ SCORES REPORT"
            ws["A2"] = f"Student: {student_name}"
            ws["A3"] = f"Section: {section} | Subject: {subject}"

            ws["A1"].font = Font(size=16, bold=True)
            ws["A2"].font = Font(size=12, bold=True)
            ws["A3"].font = Font(size=12)

            for r in range(1, 4):
                ws[f"A{r}"].alignment = Alignment(horizontal="center")

            headers = ["Subject", "Quiz #", "Score", "Date"]
            ws.append([])
            ws.append(headers)

            for cell in ws[5]:
                cell.font = Font(bold=True)

            for row in quiz_data:
                ws.append(row)

            wb.save(file_path)

            try:
                if os.name == "nt":
                    subprocess.Popen(f'explorer "{student_folder}"')
                else:
                    subprocess.Popen(["open", student_folder])
            except:
                pass

            messagebox.showinfo(
                "Export Successful",
                f"Quiz scores exported successfully!\n\n"
                f"File:\n{filename}"
            )

        except Exception as e:
            messagebox.showerror("Export Failed", str(e))

    ttk.Button(
        button_frame,
        text="📄 Export to Excel",
        style="Curved.TButton",
        command=lambda: export_to_excel(student_name, current_section, current_subject)
    ).grid(row=0, column=0, padx=10)

    ttk.Button(button_frame, text="🔙 Back", style="Curved.TButton",
               command=lambda: show_student_landing_page(student_name, current_section, current_subject)).grid(row=0,
                                                                                                               column=1,
                                                                                                               padx=10)


def show_quiz_page(quiz_number, quiz_title, time_limit, questions):
    for widget in window.winfo_children():
        widget.destroy()

    ttk.Label(window, text=f"Quiz {quiz_number} - {quiz_title}", font=("Helvetica", 16)).pack(pady=10)

    if not questions:
        messagebox.showerror("Error", "No questions found for this quiz!")
        return

    quiz_frame = tk.Frame(window)
    quiz_frame.pack(pady=10, padx=20, fill="both", expand=True)

    canvas = tk.Canvas(quiz_frame)
    scrollbar = ttk.Scrollbar(quiz_frame, orient="vertical", command=canvas.yview)
    scrollable_frame = ttk.Frame(canvas)

    scrollable_frame.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
    canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
    canvas.configure(yscrollcommand=scrollbar.set)

    answer_vars = []

    for i, q in enumerate(questions):
        ttk.Label(scrollable_frame, text=q.get('question', ''), font=("Helvetica", 14)).pack(anchor="w", pady=5)

        qtype = str(q.get("type", "MCQ")).upper()

        if qtype == "MCQ":
            ans_var = tk.StringVar()
            answer_vars.append({"type": "MCQ", "var": ans_var})

            for choice in q.get("choices", []):
                ttk.Radiobutton(scrollable_frame, text=choice, variable=ans_var, value=choice).pack(anchor="w")

        else:  # ID
            ans_var = tk.StringVar()
            answer_vars.append({"type": "ID", "var": ans_var})
            ttk.Entry(scrollable_frame, textvariable=ans_var, width=45).pack(anchor="w", pady=4)
            ttk.Label(scrollable_frame, text="(Type your answer)", foreground="gray").pack(anchor="w", pady=(0, 10))

    canvas.pack(side="left", fill="both", expand=True)
    scrollbar.pack(side="right", fill="y")

    # Timer label
    timer_label = ttk.Label(window, font=("Helvetica", 14))
    timer_label.pack(pady=10)

    if time_limit and time_limit > 0:
        time_remaining = time_limit * 60

        def update_timer():
            nonlocal time_remaining
            if time_remaining <= 0:
                submit_quiz(quiz_number, questions, answer_vars)
                return

            mins = time_remaining // 60
            secs = time_remaining % 60
            timer_label.config(text=f"Time Left: {mins:02d}:{secs:02d}")
            time_remaining -= 1
            window.after(1000, update_timer)

        update_timer()
    else:
        timer_label.config(text="No Timer")

    submit_button = ttk.Button(
        window, text="Submit",
        command=lambda: submit_quiz(quiz_number, questions, answer_vars)
    )
    submit_button.pack(pady=10)


def bind_mousewheel_to_canvas_for_window(win, canvas):
    """
    Mousewheel scroll for a Canvas INSIDE a specific Toplevel only.
    Works on Windows/macOS (MouseWheel) + Linux (Button-4/5).
    """

    def _y_scroll(delta_units):
        canvas.yview_scroll(delta_units, "units")
        return "break"

    # Windows / macOS
    def _on_mousewheel(event):
        # On Windows event.delta is usually +/-120 multiples
        # On macOS it's small values
        if event.delta == 0:
            return "break"
        direction = -1 if event.delta > 0 else 1
        return _y_scroll(direction * 3)

    # Linux
    def _on_linux_up(event):
        return _y_scroll(-3)

    def _on_linux_down(event):
        return _y_scroll(3)

    # Bind when mouse enters the canvas (or any child), unbind when leaves
    def _bind(_=None):
        win.bind("<MouseWheel>", _on_mousewheel)
        win.bind("<Button-4>", _on_linux_up)
        win.bind("<Button-5>", _on_linux_down)

    def _unbind(_=None):
        win.unbind("<MouseWheel>")
        win.unbind("<Button-4>")
        win.unbind("<Button-5>")

    # Bind on enter/leave of canvas area
    canvas.bind("<Enter>", _bind)
    canvas.bind("<Leave>", _unbind)

    # Also: bind to inner frame so wheel works when hovering over labels
    # (you’ll pass `inner` and bind Enter/Leave there too if you want)
    return _bind, _unbind


def show_quiz_window(time_limit, quiz_number, quiz_title, questions):
    # ── Zebra stripe colours ──────────────────────────────────────────────
    STRIPE_ODD  = "#FFFFFF"        # white  (odd  questions, 1-based)
    STRIPE_EVEN = "#EEF4FF"        # light blue tint (even questions)
    DARK_BLUE   = "#003580"        # header / footer / submit button

    def enable_mousewheel(widget):
        def _on_mousewheel(event):
            widget.yview_scroll(int(-1 * (event.delta / 120)), "units")
            return "break"
        widget.bind("<Enter>", lambda e: widget.bind_all("<MouseWheel>", _on_mousewheel))
        widget.bind("<Leave>", lambda e: widget.unbind_all("<MouseWheel>"))

    quiz_window = tk.Toplevel(window)
    quiz_window.title(f"Quiz #{quiz_number} - {quiz_title}")
    quiz_window.geometry("1000x700")
    quiz_window.configure(bg="white")

    # ── Dark-blue header bar ──────────────────────────────────────────────
    header = tk.Frame(quiz_window, bg=DARK_BLUE, height=60)
    header.pack(fill="x")

    tk.Label(
        header,
        text=f"Quiz #{quiz_number} - {quiz_title}",
        font=("Helvetica", 20, "bold"),
        bg=DARK_BLUE, fg="white"
    ).pack(side="left", padx=20)

    timer_label = tk.Label(
        header,
        text="Time Left",
        font=("Helvetica", 16, "bold"),
        bg=DARK_BLUE, fg="white"
    )
    timer_label.pack(side="right", padx=20)

    # ── Scrollable body (white) ───────────────────────────────────────────
    body = tk.Frame(quiz_window, bg="white")
    body.pack(fill="both", expand=True)

    canvas = tk.Canvas(body, bg="white", highlightthickness=0)
    scrollbar = ttk.Scrollbar(body, orient="vertical", command=canvas.yview)
    canvas.pack(side="left", fill="both", expand=True)
    scrollbar.pack(side="right", fill="y")
    canvas.configure(yscrollcommand=scrollbar.set)

    scroll_frame = tk.Frame(canvas, bg="white")
    window_id = canvas.create_window((0, 0), window=scroll_frame, anchor="n")

    def resize_frame(event):
        canvas.itemconfig(window_id, width=event.width)

    canvas.bind("<Configure>", resize_frame)
    scroll_frame.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))

    answer_vars = []

    # ── Question cards with zebra striping ───────────────────────────────
    for i, q in enumerate(questions, start=1):
        card_bg = STRIPE_ODD if i % 2 != 0 else STRIPE_EVEN
        card = tk.Frame(scroll_frame, bg=card_bg, bd=1, relief="ridge")
        card.pack(fill="x", padx=60, pady=8)

        # Question number + text in black on card background
        tk.Label(
            card,
            text=f"{i}. {q.get('question', '')}",
            font=("Helvetica", 14, "bold"),
            bg=card_bg, fg="black",
            wraplength=800, justify="left"
        ).pack(anchor="w", padx=16, pady=(12, 6))

        qtype = str(q.get("type", "MCQ")).upper()

        if qtype == "MCQ":
            ans_var = tk.StringVar()
            answer_vars.append({"type": "MCQ", "var": ans_var})
            for choice in q.get("choices", []):
                tk.Radiobutton(
                    card,
                    text=choice,
                    variable=ans_var,
                    value=choice,
                    bg=card_bg, fg="black",
                    activebackground=card_bg,
                    font=("Helvetica", 12),
                    anchor="w"
                ).pack(anchor="w", padx=32, pady=3)
        else:  # ID / short-answer
            ans_var = tk.StringVar()
            answer_vars.append({"type": "ID", "var": ans_var})
            tk.Entry(
                card,
                textvariable=ans_var,
                width=50,
                font=("Helvetica", 12),
                bg="white", fg="black",
                insertbackground="black"
            ).pack(anchor="w", padx=16, pady=8)
            tk.Label(card, text="(Type your answer)", bg=card_bg, fg="gray",
                     font=("Helvetica", 10, "italic")).pack(anchor="w", padx=16, pady=(0, 10))

    # ── Dark-blue footer with styled Submit button ────────────────────────
    footer = tk.Frame(quiz_window, bg=DARK_BLUE)
    footer.pack(fill="x")

    tk.Button(
        footer,
        text="✅  Submit Quiz",
        font=("Helvetica", 14, "bold"),
        bg=DARK_BLUE,
        fg="white",
        activebackground="#002060",
        activeforeground="white",
        relief="flat",
        cursor="hand2",
        padx=30, pady=10,
        command=lambda: confirm_submit(quiz_window, quiz_number, questions, answer_vars)
    ).pack(pady=12)

    # ── Mouse-wheel scrolling ─────────────────────────────────────────────
    def bind_mousewheel_to_canvas(win, canvas):
        def _on_mousewheel(event):
            canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")
            return "break"
        def _on_shift_mousewheel(event):
            canvas.xview_scroll(int(-1 * (event.delta / 120)), "units")
            return "break"
        win.bind_all("<MouseWheel>", _on_mousewheel)
        win.bind_all("<Shift-MouseWheel>", _on_shift_mousewheel)
        win.bind_all("<Button-4>", lambda e: canvas.yview_scroll(-1, "units"))
        win.bind_all("<Button-5>", lambda e: canvas.yview_scroll(1, "units"))

    bind_mousewheel_to_canvas(quiz_window, canvas)

    # ── Timer countdown ───────────────────────────────────────────────────
    if time_limit and time_limit > 0:
        time_remaining = time_limit * 60

        def update_timer():
            nonlocal time_remaining
            if not quiz_window.winfo_exists():
                return
            if time_remaining <= 0:
                submit_quiz(quiz_number, questions, answer_vars)
                quiz_window.destroy()
                return
            mins = time_remaining // 60
            secs = time_remaining % 60
            timer_label.config(text=f"Time Left: {mins:02d}:{secs:02d}")
            time_remaining -= 1
            quiz_window.after(1000, update_timer)

        update_timer()
    else:
        timer_label.config(text="No Timer")

    def cleanup_bindings():
        try:
            quiz_window.unbind_all("<MouseWheel>")
            quiz_window.unbind_all("<Shift-MouseWheel>")
            quiz_window.unbind_all("<Button-4>")
            quiz_window.unbind_all("<Button-5>")
        except:
            pass

    quiz_window.protocol("WM_DELETE_WINDOW", lambda: (cleanup_bindings(), quiz_window.destroy()))


def confirm_submit(parent, quiz_number, questions, answer_vars):
    review = tk.Toplevel(parent)
    review.title("Review Answers")
    review.geometry("700x500")

    frame = tk.Frame(review)
    frame.pack(fill="both", expand=True, padx=10, pady=10)

    canvas = tk.Canvas(frame)
    scrollbar = ttk.Scrollbar(frame, orient="vertical", command=canvas.yview)
    canvas.pack(side="left", fill="both", expand=True)
    scrollbar.pack(side="right", fill="y")
    canvas.configure(yscrollcommand=scrollbar.set)

    inner = tk.Frame(canvas)
    canvas.create_window((0, 0), window=inner, anchor="nw")
    inner.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))

    # ✅ enable mousewheel scrolling (review window only)
    bind_mousewheel_to_canvas_for_window(review, canvas)

    # OPTIONAL: also bind when hovering labels (so it never “stops”)
    inner.bind("<Enter>", lambda e: bind_mousewheel_to_canvas_for_window(review, canvas)[0]())
    inner.bind("<Leave>", lambda e: bind_mousewheel_to_canvas_for_window(review, canvas)[1]())

    for i, q in enumerate(questions):
        ans = answer_vars[i]["var"].get().strip() or "❌ No answer selected"

        tk.Label(inner, text=f"{i + 1}. {q.get('question', '')}",
                 font=("Helvetica", 12, "bold"),
                 wraplength=600, justify="left").pack(anchor="w", pady=(8, 2))

        tk.Label(inner, text=f"➡ Your answer: {ans}",
                 font=("Helvetica", 11), fg="blue").pack(anchor="w")

    btn_frame = tk.Frame(review)
    btn_frame.pack(pady=10)

    ttk.Button(
        btn_frame, text="✅ Confirm & Submit",
        command=lambda: (
            submit_quiz(quiz_number, questions, answer_vars),
            review.destroy(),
            parent.destroy()
        )
    ).pack(side="left", padx=10)

    ttk.Button(btn_frame, text="⬅ Go Back", command=review.destroy).pack(side="left", padx=10)



def submit_quiz(quiz_number, questions, answer_vars):
    global student_name, current_section, current_subject, current_instructor, client_socket

    correct = 0
    total = len(questions)

    def norm(s: str) -> str:
        return str(s).strip().lower()

    for i, q in enumerate(questions):
        qtype = str(q.get("type", "MCQ")).upper()
        selected = answer_vars[i]["var"].get().strip()

        if qtype == "MCQ":
            # selected looks like: "A. blah" or "A. ..."
            selected_letter = ""
            if selected:
                selected_letter = selected.split(".", 1)[0].strip().upper()

            correct_letter = str(q.get("answer", "")).strip().upper()
            if selected_letter and selected_letter == correct_letter:
                correct += 1

        else:  # ID
            # compare text (case-insensitive)
            correct_text = norm(q.get("answer", ""))
            if selected and norm(selected) == correct_text:
                correct += 1

    if total == 0:
        messagebox.showerror("Error", "No questions found.")
        return

    score = (correct / total) * 100.0
    date_taken = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    # ✅ send to server (include instructor too, so DB can filter later if you upgrade)
    payload = {
        "action": "submit_quiz_score",
        "student_name": student_name,
        "quiz_number": quiz_number,
        "score": score,
        "section": current_section,
        "subject": current_subject,
        "date_taken": date_taken,
        "instructor": current_instructor
    }

    try:
        client_socket.sendall(json.dumps(payload).encode("utf-8"))
        print("✅ Quiz score sent to server.")
    except Exception as e:
        print(f"❌ Failed to send score to server: {e}")
        messagebox.showerror("Error", f"Failed to send score to instructor.\n\n{e}")
        return

    # ✅ save locally (unique)
    try:
        ensure_student_quiz_db()
        conn = sqlite3.connect(os.path.join(os.path.abspath(os.path.dirname(__file__)), 'student_quiz_scores.db'))
        c = conn.cursor()
        c.execute("""
                  INSERT INTO quiz_scores (student_name, instructor, section, subject, quiz_number, score, date_taken)
                  VALUES (?, ?, ?, ?, ?, ?, ?)
                  """,
                  (student_name, current_instructor, current_section, current_subject, quiz_number, score, date_taken))
        conn.commit()
        conn.close()
        print("✅ Score saved locally.")
    except Exception as e:
        print(f"⚠️ Failed to save locally: {e}")

    messagebox.showinfo(
        "Quiz Completed",
        f"You answered {correct} out of {total} correctly.\nYour score: {score:.2f}%"
    )
    show_quiz_scores_page()


def get_available_quizzes():
    global client_socket, current_instructor, current_section, current_subject

    try:
        request = {
            "action": "get_quizzes",
            "instructor": current_instructor,
            "section": current_section,
            "subject": current_subject
        }
        client_socket.sendall(json.dumps(request).encode("utf-8"))

        response = client_socket.recv(8192).decode("utf-8")
        data = json.loads(response)
        return data.get("quizzes", [])
    except Exception as e:
        print(f"❌ Error fetching quizzes from instructor: {e}")
        return []


quiz_dropdown_data = {}  # <-- put this GLOBAL near the top


# Make sure this exists ONCE globally (outside the function)
# quiz_dropdown_data = {}

def show_quiz_selection_page():
    global quiz_dropdown_data, quiz_dropdown, quiz_number_var
    global current_instructor, current_section, current_subject
    global student_name

    for widget in window.winfo_children():
        widget.destroy()

    # ========= BACKGROUND =========
    script_dir = os.path.dirname(os.path.abspath(__file__))
    bg_path = os.path.join(script_dir, "bg.jpg")
    original_bg_image = Image.open(bg_path)

    def resize_bg(event=None):
        new_width = window.winfo_width()
        new_height = window.winfo_height()
        resized_bg = original_bg_image.resize((new_width, new_height), Image.LANCZOS)
        bg_image = ImageTk.PhotoImage(resized_bg)
        background_label.config(image=bg_image)
        background_label.image = bg_image

    background_label = tk.Label(window, bg="white")
    background_label.place(x=0, y=0, relwidth=1, relheight=1)
    resize_bg()
    window.bind("<Configure>", resize_bg)

    # ========= MAIN FRAME =========
    quiz_frame = ttk.Frame(window, style="CustomFrame.TFrame")
    quiz_frame.place(relx=0.5, rely=0.5, anchor="center", width=700, height=450)

    ttk.Label(
        quiz_frame,
        text="Select a quiz to start:",
        font=("Helvetica", 14),
        style="Label.TLabel"
    ).pack(pady=10)

    quiz_number_var = tk.StringVar()

    quiz_dropdown = ttk.Combobox(
        quiz_frame,
        font=("Helvetica", 12),
        textvariable=quiz_number_var,
        state="readonly",
        width=55
    )
    quiz_dropdown.pack(pady=10)

    # ========= REFRESH FUNCTION (single source of truth) =========
    def refresh_quiz_dropdown():
        available_quizzes = get_available_quizzes()

        formatted = []
        quiz_dropdown_data.clear()

        for quiz in available_quizzes:
            num = quiz.get("quiz_number")
            subject = quiz.get("subject", "Unknown")
            status = quiz.get("status", "Closed")
            timestamp = quiz.get("timestamp", "Unknown Date")

            title = quiz.get("quiz_title", "Untitled")
            time_limit = int(quiz.get("time_limit", 0) or 0)
            tl_txt = f"{time_limit}m" if time_limit > 0 else "No Timer"

            display = f"[{status}] {subject} - Quiz {num}: {title} ({tl_txt}) (Uploaded: {timestamp})"

            if server_has_taken_quiz(student_name, current_instructor, current_section, current_subject, num):
                display += " ✅ Done"

            formatted.append(display)
            quiz_dropdown_data[display] = quiz

        quiz_dropdown["values"] = formatted
        quiz_number_var.set(formatted[0] if formatted else "")

    # ========= BUTTONS =========
    button_style = {"style": "Orange.TButton", "width": 25}

    ttk.Button(
        quiz_frame,
        text="Start Quiz",
        **button_style,
        command=start_quiz
    ).pack(pady=10)

    ttk.Button(
        quiz_frame,
        text="View My Quiz Scores",
        **button_style,
        command=show_quiz_scores_page
    ).pack(pady=10)

    ttk.Button(
        quiz_frame,
        text="🔄 Refresh Quiz List",
        style="Orange.TButton",
        width=25,
        command=refresh_quiz_dropdown
    ).pack(pady=5)

    ttk.Button(
        quiz_frame,
        text="Back",
        style="Orange.TButton",
        width=15,
        command=lambda: show_student_landing_page(student_name, current_section, current_subject)
    ).pack(pady=10)

    refresh_quiz_dropdown()


def show_student_landing_page(name, section, subject):
    global style

    style.configure("Custom.TLabel", foreground="#8A2BE2", background="white", font=("Helvetica", 16))
    style.configure("CustomFrame.TFrame", background="white")

    # ✅ Configure Bootstrap Orange Theme for Buttons
    style.configure("Orange.TButton",
                    font=("Helvetica", 14, "bold"),
                    padding=10,
                    relief="flat",
                    background="#FF8C00",  # Dark Orange
                    foreground="white",
                    borderwidth=2,
                    bordercolor="white",
                    borderradius=15)

    style.map("Orange.TButton",
              foreground=[('hover', 'white')],
              background=[('hover', '#3BB143')],  # Slightly darker orange on hover
              bordercolor=[('hover', '#3BB143')])

    for widget in window.winfo_children():
        widget.destroy()

    # Load original background only once
    script_dir = os.path.dirname(os.path.abspath(__file__))
    bg_path = os.path.join(script_dir, "bg.jpg")
    original_bg_image = Image.open(bg_path)

    def resize_bg(event=None):
        new_width = window.winfo_width()
        new_height = window.winfo_height()

        resized_bg = original_bg_image.resize((new_width, new_height), Image.LANCZOS)
        bg_image = ImageTk.PhotoImage(resized_bg)

        background_label.config(image=bg_image)
        background_label.image = bg_image  # keep reference to prevent garbage collection

    background_label = tk.Label(window, bg="white")
    background_label.place(x=0, y=0, relwidth=1, relheight=1)

    resize_bg()

    window.bind("<Configure>", resize_bg)

    main_frame = ttk.Frame(window, style="CustomFrame.TFrame", width=600, height=500)
    main_frame.pack(pady=20, padx=40, fill=tk.BOTH, expand=True)
    main_frame.pack_propagate(False)

    welcome_label = ttk.Label(main_frame, text=f"Welcome, {name}!", style="Custom.TLabel")
    welcome_label.pack(pady=20)

    prompt_label = ttk.Label(main_frame, text="What would you like to do?", style="Custom.TLabel")
    prompt_label.pack(pady=20)

    button_width = 25
    quiz_button = ttk.Button(main_frame, text="Start the Quiz", style="Orange.TButton",
                             width=button_width, command=show_quiz_selection_page)
    quiz_button.pack(pady=10)

    activity_button = ttk.Button(main_frame, text="Start the Activity", style="Orange.TButton",
                                 width=button_width, command=start_activity)
    activity_button.pack(pady=10)



    logs_button = ttk.Button(main_frame, text="View Attendance Logs", style="Orange.TButton",
                             width=button_width, command=show_attendance_logs)
    logs_button.pack(pady=10)

    def logout():
        if messagebox.askyesno("Logout Confirmation", "Are you sure you want to logout?"):
            show_sign_in_page()

    logout_button = ttk.Button(main_frame, text="Logout", style="Orange.TButton",
                               width=button_width, command=logout)
    logout_button.pack(pady=10)

    student_info_label = ttk.Label(
        main_frame,
        text=f"Student: {name} | Section: {section} | Subject: {subject}",
        style="Custom.TLabel"
    )
    student_info_label.pack(side='bottom', pady=20)

    motto_label = tk.Label(window, text="Truth. Excellence. Service.\nCavite State University Tanza Campus",
                           fg='white', font=("Arial", 12), bg='#6A329D')
    motto_label.pack(pady=10, side=tk.BOTTOM)


def listen_for_commands():
    global client_socket
    while True:
        try:
            ready_to_read, _, _ = select.select([client_socket], [], [], 0.5)
            if ready_to_read:
                command = client_socket.recv(1024).decode('utf-8')
                if not command:
                    break

                print("📩 Received:", command)
                if command.strip().startswith("{"):
                    data = json.loads(command)
                    action = data.get("action", "").lower()
                    timer = data.get("timer", 0)

                    if action == "welcome_pc":
                        global client_pc_id
                        client_pc_id = data.get("client_id")
                        label = data.get("label", "Unknown")
                        update_pc_label(label)
                    elif action in ["shutdown", "sleep"]:
                        if timer > 0:
                            show_shutdown_timer(timer, action.capitalize())
                        else:
                            shutdown_system() if action == "shutdown" else sleep_system()

                    elif action == "launch_program":
                        program = data.get("program", "")
                        print(f"🖥 Launching requested program: {program}")
                        launch_program(program)

                    else:
                        print("⚠️ Unknown action received.")
                else:
                    print("⚠️ Received non-JSON data:", command)
        except Exception as e:
            print(f"❌ Error in listener: {e}")
            break


import os
import subprocess
import shutil


def launch_program(program):
    try:
        program = program.upper()

        # Try to detect the executable using known names and PATH
        common_executables = {
            "CHROME": "chrome.exe",
            "ECLIPSE": "eclipse.exe",
            "XAMPP": "xampp-control.exe",
            "VISUAL STUDIO": "devenv.exe",
            "MYSQL": "MySQLWorkbench.exe",
            "PYTHONIDE": "python.exe",
            "NETBEANS": "netbeans64.exe",
            "NODEJS": "node.exe",
            "PENCILDRAW": "Pencil2D.exe",
            "C++": "codeblocks.exe"  # Or change to Dev-C++ if needed
        }

        exe_name = common_executables.get(program)

        if not exe_name:
            print(f"⚠️ Program '{program}' not recognized.")
            return

        # Try finding it in system PATH first
        path = shutil.which(exe_name)

        # If not found, search common locations
        if not path:
            search_dirs = [
                os.environ.get("ProgramFiles", r"C:\Program Files"),
                os.environ.get("ProgramFiles(x86)", r"C:\Program Files (x86)"),
                r"C:\xampp",  # Specific known paths
            ]

            for root_dir in search_dirs:
                for root, dirs, files in os.walk(root_dir):
                    if exe_name in files:
                        path = os.path.join(root, exe_name)
                        break
                if path:
                    break

        if path:
            subprocess.Popen(path)
            print(f"🚀 Launched {program} at: {path}")
        else:
            print(f"❌ {program} not found on this system.")

    except Exception as e:
        print(f"❌ Failed to launch {program}: {e}")

if __name__ == "__main__":
    create_persistent_label()
    show_landing_page()
    window.mainloop()